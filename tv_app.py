#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re, io, warnings, requests, time
from datetime import datetime

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────
# 應用版本資訊
# ─────────────────────────────────────────────────────────────────
APP_VERSION   = "v9.28"
APP_UPDATED   = "2026-05-14 11:00"
APP_NOTES     = (
    "🆕 detail card 加 SEPA / VCP / RS 詳細診斷 section（8 條件逐項打勾）"
    "  ── 動態進出場建議：完整 setup → 強烈進場；跌破 SMA50/200 → 出場 ｜ "
    "🆕 綜合決策得分系統（SEPA+VCP+T1+B+recipes 加總，-2~+5 分）｜ "
    "🆕 SEPA 投組回測 OOS：🇹🇼 fifo+pos10 CAGR +60.6% Sharpe 1.79 MDD -2.76% ⭐⭐ "
    "  🇺🇸 rs_high+pos50 CAGR +20% Sharpe 1.52 MDD -7.08% ｜ "
    "🆕 SEPA 候選 LINE 通知（Minervini 完整 setup / SEPA+RS70 / Pivot breakout）｜ "
    "—— 上版 v9.19.1 ——｜ "
    "🆕 SEPA OOS 驗證（TW SEPA+VCP+RS70+90d: 47% win, +11.2% mean）"
)
APP_VALIDATIONS = (
    "🆕 BB 全套（OANDA 10 種判斷）alpha 驗證:"
    "  🐻空頭+BB Expansion win 56.1%/+3.14%, 🐻空頭+%B<0 win 55.5%/+2.84%, 🐂多頭+%B<0 win 54.9%/+3.32% ｜ "
    "🚀 倒鎚 OOS 2024+ 71.8% 勝率 / +9.35% 30d / PF 5.5（無過擬合）｜ "
    "🎯 T1_V7 OOS CAGR：🇹🇼 h30 +14.78% / 🇺🇸 h60 +16.40%（跨市場最佳 hold 不同）｜ "
    "⚠️ T1_V7 hold=60 在 TW 是過擬合 trap — TW 勿用 60d ｜ "
    "💎 max_pos=50 + drop_deep priority → 倒鎚 Sharpe 0.54→1.99 / MDD -3.44% ｜ "
    "🛑 止損研究：倒鎚不該止損 / T1_V7 fixed_10 略好 / Trailing 全失敗 ｜ "
    "📅 月份效應：🇹🇼 3月-7.94%/4月+15.91% ｜ 🇺🇸 5月-3.17%/10月+5.77% ｜ "
    "🎲 混合策略沒奇跡（相關性高，選一個純策略最簡單清楚）｜ "
    "📊 大盤對照：🇹🇼 TWII +20.35% / 🇺🇸 SPY +13.31%（CAGR 輸但 Sharpe + MDD 勝）｜ "
    "🪙 v8 不適用加密貨幣"
)

import numpy as np
import pandas as pd
import ta
import yfinance as yf

# 設定 yfinance 請求 headers，避免被 Yahoo Finance 封鎖
try:
    import requests as _req
    _session = _req.Session()
    _session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
    })
    yf.set_config(session=_session)
except Exception:
    pass

# twstock loaded on demand via cached function
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Indicator Scanner", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');
html,body,[class*="css"]{font-family:'IBM Plex Sans',sans-serif;}
.tv-header{background:linear-gradient(135deg,#0a0f1e 0%,#0d2137 60%,#0a1628 100%);border-bottom:1px solid #1e3a5f;padding:20px 28px 16px;margin:-1rem -1rem 1.5rem -1rem;display:flex;align-items:center;gap:14px;}
.tv-header h1{font-family:'IBM Plex Mono',monospace;font-size:1.3rem;font-weight:600;color:#e8f4fd;margin:0;letter-spacing:.04em;}
.tv-header .sub{font-size:.75rem;color:#5a8ab0;margin-top:2px;}
.cards-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:1.4rem;}
.card{flex:1;min-width:160px;background:#0d1b2e;border:1px solid #1e3a5f;border-radius:10px;padding:12px 14px;text-align:center;display:flex;flex-direction:column;align-items:center;}
.card .c-label{font-size:.7rem;letter-spacing:.08em;text-transform:uppercase;}
.card .c-value{font-size:1.6rem;font-weight:700;margin-top:2px;}
.card.total{background:#0d1b2e;}
.card.total .c-label{color:#5a8ab0;} .card.total .c-value{color:#f0f4ff;}
/* 🆕 進場：綠 */
.card.entry{background:#0a1e10;border-color:#3dbb6a55;}
.card.entry .c-label{color:#3dbb6a;} .card.entry .c-value{color:#3dbb6a;}
/* 🆕 出場：紅 */
.card.exit{background:#1a0808;border-color:#ff555555;}
.card.exit .c-label{color:#ff7777;} .card.exit .c-value{color:#ff7777;}
/* 🆕 持倉：藍 */
.card.hold{background:#0a1830;border-color:#5a8ab055;}
.card.hold .c-label{color:#7abadd;} .card.hold .c-value{color:#7abadd;}
/* 觀望：灰 */
.card.neu{background:#0d1b2e;}
.card.neu .c-label{color:#7a8899;} .card.neu .c-value{color:#7a8899;}
/* 舊 buy / sell 對齊（avoid breaking other usages）*/
.card.buy .c-value{color:#3dbb6a;}.card.sell .c-value{color:#ff7777;}
.res-table{width:100%;border-collapse:collapse;font-size:.82rem;}
.res-table th{background:#0a1628;color:#5a8ab0;font-size:.68rem;font-weight:600;text-transform:uppercase;letter-spacing:.07em;padding:8px 10px;text-align:center;border-bottom:2px solid #1e3a5f;white-space:nowrap;}
.res-table td{padding:7px 10px;text-align:center;border-bottom:1px solid #0f1f33;white-space:nowrap;font-family:'IBM Plex Mono',monospace;font-size:.78rem;}
.res-table tr:hover td{background:rgba(30,58,95,.35);}
.ticker-cell{font-weight:700;color:#e8f4fd !important;font-size:.9rem !important;}
.market-cell{color:#7ab0d0 !important;font-size:.72rem !important;}
.j-buy{color:#3b9eff;font-weight:600;}.j-sell{color:#ff5555;font-weight:600;}.j-neutral{color:#8899aa;}.j-na{color:#4a6070;font-style:italic;}
.badge{display:inline-block;padding:2px 9px;border-radius:20px;font-size:.72rem;font-weight:700;font-family:'IBM Plex Sans',sans-serif;letter-spacing:.03em;}
.badge-strong-buy{background:#0d3b6e;color:#3b9eff;border:1px solid #1a5fa8;}
.badge-buy{background:#0d2e50;color:#60b3ff;border:1px solid #154d84;}
.badge-strong-sell{background:#4a0a0a;color:#ff6b6b;border:1px solid #8b1a1a;}
.badge-sell{background:#3b0d0d;color:#ff8080;border:1px solid #6b1515;}
.badge-neutral{background:#1a2030;color:#8899aa;border:1px solid #2a3545;}
.badge-buy-limit{background:#3a2a00;color:#f0c030;border:1px solid #8a6000;}
.badge-overheat{background:#3a1800;color:#ff8830;border:1px solid #8a3800;}
.badge-bearish{background:#3a0808;color:#ff5555;border:1px solid #7a1010;}
.ind-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:8px;margin:10px 0;}
.ind-item{background:#0d1b2e;border:1px solid #1e3550;border-radius:8px;padding:10px 13px;display:flex;flex-direction:column;gap:5px;}
.ind-label{color:#7aaac8;font-size:.72rem;font-weight:600;letter-spacing:.03em;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.ind-val{font-family:'IBM Plex Sans',sans-serif;font-size:.82rem;font-weight:600;word-break:break-word;line-height:1.35;}
.ind-buy .ind-val{color:#3b9eff;}.ind-sell .ind-val{color:#ff5555;}.ind-neu .ind-val{color:#9aaabb;}
.section-title{font-size:.65rem;color:#3a5a7a;text-transform:uppercase;letter-spacing:.1em;font-weight:700;padding:6px 0 4px;margin-top:8px;border-top:1px solid #0f1f33;}
section[data-testid="stSidebar"]{background:#080e1a;border-right:1px solid #1e3a5f;}
section[data-testid="stSidebar"] .stTextArea textarea{background:#0d1b2e !important;color:#c8dff0 !important;border:1px solid #1e3a5f !important;font-family:'IBM Plex Mono',monospace !important;font-size:.82rem !important;}
.stButton button{background:linear-gradient(135deg,#0d4a8a,#0a6dd4) !important;color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important;}
.stDownloadButton button{background:linear-gradient(135deg,#0d5c30,#0a8040) !important;color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important;}
.stProgress > div > div{background:#0a6dd4 !important;}
.main{background:#060c18;}
.stExpander{border:1px solid #1a2f48 !important;border-radius:10px !important;background:#080e1a !important;}
.res-table a:hover{text-decoration:underline !important;opacity:.85;}

/* 🆕 v9.18.1：手機優化 — 避免下拉選單觸發鍵盤佔用畫面 */
/* iOS 自動 zoom 防護（input font-size ≥ 16px 不 zoom）*/
[data-baseweb="select"] input,
[data-baseweb="combobox"] input {
    font-size: 16px !important;
    caret-color: transparent !important;
}
/* 觸控裝置：input 完全不接收輸入焦點 */
@media (pointer: coarse), (max-width: 768px) {
    [data-baseweb="select"] input,
    [data-baseweb="combobox"] input {
        pointer-events: none !important;
        user-select: none !important;
        -webkit-user-select: none !important;
    }
    /* 但 multiselect 移除 chip 的 X 仍要可點 */
    [data-baseweb="select"] [role="button"],
    [data-baseweb="select"] svg {
        pointer-events: auto !important;
    }
}
</style>""", unsafe_allow_html=True)

# 🆕 v9.18.1：JS 注入 — 把所有 selectbox/multiselect 的 input 設成 readonly + inputmode=none
# 這是手機端阻止鍵盤的最可靠方法，CSS 只能輔助
import streamlit.components.v1 as _components
_components.html("""
<script>
(function() {
    const isTouch = window.matchMedia('(pointer: coarse)').matches
                     || window.matchMedia('(max-width: 768px)').matches;
    if (!isTouch) return;

    function disableKeyboard(root) {
        // 找所有 selectbox / multiselect 的 input
        const inputs = root.querySelectorAll(
            '[data-baseweb="select"] input, [data-baseweb="combobox"] input'
        );
        inputs.forEach(inp => {
            if (!inp.dataset.kbDisabled) {
                inp.setAttribute('readonly', 'readonly');
                inp.setAttribute('inputmode', 'none');
                inp.setAttribute('autocomplete', 'off');
                inp.setAttribute('autocorrect', 'off');
                inp.setAttribute('autocapitalize', 'off');
                inp.setAttribute('spellcheck', 'false');
                inp.dataset.kbDisabled = '1';
            }
        });
    }

    // 等 parent document 載完，操作 parent（Streamlit iframe 的父）
    function applyToParent() {
        try {
            const doc = window.parent.document;
            disableKeyboard(doc);
            const observer = new MutationObserver(() => disableKeyboard(doc));
            observer.observe(doc.body, { childList: true, subtree: true });
        } catch(e) {}
    }
    if (document.readyState === 'complete') applyToParent();
    else window.addEventListener('load', applyToParent);
})();
</script>
""", height=0, width=0)

# ─────────────────────────────────────────────────────────────────
# LABELS
# ─────────────────────────────────────────────────────────────────
OSC_LABELS = ["RSI(14)","隨機%K","CCI(20)","ADX(14)","AO","動量(10)","MACD",
              "StochRSI","威廉%R","牛熊力度","終極震盪","布林%B"]
MA_LABELS  = ["EMA(10)","SMA(10)","EMA(20)","SMA(20)","EMA(30)","SMA(30)",
              "EMA(50)","SMA(50)","EMA(60)","SMA(60)","EMA(100)","SMA(100)",
              "EMA(200)","SMA(200)","一目均衡基準線","VWMA(20)","Hull MA(9)"]

# ── v9.30：個股詳細卡相關常數+函數已抽到 detail_card_render.py ─────
# tv_app 與 pages/01_intraday.py 共用同一套渲染邏輯
from detail_card_render import (
    GROUP_NAMES, GROUP_WEIGHTS, GROUP_COLORS,
    TREND_W, POSITION_W, MOMENTUM_W, AUX_W,
    _INVERSE_ETF_TICKERS, _CONCEPT_COLORS,
    _j, _jz, _jadx, _rec, fmt,
    judge_oscillators, judge_mas, calc_summary,
    judge_trend, judge_position, judge_momentum,
    compute_momentum_grade, judge_aux, _calc_aux_summary,
    apply_cap, badge, get_rec_label,
    _load_concept_tags, _get_concepts, _concept_chip_html,
    render_detail as _render_detail_core,
)

# ── v9.31：get_operation_advice + 8 helpers 抽到 operation_advice.py ─────
from operation_advice import (
    _get_us_overnight, _load_us_impact, _load_per_stock_wf, _load_clusters,
    _get_proximity_alerts, _get_inverse_etf_advice, _SPECIAL_TICKER_WARN,
    render_confidence_dots, get_operation_advice,
)


def render_detail(ticker, d, groups, group_summs, tsumm, cap, market: str = "") -> str:
    """tv_app 版 render_detail wrapper — 把 get_operation_advice callback
    注入 detail_card_render.render_detail。
    🆕 v9.32：取消新聞情感區塊（news_fn=None）— 不再呼叫網路抓新聞
    🆕 v9.32：標記 _intraday_tf='1d' 讓 operation_advice 跳過靜態 ZigZag PNG，
            由主流程改用 intraday plotly chart 渲染。"""
    # 觸發 operation_advice 跳過內嵌靜態 ZigZag（外部 plotly 接手）
    if isinstance(d, dict):
        d.setdefault('_intraday_tf', '1d')
        d.setdefault('_bar_unit', '天')
    return _render_detail_core(
        ticker, d, groups, group_summs, tsumm, cap, market=market,
        advice_fn=get_operation_advice,
        news_fn=None,                  # 🆕 v9.32：取消新聞讀取
        concepts_fn=_get_concepts,
        concept_chip_fn=_concept_chip_html,
    )


# Excel 匯出仍使用舊格式（保留相容性）
OSC_WEIGHTS = [2.0] * 12
MA_WEIGHTS  = [1.5]*6 + [1.2]*4 + [1.0]*7

# ── 指數 / 特殊代號別名對照 ──────────────────────────────────────
SYMBOL_ALIASES = {
    "DJI":"^DJI","DJIA":"^DJI","SPX":"^GSPC","SP500":"^GSPC",
    "NDX":"^NDX","NASDAQ":"^IXIC","COMP":"^IXIC",
    "VIX":"^VIX","RUT":"^RUT",
    "TWII":"^TWII","TWI":"^TWII",
    "N225":"^N225","NIKKEI":"^N225",
    "HSI":"^HSI","KOSPI":"^KS11",
    "SSE":"000001.SS","SHCOMP":"000001.SS",
    "DAX":"^GDAXI","FTSE":"^FTSE","CAC":"^FCHI",
    "GOLD":"GC=F","OIL":"CL=F","WTI":"CL=F",
    "USDJPY":"JPY=X","EURUSD":"EURUSD=X","USDTWD":"TWD=X",
}

INDEX_NAMES = {
    "^DJI":"道瓊工業指數","^GSPC":"S&P 500","^IXIC":"NASDAQ 綜合指數",
    "^NDX":"NASDAQ 100","^VIX":"VIX 恐慌指數","^RUT":"羅素 2000",
    "^TWII":"台灣加權指數","^N225":"日經 225","^HSI":"恒生指數",
    "^KS11":"韓國 KOSPI","000001.SS":"上證指數",
    "^GDAXI":"德國 DAX","^FTSE":"英國富時 100","^FCHI":"法國 CAC 40",
    "GC=F":"黃金期貨","CL=F":"WTI 原油期貨",
    "JPY=X":"USD/JPY","EURUSD=X":"EUR/USD","TWD=X":"USD/TWD",
}

# ── 台股代號識別 ─────────────────────────────────────────────────
def is_tw_stock(ticker: str) -> bool:
    """台股：純數字 或 數字+單一英文字母結尾（00632R、006205L）"""
    import re
    return bool(re.match(r'^\d+[A-Z]?$', ticker))

def get_yf_symbol(ticker: str) -> str:
    """將使用者輸入代號轉換為 yfinance 查詢格式"""
    if ticker in SYMBOL_ALIASES:
        return SYMBOL_ALIASES[ticker]
    if is_tw_stock(ticker):
        # 🆕 OTC 上櫃股直接用 .TWO（避免 .TW fallback 浪費時間）
        if ticker in _OTC_TICKERS:
            return ticker + ".TWO"
        return ticker + ".TW"
    return ticker


# 🆕 OTC 上櫃股清單（加速 fetch — 跳過 .TW 失敗 retry）
@st.cache_data
def _load_otc_tickers():
    try:
        from pathlib import Path as _P
        import json as _json
        p = _P(__file__).parent / 'otc_tickers.json'
        if p.exists():
            return set(_json.load(open(p, encoding='utf-8')))
    except Exception:
        pass
    return set()
_OTC_TICKERS = _load_otc_tickers()


# 🆕 VWAPEXEC 適用清單（從 build_applicable_list.py 產出）
@st.cache_data(ttl=3600)
def _load_vwap_applicable():
    """載入 VWAPEXEC 適用分級：{ticker: {tier: 'TOP'/'OK'/'NA', delta: float}}"""
    try:
        from pathlib import Path as _P
        import json as _json
        p = _P(__file__).parent / 'vwap_applicable.json'
        if p.exists():
            with open(p, encoding='utf-8') as f:
                return _json.load(f)
    except Exception:
        pass
    return {}


def get_vwap_tier(ticker: str) -> dict:
    """取得 VWAPEXEC 適用分級（tier + delta）"""
    data = _load_vwap_applicable()
    return data.get(ticker, {})

# ── TradingView 圖表連結對照（ticker → TV symbol） ───────────────
TV_CHART_MAP = {
    "DJI":"DJ:DJI","DJIA":"DJ:DJI",
    "SPX":"SP:SPX","SP500":"SP:SPX",
    "NDX":"NASDAQ:NDX",
    "NASDAQ":"NASDAQ:COMP","COMP":"NASDAQ:COMP",
    "VIX":"CBOE:VIX",
    "RUT":"TVC:RUT",
    "TWII":"TWSE:TAIEX","TWI":"TWSE:TAIEX",
    "N225":"INDEX:NKY","NIKKEI":"INDEX:NKY",
    "HSI":"HSI:HSI",
    "KOSPI":"KRX:KOSPI",
    "SSE":"SSE:000001","SHCOMP":"SSE:000001",
    "DAX":"XETR:DAX","FTSE":"LSE:UKX","CAC":"EURONEXT:PX1",
    "GOLD":"TVC:GOLD","OIL":"TVC:USOIL","WTI":"TVC:USOIL",
    "USDJPY":"FX:USDJPY","EURUSD":"FX:EURUSD","USDTWD":"FX:USDTWD",
}

def get_tv_url(ticker: str, market: str) -> str:
    """產生 TradingView 圖表連結"""
    base = "https://www.tradingview.com/chart/?symbol="
    if ticker in TV_CHART_MAP:
        return base + TV_CHART_MAP[ticker]
    if market == "台股":
        # 上櫃(.TWO)用 TPEX，上市(.TW)用 TWSE
        tw_names = _get_tw_names()
        # twstock codes 包含市場資訊
        try:
            import twstock
            info = twstock.codes.get(ticker)
            if info and getattr(info, 'market', '') in ('上櫃', 'OTC'):
                return base + f"TPEX:{ticker}"
        except Exception:
            pass
        return base + f"TWSE:{ticker}"
    if market in ("NASDAQ","NYSE","AMEX","OTC"):
        return base + f"{market}:{ticker}"
    # 🆕 v9.13：US 股票無明確交易所 → 不加前綴讓 TradingView 自動偵測
    # 修：原本 default 'NASDAQ' 對 NYSE 股（PL/JPM/BAC 等）會產生不存在 URL
    return base + ticker

def _build_company_info_prompt(ticker: str, name: str, d: dict) -> str:
    """🆕 v9.18：點 ticker 改為查公司基本資訊 + 業務 + 影響因素 + 連動概念
    返回 Perplexity / 其他 AI 用的純文字提示詞。"""
    close  = d.get("close")  or 0
    sma50  = d.get("sma50")  or 0
    sma200 = d.get("sma200") or 0
    is_tw = ticker.isdigit() and 4 <= len(ticker) <= 6 and not ticker.startswith('00')
    is_etf = ticker.startswith('00')
    market_label = ('台股 ETF' if is_etf else '台股' if is_tw else '美股')

    # 已知概念股標籤
    try:
        concepts = _get_concepts(ticker, max_n=8)
        concept_str = '、'.join(concepts) if concepts else '（無）'
    except Exception:
        concept_str = '（無資料）'

    display = f'{ticker}（{name}）' if name and name != ticker else ticker

    lines = [
        f"請提供 {market_label} {display} 的全面公司分析報告，重點放在「股價會被什麼影響」。",
        "",
        "【一、公司基本資料】",
        "- 公司全名、成立年份、總部地點",
        "- 所屬產業 / 子產業類別",
        "- 主要業務（請具體列出產品線或服務項目）",
        "- 市值、員工數、最新一年營收規模",
        "- 主要股東結構（前 3-5 大持股 + 持股比例）",
        "",
        "【二、目前營運狀況】",
        "- 最新一季 / 一年財報重點：營收、毛利率、EPS、年增/季增",
        "- 主要產品線的營收佔比",
        "- 近一年股價走勢與重大事件",
        "",
        "【三、發展方向與策略】",
        "- 未來 1-3 年的成長策略重點（明確列出具體計畫）",
        "- 主要產品 / 服務 roadmap",
        "- 資本支出、研發投入、新市場擴張",
        "- M&A、轉型、品牌策略（若有）",
        "",
        "【四、合作對象與生態圈】",
        "- 主要客戶（前 5 大，含營收佔比 % 若可得）",
        "- 主要供應商 / 上下游夥伴",
        "- 策略合作夥伴 / 聯盟",
        "- 主要競爭對手（國內 + 國際同業）",
        "",
        "【五、影響股價的關鍵因素 ⭐核心⭐】",
        "- 受哪些大盤指數連動高（如 ^GSPC SPY / TWII / ^SOX 半導體）",
        "- 受哪些總體因素影響（利率、匯率、原物料、地緣政治）",
        "- 受哪些產業趨勢驅動（請具體：AI、半導體、5G、ESG、電動車、晶圓代工、雲端...）",
        "- 受哪些公司股價連動高（β 高的標的，例如台積電、輝達、Apple）",
        "- 特定 catalyst / 風險事件（法規、訴訟、產能擴張、訂單）",
        "- 季節性因素（旺季 / 淡季 / 法說會 / 財報季 / 配息）",
        "",
        "【六、概念股 / 技術連動分析】",
        f"- 已知概念股標籤（用戶系統內）：{concept_str}",
        "- 還有哪些主流概念可以歸類進去？（補充缺漏）",
        "- 是否為 ETF 成分股（哪些 ETF？影響 passive 資金流向）",
        "- 與其他同類股的比較（誰是 leader、誰落後）",
        "",
        "【七、目前股價技術面 quick check】",
        f"- 現價約 {close:.2f}，50 日均線 {sma50:.2f}，200 日均線 {sma200:.2f}",
        "- 趨勢判斷（多 / 空 / 盤整）",
        "- 近期關鍵支撐 / 壓力位",
        "",
        "【八、短中長期展望 + 主要風險】",
        "- 短期（1-3 個月）關注的重點 catalyst",
        "- 中長期（6-12 個月）展望",
        "- 主要下行風險（請具體列出）",
        "",
        "輸出格式要求：",
        "- 繁體中文",
        "- 每個重點請用條列 + 簡要表格呈現",
        "- 資訊請以「最新可得的日期」為主，並在每節末尾註明資料時間",
        "- 避免空洞形容詞，重點放在「具體事實」與「可行動的洞察」",
        "- 若資料不足或不確定，請明確標示「未確認」",
    ]
    return "\n".join(lines)


def get_ai_url(ticker: str, name: str, d: dict,
               platform_url_tpl: str = "https://www.perplexity.ai/search?q={prompt}") -> str:
    """v9.18：改成公司資訊問答 prompt"""
    import urllib.parse
    prompt = _build_company_info_prompt(ticker, name, d)
    encoded = urllib.parse.quote(prompt)
    return platform_url_tpl.replace("{prompt}", encoded)


def get_prompt_text(ticker: str, name: str, d: dict) -> str:
    """回傳純文字提示詞（供複製用，與 get_ai_url 共用同一份 prompt）"""
    return _build_company_info_prompt(ticker, name, d)


# 🆕 v9.18：保留舊版技術分析 prompt（detail card AI expander 仍需用）
def get_technical_analysis_prompt(ticker: str, name: str, d: dict) -> str:
    """舊版技術分析 prompt（布林通道 + 均線 + 買賣點）— 用於 detail card AI 解讀"""
    close  = d.get("close")  or 0
    sma50  = d.get("sma50")  or 0
    sma200 = d.get("sma200") or 0
    bbu    = d.get("bbu")    or 0
    bbl    = d.get("bbl")    or 0
    bbm    = d.get("ema20")  or 0
    display = f"{ticker}（{name}）" if name and name != ticker else ticker
    lines = [
        f"你是一位專業量化交易員，請針對 {display} 用日線 + 布林通道為主做技術分析，判斷短中期買賣點。",
        "",
        f"現價約 {close:.2f}，50 日均線 {sma50:.2f}，200 日均線 {sma200:.2f}，布林上軌 {bbu:.2f}、中軌 {bbm:.2f}、下軌 {bbl:.2f}。",
        "",
        "分析項目：1) 趨勢與均線位置 2) 布林通道狀態（張口/收口） 3) 買點條件 4) 賣點條件 5) 停損停利建議",
        "繁體中文，條列 + 表格呈現。",
    ]
    return "\n".join(lines)


# ── 🆕 v9.10t：美股盤後快訊（即時抓昨夜美股報酬）──────






# ── 🆕 v9.10u：5 大發現整合 ────────────────────────────────────


@st.cache_data(ttl=86400, show_spinner=False)
def _get_sector_ranking_recent() -> list:
    """過去 1 個月各產業平均報酬排名（用 vwap_applicable 中的個股月報酬）
    回傳 [(sector, monthly_return%, n_stocks), ...] 由強到弱排序"""
    from pathlib import Path as _P
    import json as _json
    # 讀行業
    industry_map = {}
    p = _P(__file__).parent / 'tw_universe.txt'
    if not p.exists(): return []
    for line in p.read_text(encoding='utf-8').splitlines():
        if not line or line.startswith('#'): continue
        parts = line.split('|')
        if len(parts) >= 5 and parts[4]:
            industry_map[parts[0].strip()] = parts[4].strip()

    # 用 yfinance 抓近 30 日資料（避免依賴 data_cache）
    sectors = {}
    for tk, ind in industry_map.items():
        sectors.setdefault(ind, []).append(tk)

    # 只跑流動性夠的（取 vwap_applicable）
    vwap_path = _P(__file__).parent / 'vwap_applicable.json'
    top_set = set()
    if vwap_path.exists():
        try:
            vp = _json.loads(vwap_path.read_text(encoding='utf-8'))
            top_set = set(t for t, info in vp.items()
                          if info.get('tier') in ('TOP', 'OK'))
        except: pass

    # 對每產業計算月報酬 (用 yfinance batch)
    out = []
    for sec, members in sectors.items():
        members = [t for t in members if t in top_set][:20]  # 每產業最多 20 檔
        if len(members) < 5: continue
        rets = []
        try:
            yf_tk = [f'{t}.TW' for t in members]
            df = yf.download(yf_tk, period='1mo', interval='1d',
                             progress=False, group_by='ticker', threads=True,
                             auto_adjust=False)
            if df is None or df.empty: continue
            for t in yf_tk:
                try:
                    sub = df[t] if len(yf_tk) > 1 else df
                    sub = sub.dropna(how='all')
                    if len(sub) < 15: continue
                    ret = (sub['Close'].iloc[-1] - sub['Close'].iloc[0]) / sub['Close'].iloc[0] * 100
                    rets.append(ret)
                except: continue
        except Exception:
            continue
        if rets:
            out.append((sec, float(np.mean(rets)), len(rets)))
    out.sort(key=lambda x: -x[1])
    return out


@st.cache_data(ttl=3600, show_spinner=False)
# ── 🆕 市場廣度警報（D：偵測「指數漲、廣度差」失效市況）──────
@st.cache_data(ttl=3600, show_spinner=False)
def _get_market_breadth() -> dict:
    """從 yfinance 取台股大盤 + 計算廣度。
    回傳 {twii_60d_chg, breadth_pct, alert_level, msg}
    """
    try:
        twii = yf.Ticker('^TWII').history(period='3mo', auto_adjust=True)
        if len(twii) < 60: return {'has_data': False}
        cur = float(twii['Close'].iloc[-1])
        c60 = float(twii['Close'].iloc[-60])
        twii_60d_chg = (cur - c60) / c60 * 100

        # 用幾檔權值股代理計算廣度（取重要 30 檔，每檔看是否仍 > EMA60）
        # 這是廣度的快速近似，避免抓全市場
        proxies = ['2330','2317','2454','2412','2882','2891','2308','2382',
                   '2884','2603','2885','2886','3008','2880','2207','1101',
                   '1216','1303','2002','2615','2609','2610','3711','2474',
                   '6505','3034','2618','3037','2890','2887']
        above_ema60 = 0; total = 0
        for t in proxies:
            try:
                df = yf.Ticker(f'{t}.TW').history(period='3mo',
                                                    auto_adjust=True)
                if len(df) < 60: continue
                ema60 = df['Close'].ewm(span=60, adjust=False).mean().iloc[-1]
                cur_p = df['Close'].iloc[-1]
                if cur_p > ema60: above_ema60 += 1
                total += 1
            except Exception:
                continue
        if total < 10: return {'has_data': False}
        breadth_pct = above_ema60 / total * 100

        # 警戒級別
        if twii_60d_chg > 5 and breadth_pct < 40:
            level = 'red'
            msg = (f'⚠️ 市場廣度警示：大盤 60 日 +{twii_60d_chg:.1f}% '
                   f'但僅 {breadth_pct:.0f}% 權值股站上 EMA60。'
                   f'指數由少數股拉動、多數股盤整下跌——v8 系統可能結構性失效（2024 範式）')
        elif twii_60d_chg > 5 and breadth_pct < 60:
            level = 'yellow'
            msg = (f'⚠️ 市場廣度偏窄：大盤 60 日 +{twii_60d_chg:.1f}% '
                   f'但僅 {breadth_pct:.0f}% 權值股站上 EMA60，主升段集中於少數股')
        elif twii_60d_chg < -5 and breadth_pct < 40:
            level = 'red_bear'
            msg = (f'⚠️ 全面空頭：大盤 60 日 {twii_60d_chg:+.1f}% + '
                   f'僅 {breadth_pct:.0f}% 權值股站上 EMA60')
        else:
            level = 'green'
            msg = (f'✓ 市場廣度健康：大盤 60 日 {twii_60d_chg:+.1f}%、'
                   f'{breadth_pct:.0f}% 權值股站上 EMA60')
        return {
            'has_data': True,
            'twii_60d_chg': twii_60d_chg,
            'breadth_pct': breadth_pct,
            'level': level, 'msg': msg,
        }
    except Exception:
        return {'has_data': False}


@st.cache_data(ttl=86400, show_spinner=False)
def _get_tw_names() -> dict:
    """台股中文名稱字典：合併 twstock + 靜態 tw_universe.txt（雲端可靠）"""
    out = {}
    # 1. 靜態檔（雲端首選，本地 fallback）
    try:
        from pathlib import Path as _P
        f = _P(__file__).parent / 'tw_universe.txt'
        if f.exists():
            for line in f.read_text(encoding='utf-8').splitlines():
                if not line or line.startswith('#'): continue
                parts = line.split('|')
                if len(parts) >= 2:
                    out[parts[0].strip()] = parts[1].strip()
    except Exception:
        pass
    # 2. twstock 動態（本地有，會覆蓋為最新）
    try:
        import twstock
        for code, info in twstock.codes.items():
            out[str(code)] = info.name
    except Exception:
        pass
    return out

# ── 概念股標籤 ─────────────────────────────────────────────────
# 美股 GICS sector 中文對照
_US_SECTOR_ZH = {
    "Information Technology": "資訊科技",
    "Financials": "金融",
    "Health Care": "醫療保健",
    "Consumer Discretionary": "非必需消費",
    "Industrials": "工業",
    "Communication Services": "通訊服務",
    "Consumer Staples": "必需消費",
    "Energy": "能源",
    "Utilities": "公用事業",
    "Real Estate": "不動產",
    "Materials": "原材料",
    "ETF": "ETF",
}

# 🆕 v9.30：_load_concept_tags / _get_concepts 已搬到 detail_card_render.py（上方 import）

@st.cache_data(ttl=86400, show_spinner=False)
def _load_industry_map() -> dict:
    """單一字串產業類別（簡短）：
    台股：tw_universe.txt 第 5 欄；ETF/ETN/特別股顯示類型
    美股：us_sectors.txt sector（中譯）"""
    from pathlib import Path as _P
    base = _P(__file__).parent
    out = {}
    p2 = base / 'tw_universe.txt'
    if p2.exists():
        try:
            for line in p2.read_text(encoding='utf-8').splitlines():
                if not line or line.startswith('#'): continue
                parts = line.split('|')
                if len(parts) < 5: continue
                ticker, _n, _typ, _mkt, industry = parts[:5]
                if industry:
                    out[ticker] = industry
                elif _typ in ('ETF','ETN','特別股','臺灣存託憑證(TDR)'):
                    out[ticker] = ('TDR' if _typ=='臺灣存託憑證(TDR)'
                                   else _typ)
        except Exception:
            pass
    p3 = base / 'us_sectors.txt'
    if p3.exists():
        try:
            for line in p3.read_text(encoding='utf-8').splitlines():
                if not line or line.startswith('#'): continue
                parts = line.split('|')
                if len(parts) < 4: continue
                ticker, _name, sector, _sub = parts[:4]
                if sector and sector != 'NO_SECTOR':
                    out[ticker] = _US_SECTOR_ZH.get(sector, sector)
        except Exception:
            pass
    return out

def _get_industry(ticker: str) -> str:
    return _load_industry_map().get(ticker, "")

# 🆕 v9.30：_CONCEPT_COLORS / _concept_chip_html 已搬到 detail_card_render.py

# 常用美股/ETF 靜態名稱對照（避免 get_info() 在雲端失敗）
US_NAMES = {
    "BOTZ": "Global X Robotics & AI ETF",
    "NVDA": "NVIDIA Corporation",
    "AAPL": "Apple Inc.",
    "MSFT": "Microsoft Corporation",
    "TSLA": "Tesla Inc.",
    "AMZN": "Amazon.com Inc.",
    "GOOGL": "Alphabet Inc.",
    "META": "Meta Platforms Inc.",
    "SPY": "SPDR S&P 500 ETF",
    "QQQ": "Invesco QQQ Trust",
    "ARKK": "ARK Innovation ETF",
    "GLD": "SPDR Gold Shares",
    "TLT": "iShares 20+ Year Treasury Bond ETF",
}

@st.cache_data(ttl=86400, show_spinner=False)
def _get_us_names_static() -> dict:
    """從 us_names.txt 載入靜態名稱對照（~8000 檔）"""
    out = {}
    try:
        from pathlib import Path as _P
        f = _P(__file__).parent / 'us_names.txt'
        if f.exists():
            for line in f.read_text(encoding='utf-8').splitlines():
                if not line or line.startswith('#'): continue
                parts = line.split('|', 1)
                if len(parts) >= 2:
                    out[parts[0].strip()] = parts[1].strip()
    except Exception:
        pass
    return out

def _get_stock_name(ticker: str, symbol: str) -> str:
    """取得股票中文/英文名稱"""
    try:
        # 1. 指數/特殊代號靜態對照
        if symbol in INDEX_NAMES:
            return INDEX_NAMES[symbol]
        # 2. 台股 twstock + 靜態 tw_universe.txt
        if is_tw_stock(ticker):
            tw_names = _get_tw_names()
            return tw_names.get(ticker, ticker)
        # 3. 常用美股靜態對照（內建 13 檔 + us_names.txt 8000+ 檔）
        if ticker in US_NAMES:
            return US_NAMES[ticker]
        us_static = _get_us_names_static()
        if ticker in us_static:
            return us_static[ticker]
        # 4. yfinance 動態查詢（最後 fallback）
        try:
            t = yf.Ticker(symbol)
            info = t.get_info()
            return info.get("longName") or info.get("shortName") or ticker
        except Exception:
            return ticker
    except Exception:
        return ticker

def hull_ma(series: pd.Series, n: int = 9) -> pd.Series:
    """Hull Moving Average"""
    half = max(n // 2, 1)
    w1 = series.rolling(half).apply(
        lambda x: np.average(x, weights=range(1, len(x)+1)), raw=True)
    w2 = series.rolling(n).apply(
        lambda x: np.average(x, weights=range(1, len(x)+1)), raw=True)
    diff = 2 * w1 - w2
    sqn = max(int(np.sqrt(n)), 1)
    return diff.rolling(sqn).apply(
        lambda x: np.average(x, weights=range(1, len(x)+1)), raw=True)

def vwma(close: pd.Series, volume: pd.Series, n: int = 20) -> pd.Series:
    """Volume Weighted Moving Average"""
    return (close * volume).rolling(n).sum() / volume.rolling(n).sum()


@st.cache_data(ttl=1800, show_spinner=False)
def fetch_news(ticker: str, market: str) -> list:
    """取得最新 3 條新聞：台股用 Google News RSS，美股用 yfinance"""
    if is_tw_stock(ticker):
        return _fetch_news_google(ticker, market)
    try:
        symbol = get_yf_symbol(ticker)
        news = yf.Ticker(symbol).news or []
        result = []
        for item in news[:3]:
            content = item.get("content", {})
            title   = content.get("title") or item.get("title", "")
            link    = (content.get("canonicalUrl", {}).get("url")
                       or content.get("clickThroughUrl", {}).get("url")
                       or item.get("link", ""))
            pub     = (content.get("provider", {}).get("displayName")
                       or item.get("publisher", ""))
            if title and link:
                result.append({"title": title, "link": link, "publisher": pub})
        return result if result else _fetch_news_google(ticker, market)
    except Exception:
        return _fetch_news_google(ticker, market)


# ── 🆕 v3 金融領域情感系統（純規則，移除 SnowNLP）─────────────────
# 設計原則：
#   1. SnowNLP 訓練資料是電商評論，對金融標題誤判嚴重 → 完全移除
#   2. 純金融字典（正/負各 100+ 詞）
#   3. regex 數字強度 + 詞彙否定處理
#   4. 短語優先於單字（避免「成長」匹配到「衰退成長率」）

import re as _re

# ─── 強烈正面（每個 +0.30 ~ +0.40）─────────────────
_FIN_POS_STRONG = {
    # 價格行為
    '大漲': 0.35, '飆漲': 0.40, '飆升': 0.40, '飆破': 0.35, '創新高': 0.40,
    '創歷史新高': 0.45, '締造紀錄': 0.35, '長紅': 0.30, '攻頂': 0.30,
    # 業績/財報
    '業績亮眼': 0.40, '業績超標': 0.40, '業績爆發': 0.45, '業績大躍進': 0.45,
    '獲利大增': 0.35, '獲利躍進': 0.40, '獲利暴衝': 0.40, '營收創新高': 0.40,
    '盈餘上修': 0.35, '財報優': 0.35, '財測上修': 0.35,
    # 訂單/需求
    '訂單滿載': 0.40, '訂單爆量': 0.40, '需求強勁': 0.35, '需求暢旺': 0.35,
    '出貨暢旺': 0.35, '搶到訂單': 0.35, '取得大單': 0.35, '簽下訂單': 0.30,
    # 商機/題材
    '爆發商機': 0.40, '商機爆發': 0.40, '商機湧現': 0.35, '全面引爆': 0.40,
    '激增': 0.35, '暴增': 0.35, '蜂擁': 0.30, '搶買': 0.30, '搶購': 0.30,
}

# ─── 中等正面（每個 +0.15 ~ +0.25）─────────────────
_FIN_POS_MEDIUM = {
    # 價格
    '上漲': 0.18, '收漲': 0.18, '勁揚': 0.25, '走揚': 0.20, '反彈': 0.15,
    '回升': 0.18, '止跌': 0.18, '突破': 0.20, '攻上': 0.22, '直衝': 0.22,
    '創高': 0.25, '改寫紀錄': 0.25, '達陣': 0.20, '達標': 0.20,
    # 評等
    '看好': 0.25, '看多': 0.25, '加碼': 0.22, '增持': 0.22, '買進': 0.20,
    '推薦': 0.18, '優於大盤': 0.22, 'Outperform': 0.22, '持續看好': 0.25,
    # 利多
    '利多': 0.25, '利好': 0.25, '受惠': 0.20, '受惠商機': 0.30, '受惠政策': 0.20,
    '搶單': 0.25, '搶先': 0.18, '卡位': 0.18, '搶占': 0.20, '搶攻': 0.20,
    # 業績
    '亮眼': 0.22, '亮麗': 0.22, '佳音': 0.20, '報喜': 0.22, '看漲': 0.22,
    '獲利成長': 0.22, '營收成長': 0.22, '上修': 0.22, '調升': 0.20,
    '營收年增': 0.22, '展望樂觀': 0.20,
    # 法人
    '法人加碼': 0.22, '法人買超': 0.20, '外資買超': 0.22, '外資加碼': 0.22,
    '投信買超': 0.18, '主力買超': 0.18,
    # 訂單能見度
    '訂單能見度': 0.20, '能見度高': 0.22, '產能滿載': 0.30,
    # 題材/趨勢
    '藍海': 0.25, '商機': 0.18, '爆發': 0.18, '引爆': 0.20,
    'AI 概念': 0.15, 'AI受惠': 0.22, '熱錢': 0.15, '吸金': 0.18,
}

# ─── 微弱正面（每個 +0.05 ~ +0.10）─────────────────
_FIN_POS_WEAK = {
    '高達': 0.12, '達到': 0.05, '創高': 0.10, '增': 0.05, '揚': 0.08,
    '年增率': 0.12, '年增': 0.12, '月增': 0.08, '季增': 0.10,
    '提升': 0.10, '改善': 0.12, '回神': 0.12, '熱絡': 0.12,
    '入列': 0.10, '入選': 0.10, '受邀': 0.10, '指名': 0.12, '點名': 0.10,
    '看好': 0.18, '受惠': 0.15, '搶手': 0.15, '搶卡': 0.15,
    '佳績': 0.18, '佳音': 0.18, '佳評': 0.15, '長期看好': 0.20,
}

# ─── 強烈負面（每個 -0.30 ~ -0.40）─────────────────
_FIN_NEG_STRONG = {
    # 價格行為
    '大跌': -0.35, '崩跌': -0.40, '崩盤': -0.45, '暴跌': -0.40, '重挫': -0.35,
    '跌停': -0.45, '跌停板': -0.45, '停板': -0.35, '創新低': -0.30,
    '探底': -0.25, '腰斬': -0.40,
    # 業績/財報
    '虧損': -0.35, '財報差': -0.35, '獲利衰退': -0.35, '獲利大減': -0.35,
    '營收下滑': -0.30, '營收衰退': -0.35, '盈餘下修': -0.35, '財測下修': -0.35,
    # 公司危機
    '違約': -0.40, '掏空': -0.45, '破產': -0.45, '下市': -0.45,
    '停牌': -0.40, '查封': -0.40, '裁員': -0.30, '縮編': -0.25,
    # 情緒
    '恐慌': -0.30, '崩潰': -0.35, '股災': -0.40, '股殤': -0.40,
    # 利空
    '利空': -0.30, '警訊': -0.25, '警告': -0.25, '危機': -0.30,
}

# ─── 中等負面（每個 -0.15 ~ -0.25）─────────────────
_FIN_NEG_MEDIUM = {
    # 價格
    '下跌': -0.18, '收跌': -0.18, '走低': -0.18, '挫低': -0.22, '修正': -0.15,
    '跌深': -0.22, '跌破': -0.20, '反轉': -0.18, '跌幅擴大': -0.22,
    # 評等
    '看壞': -0.25, '看空': -0.25, '減碼': -0.22, '降評': -0.22, '不推': -0.22,
    '賣超': -0.15, '出脫': -0.18, '劣於大盤': -0.22, 'Underperform': -0.22,
    # 業績
    '衰退': -0.25, '疲軟': -0.20, '低迷': -0.22, '陰霾': -0.20, '失利': -0.20,
    '營收年減': -0.22, '下修': -0.22, '調降': -0.20, '展望保守': -0.18,
    '展望黯淡': -0.25, '展望疲弱': -0.22,
    # 法人
    '法人賣超': -0.22, '法人撤': -0.20, '外資賣超': -0.22, '外資撤': -0.22,
    '主力賣超': -0.18,
    # 套牢
    '套牢': -0.22, '失血': -0.22, '出走': -0.18, '逃命': -0.20, '出清': -0.18,
    # 失守/警訊
    '失守': -0.20, '空頭': -0.22, '空襲': -0.22, '失靈': -0.20,
}

# ─── 微弱負面（每個 -0.05 ~ -0.10）─────────────────
_FIN_NEG_WEAK = {
    '減': -0.05, '降': -0.05, '少': -0.03, '弱': -0.08,
    '年減': -0.15, '月減': -0.10, '季減': -0.10, '月減率': -0.12,
    '滑落': -0.12, '下滑': -0.12, '收斂': -0.10,
    '盤整': -0.05, '震盪': -0.05, '觀望': -0.05,
    '持平': 0.00,
}

# 合併
_FIN_POS_KEYWORDS = {**_FIN_POS_STRONG, **_FIN_POS_MEDIUM, **_FIN_POS_WEAK}
_FIN_NEG_KEYWORDS = {**_FIN_NEG_STRONG, **_FIN_NEG_MEDIUM, **_FIN_NEG_WEAK}

# 否定詞（出現在關鍵字前 → 反向）
_NEGATION_WORDS = ['不', '未', '沒', '無', '非']

# ─── 數字模式（regex）─────────────────────────
_PCT_POS_PATTERNS = [
    (r'年增[率]?[\s高達]*(\d+(?:\.\d+)?)\s*%',     0.006),  # 年增率 146% → +0.50（封頂）
    (r'(?:月增|季增)[率]?(\d+(?:\.\d+)?)\s*%',     0.004),
    (r'營收[漲增上]\s*(\d+(?:\.\d+)?)\s*%',         0.005),
    (r'(?:大漲|飆升|飆漲|急漲|暴漲)\s*(\d+(?:\.\d+)?)\s*%', 0.012),
    (r'獲利(?:成長|大增)\s*(\d+(?:\.\d+)?)\s*[倍%]', 0.05),  # N 倍 → 強放大
]
_PCT_NEG_PATTERNS = [
    (r'(?:年減|月減|季減)[率]?(\d+(?:\.\d+)?)\s*%', -0.006),
    (r'營收[減衰退]\s*(\d+(?:\.\d+)?)\s*%',         -0.005),
    (r'(?:大跌|暴跌|崩跌|重挫|急跌)\s*(\d+(?:\.\d+)?)\s*%', -0.012),
    (r'獲利(?:衰退|減少)\s*(\d+(?:\.\d+)?)\s*%',    -0.006),
]


# 🆕 v9.9e：BERT 中文情感（IDEA-CCNL/Erlangshen-Roberta-110M-Sentiment）
# 與規則混合（規則 70% + BERT 30%），規則對金融術語更準
@st.cache_resource(show_spinner=False)
def _load_sentiment_bert():
    try:
        from transformers import AutoTokenizer, AutoModelForSequenceClassification
        import torch
        mid = 'IDEA-CCNL/Erlangshen-Roberta-110M-Sentiment'
        tok = AutoTokenizer.from_pretrained(mid)
        mdl = AutoModelForSequenceClassification.from_pretrained(mid)
        mdl.eval()
        return tok, mdl, torch
    except Exception:
        return None, None, None


def _bert_sentiment(title: str) -> float:
    """BERT 情感分數 -1 ~ +1。失敗 → 0"""
    if not title: return 0.0
    tok, mdl, torch = _load_sentiment_bert()
    if tok is None: return 0.0
    try:
        inputs = tok(title, return_tensors='pt', truncation=True, max_length=128)
        with torch.no_grad():
            logits = mdl(**inputs).logits
        probs = torch.softmax(logits, dim=-1)[0]
        return float(probs[1] - probs[0])  # pos - neg
    except Exception:
        return 0.0


def _score_news_title(title: str) -> float:
    """v9.9e 混合評分：規則 70% + BERT 30%（金融術語規則為主，通用情緒 BERT 補強）
    回傳 -1.0 ~ +1.0
    """
    if not title: return 0.0
    score = 0.0
    matched_pos = []
    matched_neg = []

    # 1. 正向關鍵字（按長度排序，優先匹配長詞，避免「商機」優先於「商機爆發」）
    sorted_pos = sorted(_FIN_POS_KEYWORDS.items(), key=lambda x: -len(x[0]))
    sorted_neg = sorted(_FIN_NEG_KEYWORDS.items(), key=lambda x: -len(x[0]))

    consumed_pos = title  # 移除已匹配段落，避免短詞重複命中
    for kw, w in sorted_pos:
        if kw in consumed_pos:
            # 否定處理：檢查 kw 前 2 字是否含否定詞
            idx = consumed_pos.find(kw)
            prefix = consumed_pos[max(0, idx-2):idx]
            if any(neg in prefix for neg in _NEGATION_WORDS):
                # 否定 → 反向 50%
                score -= w * 0.5
                matched_neg.append((f'不{kw}', -w*0.5))
            else:
                score += w
                matched_pos.append((kw, w))
            consumed_pos = consumed_pos.replace(kw, '_'*len(kw), 1)

    consumed_neg = title
    for kw, w in sorted_neg:
        if kw in consumed_neg:
            idx = consumed_neg.find(kw)
            prefix = consumed_neg[max(0, idx-2):idx]
            if any(neg in prefix for neg in _NEGATION_WORDS):
                score -= w * 0.5  # 不跌 → 正面
                matched_pos.append((f'不{kw}', -w*0.5))
            else:
                score += w
                matched_neg.append((kw, w))
            consumed_neg = consumed_neg.replace(kw, '_'*len(kw), 1)

    # 2. 數字模式
    for pat, mul in _PCT_POS_PATTERNS:
        for m in _re.finditer(pat, title):
            try:
                pct = float(m.group(1))
                score += min(0.5, pct * mul)
            except Exception: pass
    for pat, mul in _PCT_NEG_PATTERNS:
        for m in _re.finditer(pat, title):
            try:
                pct = float(m.group(1))
                score += max(-0.5, pct * mul)
            except Exception: pass

    # 3. BERT 補強（規則 70% + BERT 30%）
    bert_s = _bert_sentiment(title)
    rule_score = score
    if matched_pos or matched_neg:
        # 有規則命中 → 規則為主，BERT 微調
        score = rule_score * 0.7 + bert_s * 0.3
    else:
        # 規則無命中 → 完全靠 BERT（縮小至 ±0.5 避免過度極端）
        score = bert_s * 0.5

    return max(-1.0, min(1.0, score))


@st.cache_data(ttl=1800)
def get_news_sentiment(ticker: str, market: str) -> dict:
    """為個股新聞算情感分數。
    回傳：{ avg_score: float, n: int, headlines: [(title, score, link), ...] }"""
    news = fetch_news(ticker, market)
    if not news:
        return {'avg_score': 0.0, 'n': 0, 'headlines': []}
    headlines = []
    for n in news:
        title = n.get('title', '')
        link = n.get('link', '')
        s = _score_news_title(title)
        headlines.append((title, s, link, n.get('publisher', '')))
    avg = sum(h[1] for h in headlines) / len(headlines)
    return {'avg_score': round(avg, 3), 'n': len(headlines), 'headlines': headlines}


def _fetch_news_google(ticker: str, market: str) -> list:
    """Google News RSS 備援（支援台股）"""
    try:
        import xml.etree.ElementTree as ET
        # 台股加上公司名稱一起搜尋，命中率更高
        name = ""
        if is_tw_stock(ticker):
            tw_names = _get_tw_names()
            name = tw_names.get(ticker, "")
        query = f"{ticker} {name}".strip() if name else ticker
        import urllib.parse
        url = (f"https://news.google.com/rss/search?q={urllib.parse.quote(query)}"
               f"&hl=zh-TW&gl=TW&ceid=TW:zh-Hant")
        resp = requests.get(url, timeout=6,
                            headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code != 200:
            return []
        root = ET.fromstring(resp.content)
        items = root.findall(".//item")
        result = []
        for item in items[:3]:
            title = (item.findtext("title") or "").strip()
            link  = (item.findtext("link") or "").strip()
            pub   = (item.findtext("source") or "").strip()
            if title and link:
                result.append({"title": title, "link": link, "publisher": pub})
        return result
    except Exception:
        return []


@st.cache_data(ttl=1800, show_spinner=False)
def fetch_indicators(ticker: str, market: str, end_date: str = "", _cache_ver: str = ""):
    """end_date: 'YYYY-MM-DD' 指定截止日期（空字串=最新）；歷史日期快取永久有效
    🆕 v9.20.6：_cache_ver 參數讓版本變更時自動 invalidate cache（避免舊 d 殘留）"""
    symbol = get_yf_symbol(ticker)
    df = None
    last_err = None
    is_tw = symbol.endswith(".TW") or symbol.endswith(".TWO")

    # 計算 end 參數（yfinance 的 end 是「不包含」那天，所以要加 1 天）
    import datetime as _dt
    if end_date:
        _end_dt   = _dt.date.fromisoformat(end_date) + _dt.timedelta(days=1)
        _end_str  = _end_dt.isoformat()
        _start_dt = _dt.date.fromisoformat(end_date) - _dt.timedelta(days=730)
        _start_str = _start_dt.isoformat()
    else:
        _end_str = _start_str = None   # 使用 period 模式

    def _try_tw_download(sym):
        """嘗試下載台股資料，回傳 DataFrame 或 None
        修正：閾值 20 → 100。20 rows 對 indicator 計算太少（需 200+ 日暖機）
        且會誤觸 .TW（上市）剛好給 20 行卻沒觸發 .TWO fallback 的 bug。
        """
        MIN_ROWS = 100
        if _start_str:
            raw = yf.download(
                sym, start=_start_str, end=_end_str, interval="1d",
                progress=False, auto_adjust=False, multi_level_index=False,
            )
            if raw is not None and len(raw) >= MIN_ROWS:
                _c = raw.get("Close", raw.get("close", pd.Series()))
                if not _c.dropna().empty:
                    return raw
            return None
        for _period, _adj in [("2y", False), ("1y", False), ("2y", True), ("1y", True)]:
            raw = yf.download(
                sym, period=_period, interval="1d",
                progress=False, auto_adjust=_adj,
                multi_level_index=False,
            )
            if raw is not None and len(raw) >= MIN_ROWS:
                _c = raw.get("Close", raw.get("close", pd.Series()))
                if not _c.dropna().empty:
                    return raw
        return None

    for attempt in range(3):
        try:
            if is_tw:
                # 先試 .TW（上市），失敗再試 .TWO（上櫃）
                df = _try_tw_download(symbol)
                if df is None and symbol.endswith(".TW"):
                    alt = ticker + ".TWO"
                    df = _try_tw_download(alt)
                    if df is not None:
                        symbol = alt  # 更新為正確代號
            else:
                if _start_str:
                    df = yf.Ticker(symbol).history(start=_start_str, end=_end_str, interval="1d")
                else:
                    df = yf.Ticker(symbol).history(period="1y", interval="1d")
                if df is not None and len(df) < 30:
                    df = None
            if df is not None and len(df) >= 30:
                break
            last_err = f"rows={len(df) if df is not None else 0}"
            df = None
        except Exception as e:
            last_err = str(e)[:120]
            df = None
        if df is None and attempt < 2:
            time.sleep(3 + attempt * 3)
    if df is None or len(df) < 20:
        rows = len(df) if df is not None else 0
        return {"_error": f"rows={rows}"}
    try:
        # 抓取股票名稱
        name = _get_stock_name(ticker, symbol)
        # 統一欄位名稱 (Ticker.history 和 yf.download 格式不同)
        # yf.download 可能有 MultiIndex (Price, Ticker) 或 (Price,) 格式
        if hasattr(df.columns, 'levels'):
            # MultiIndex - 取第一層
            df.columns = [str(col[0]).strip().capitalize()
                          if isinstance(col, tuple) else str(col).strip().capitalize()
                          for col in df.columns]
        else:
            df.columns = [str(col).strip().capitalize() for col in df.columns]
        # 確認必要欄位存在
        col_map = {c.lower(): c for c in df.columns}
        def get_col(name):
            return df[col_map[name]] if name in col_map else pd.Series(dtype=float)
        c = get_col("close")
        h = get_col("high")
        l = get_col("low")
        v = get_col("volume")
        if c.dropna().empty:
            return {"_error": "close all NaN"}

        def last(s):
            # 取最後一個非 NaN 值（避免最新行有 NaN 的問題）
            s_clean = s.dropna()
            if s_clean.empty:
                return None
            return float(s_clean.iloc[-1])

        bb      = ta.volatility.BollingerBands(c, 20, 2)
        ema13   = ta.trend.EMAIndicator(c, 13).ema_indicator()
        ichi    = ta.trend.IchimokuIndicator(h, l, 9, 26, 52)
        ema5_s  = ta.trend.EMAIndicator(c, 5).ema_indicator()
        ema20_s = ta.trend.EMAIndicator(c, 20).ema_indicator()
        ema60_s = ta.trend.EMAIndicator(c, 60).ema_indicator()
        sma200_s= ta.trend.SMAIndicator(c, 200).sma_indicator()
        adx_obj = ta.trend.ADXIndicator(h, l, c, 14)
        adx_s   = adx_obj.adx()
        macd_obj    = ta.trend.MACD(c)
        macd_hist_s = macd_obj.macd_diff()
        atr_s   = ta.volatility.AverageTrueRange(h, l, c, 14).average_true_range()
        rsi_s   = ta.momentum.RSIIndicator(c, 14).rsi()
        vol_ma20_s  = ta.trend.SMAIndicator(v, 20).sma_indicator() if not v.dropna().empty else pd.Series(dtype=float)

        # 動量: 取當前值與前一期方向比較（TV邏輯）
        stoch_obj    = ta.momentum.StochasticOscillator(h, l, c, 14, 3)
        stoch_k_s    = stoch_obj.stoch()
        stoch_d_s    = stoch_obj.stoch_signal()   # %D，TV用此判斷（crossover邏輯）

        ao_series    = ta.momentum.AwesomeOscillatorIndicator(h, l).awesome_oscillator()
        mom_series   = c - c.shift(10)

        stochrsi_obj = ta.momentum.StochRSIIndicator(c, 14, 3, 3)
        stochrsi_d_s = stochrsi_obj.stochrsi_d() * 100  # %D×100

        willr_s      = ta.momentum.WilliamsRIndicator(h, l, c, 14).williams_r()
        bbpower_s    = c - ema13

        def prev(s, n=1):
            idx = -(n+1)
            return float(s.iloc[idx]) if len(s) >= abs(idx) and pd.notna(s.iloc[idx]) else None

        # 🆕 T3 信心度：5 個指標加總（每命中 1 分）
        def _t3_confidence(close_now, ema5_now, ema20_now,
                           ema5_5d_ago, ema20_5d_ago):
            score = 0
            hits = []
            try:
                # C1: close > EMA20
                if close_now is not None and ema20_now is not None and close_now > ema20_now:
                    score += 1; hits.append('close>EMA20')
                # C3: EMA20 5d 斜率為正
                if ema20_now is not None and ema20_5d_ago is not None and ema20_now > ema20_5d_ago:
                    score += 1; hits.append('EMA20上升')
                # E5: EMA5 5d 斜率為正
                e5_up = (ema5_now is not None and ema5_5d_ago is not None and ema5_now > ema5_5d_ago)
                if e5_up:
                    score += 1; hits.append('EMA5上升')
                # E5>E20: 多頭排列
                if ema5_now is not None and ema20_now is not None and ema5_now > ema20_now:
                    score += 1; hits.append('EMA5>EMA20')
                # 雙均線都升
                e20_up = (ema20_now is not None and ema20_5d_ago is not None
                          and ema20_now > ema20_5d_ago)
                if e5_up and e20_up:
                    score += 1; hits.append('雙均線都升')
            except Exception:
                pass
            return score, hits

        # 🆕 T3 拉回天數：RSI 連續低於 50 的天數
        def _t3_pullback_days(rsi_series):
            try:
                arr = rsi_series.dropna().values
                if len(arr) == 0 or arr[-1] >= 50: return 0
                # 倒數，看連續 < 50 多少天
                cnt = 0
                for v in reversed(arr):
                    if v < 50: cnt += 1
                    else: break
                return cnt if cnt > 0 else 0
            except Exception:
                return 0

        # 🆕 T4 反彈天數：RSI < 32 且連續上升的天數
        def _t4_rising_days(rsi_series):
            try:
                arr = rsi_series.dropna().values
                if len(arr) < 3 or arr[-1] >= 32: return 0
                # 從最後一天往前數連續上升
                cnt = 1  # 包含當天
                for i in range(len(arr) - 1, 0, -1):
                    if arr[i] > arr[i-1] and arr[i] < 32:
                        cnt += 1
                    else:
                        break
                return cnt
            except Exception:
                return 0

        close_val = last(c)
        prev_close_val = prev(c)
        change_pct = ((close_val - prev_close_val) / prev_close_val * 100
                      if close_val and prev_close_val and prev_close_val != 0 else None)
        change_amt = (close_val - prev_close_val
                      if close_val is not None and prev_close_val is not None else None)

        # EMA20/60 黃金/死亡交叉距今天數（正=黃金交叉，負=死亡交叉）
        ema20_cross_days = None
        try:
            diff_s = (ema20_s - ema60_s).dropna()
            for _k in range(1, min(len(diff_s) - 1, 120)):
                d1, d0 = diff_s.iloc[-_k], diff_s.iloc[-_k - 1]
                if pd.notna(d1) and pd.notna(d0):
                    if d0 < 0 and d1 >= 0:
                        ema20_cross_days = _k; break
                    elif d0 > 0 and d1 <= 0:
                        ema20_cross_days = -_k; break
        except Exception:
            pass

        # 🆕 v9.11：T1 觸發至今的漲跌（cross day → today）
        cross_day_close = None
        cross_change_pct = None
        if ema20_cross_days is not None:
            try:
                _k = abs(ema20_cross_days)
                if _k < len(c):
                    cross_day_close = float(c.iloc[-_k - 1])  # cross 發生那天的收盤
                    if cross_day_close and cross_day_close > 0 and close_val is not None:
                        cross_change_pct = (close_val - cross_day_close) / cross_day_close * 100
            except Exception:
                pass

        # 週線結構（日線 resample，無需額外 API）
        w_close_v = w_ma10_v = w_ma20_v = None
        try:
            c_tz = c.copy()
            if hasattr(c_tz.index, 'tz') and c_tz.index.tz is not None:
                c_tz.index = c_tz.index.tz_localize(None)
            wc = c_tz.resample('W').last().dropna()
            if len(wc) >= 20:
                w_close_v = float(wc.iloc[-1])
                w_ma10_v  = last(ta.trend.SMAIndicator(wc, 10).sma_indicator())
                w_ma20_v  = last(ta.trend.SMAIndicator(wc, 20).sma_indicator())
        except Exception:
            pass

        d = {
            "name":         name,
            "close":        close_val,
            "prev_close":   prev_close_val,
            "change_pct":   change_pct,
            "change_amt":   change_amt,
            "rsi":          last(rsi_s),
            "rsi_prev":     prev(rsi_s),
            "rsi_prev2":    prev(rsi_s, 2),             # T4連續2天上升判斷用
            # 🆕 T3/T4 天數計算
            "t3_pullback_days": _t3_pullback_days(rsi_s),  # RSI<50 連續多少天
            "t4_rising_days":   _t4_rising_days(rsi_s),    # RSI<32 且上升多少天
            # 🆕 EMA5 + T3 信心度（v9.9t）
            "ema5":         last(ema5_s),
            "ema5_5d_ago":  prev(ema5_s, 5),
            "ema20_5d_ago": prev(ema20_s, 5),
            "atr14":        last(atr_s),
            "stoch_k":      last(stoch_k_s),
            "stoch_d":      last(stoch_d_s),
            "stoch_d_prev": prev(stoch_d_s),        # crossover判斷用
            "cci":          last(ta.trend.CCIIndicator(h, l, c, 20).cci()),
            "adx":          last(adx_s),
            "adx_prev":     prev(adx_s),
            "adx_pos":      last(adx_obj.adx_pos()),
            "adx_neg":      last(adx_obj.adx_neg()),
            "macd_hist":      last(macd_hist_s),
            "macd_hist_prev": prev(macd_hist_s),
            "volume":       last(v),
            "vol_ma20":     last(vol_ma20_s) if not vol_ma20_s.empty else None,
            "ao":           last(ao_series),
            "ao_prev":      prev(ao_series),
            "ao_prev2":     prev(ao_series, 2),
            "mom":          last(mom_series),
            "mom_prev":     prev(mom_series),
            "macd":         last(macd_obj.macd()),
            "stochrsi":     last(stochrsi_d_s),
            "stochrsi_prev":prev(stochrsi_d_s),     # crossover判斷用
            "willr":        last(willr_s),
            "willr_prev":   prev(willr_s),           # crossover判斷用
            "bbpower":      last(bbpower_s),
            "bbpower_prev": prev(bbpower_s),
            "uo":           last(ta.momentum.UltimateOscillator(h, l, c, 7, 14, 28).ultimate_oscillator()),
            "bbu":      last(bb.bollinger_hband()),
            "bbl":      last(bb.bollinger_lband()),
            # 🆕 v9.12：BB 完整狀態
            "bb_sma":   last(bb.bollinger_mavg()),
            "bb_pct_b": last(bb.bollinger_pband()),
            "bb_bandwidth": last(bb.bollinger_wband()),
            # 🆕 BB 歷史（120d bandwidth 算 squeeze percentile）
            "bb_squeeze_pct": (
                lambda bw_s: float(((bw_s.dropna().tail(120) <= bw_s.iloc[-1]).mean() * 100))
                              if pd.notna(bw_s.iloc[-1]) and len(bw_s.dropna()) >= 60 else None
            )(bb.bollinger_wband()),
            "ema10":    last(ta.trend.EMAIndicator(c, 10).ema_indicator()),
            "sma10":    last(ta.trend.SMAIndicator(c, 10).sma_indicator()),
            "ema20":      last(ema20_s),
            "ema20_prev": prev(ema20_s),
            "sma20":    last(ta.trend.SMAIndicator(c, 20).sma_indicator()),
            "ema30":    last(ta.trend.EMAIndicator(c, 30).ema_indicator()),
            "sma30":    last(ta.trend.SMAIndicator(c, 30).sma_indicator()),
            "ema50":    last(ta.trend.EMAIndicator(c, 50).ema_indicator()),
            "sma50":    last(ta.trend.SMAIndicator(c, 50).sma_indicator()),
            "ema60":      last(ema60_s),
            "ema60_prev": prev(ema60_s),
            "ema20_cross_days": ema20_cross_days,
            "cross_day_close":  cross_day_close,    # 🆕 v9.11：cross 那天的 close
            "cross_change_pct": cross_change_pct,   # 🆕 v9.11：cross 至今 % 變化
            "w_close":  w_close_v,
            "w_ma10":   w_ma10_v,
            "w_ma20":   w_ma20_v,
            "w_dev":    ((w_close_v - w_ma10_v) / w_ma10_v * 100
                         if w_close_v and w_ma10_v and w_ma10_v != 0 else None),
            "sma60":    last(ta.trend.SMAIndicator(c, 60).sma_indicator()),
            "ema100":   last(ta.trend.EMAIndicator(c, 100).ema_indicator()),
            "sma100":   last(ta.trend.SMAIndicator(c, 100).sma_indicator()),
            "ema200":      last(ta.trend.EMAIndicator(c, 200).ema_indicator()),
            "sma200":      last(sma200_s),
            "sma200_prev": prev(sma200_s),
            "ichimoku": last(ichi.ichimoku_base_line()),
            "vwma":     last(vwma(c, v, 20)),
            "hma":      last(hull_ma(c, 9)),
            # 60 日高點（接刀風險警告用）
            "high60":   float(h.iloc[-60:].max()) if len(h) >= 60 else None,
            # 🆕 v9.10x：60 日低點（底部反彈訊號用）
            "low60":    float(l.iloc[-60:].min()) if len(l) >= 60 else None,
        }

        # K 線型態偵測（最近 5 個交易日內出現的型態）
        try:
            from kline_patterns import detect_recent
            o = get_col("open")
            atr_s = ta.volatility.AverageTrueRange(h, l, c, 14).average_true_range()
            kdf = pd.DataFrame({
                'Open': o, 'High': h, 'Low': l, 'Close': c, 'atr': atr_s
            })
            kdf.index = c.index
            d['kline_patterns'] = detect_recent(kdf, lookback=5)
        except Exception:
            d['kline_patterns'] = []

        # 🆕 VWAP 今日值（盤中執行優化建議）— 僅台股
        try:
            if is_tw:
                from fugle_connector import get_minute_candles
                from vwap_loader import compute_daily_vwap
                # 抓近 1 天 5m 資料（即時值）
                tk = ticker.replace('.TW', '').replace('.TWO', '')
                m_df = get_minute_candles(tk, freq='5m', use_cache=False)
                if m_df is not None and not m_df.empty:
                    # 只取最後一天
                    last_date = m_df.index.normalize().max()
                    last_day = m_df[m_df.index >= last_date - pd.Timedelta(hours=8)]
                    if len(last_day) >= 3:
                        vw_df = compute_daily_vwap(last_day)
                        if vw_df is not None and not vw_df.empty:
                            d['vwap_today'] = float(vw_df['VWAP'].iloc[-1])
        except Exception:
            pass

        # 🆕 EPS / PER / PBR / 殖利率（優先 FinMind per_cache，否則 yfinance fallback）
        try:
            from pathlib import Path as _P
            tk = ticker.replace('.TW', '').replace('.TWO', '').replace('.', '')
            pe_path = _P(__file__).parent / 'per_cache' / f'{tk}.parquet'
            got_from_cache = False
            if is_tw and pe_path.exists():
                pe_df = pd.read_parquet(pe_path)
                if not pe_df.empty:
                    last_row = pe_df.iloc[-1]
                    per_v = last_row.get('PER')
                    pbr_v = last_row.get('PBR')
                    div_v = last_row.get('dividend_yield')
                    # FinMind PER=0 表示虧損（EPS ≤ 0）
                    if per_v is not None and not pd.isna(per_v):
                        if per_v > 0:
                            d['per'] = float(per_v); got_from_cache = True
                            d['per_kind'] = 'TTM'
                        else:
                            # PER=0 → 虧損
                            d['per_kind'] = 'LOSS'
                            got_from_cache = True   # 不再 fallback yfinance
                    if pbr_v and not pd.isna(pbr_v):
                        d['pbr'] = float(pbr_v)
                    if div_v and not pd.isna(div_v):
                        d['div_yield'] = float(div_v)
                    if per_v and not pd.isna(per_v) and per_v > 0:
                        close_v = d.get('close')
                        if close_v:
                            d['eps_ttm'] = round(close_v / per_v, 2)
                    # PER 60 日動量
                    if len(pe_df) >= 60:
                        pe_60d = pe_df['PER'].iloc[-60]
                        if pe_60d and per_v and not pd.isna(pe_60d) and not pd.isna(per_v) and pe_60d > 0:
                            d['per_60d_chg_pct'] = round((per_v - pe_60d) / pe_60d * 100, 1)
                    # PER 相對 90 日中位
                    if len(pe_df) >= 90:
                        pe_med90_arr = pe_df['PER'].iloc[-90:].dropna()
                        if len(pe_med90_arr) > 0:
                            pe_med90 = float(pe_med90_arr.median())
                            if pe_med90 and per_v:
                                d['per_vs_med90'] = round((per_v - pe_med90) / pe_med90 * 100, 1)

            # 🆕 券資比（margin_cache）— 僅台股
            if is_tw:
                ms_path = _P(__file__).parent / 'margin_cache' / f'{tk}.parquet'
                if ms_path.exists():
                    ms_df = pd.read_parquet(ms_path)
                    if not ms_df.empty:
                        last = ms_df.iloc[-1]
                        ms_now = last.get('msratio')
                        if ms_now is not None and not pd.isna(ms_now):
                            d['msratio'] = round(float(ms_now), 2)
                        if 'margin_balance' in ms_df.columns:
                            mb = last.get('margin_balance')
                            if mb is not None and not pd.isna(mb):
                                d['margin_balance'] = int(mb)
                        if 'short_balance' in ms_df.columns:
                            sb = last.get('short_balance')
                            if sb is not None and not pd.isna(sb):
                                d['short_balance'] = int(sb)
                        # 60 日動量
                        if len(ms_df) >= 60:
                            ms_60d = ms_df['msratio'].iloc[-60]
                            if ms_60d and ms_now and not pd.isna(ms_60d) and not pd.isna(ms_now) and ms_60d > 0:
                                d['msratio_60d_chg_pct'] = round((ms_now - ms_60d) / ms_60d * 100, 1)

            # 🆕 yfinance fallback（雲端 / 美股 / per_cache 缺檔時）
            if not got_from_cache:
                try:
                    info = yf.Ticker(symbol).info
                    # 優先 trailingPE，無則 forwardPE（標 F 區分）
                    per_yf = info.get('trailingPE')
                    if per_yf and not pd.isna(per_yf) and 0 < per_yf < 1000:
                        d['per'] = float(per_yf)
                        d['per_kind'] = 'TTM'
                    else:
                        fwd_pe = info.get('forwardPE')
                        if fwd_pe and not pd.isna(fwd_pe) and 0 < fwd_pe < 1000:
                            d['per'] = float(fwd_pe)
                            d['per_kind'] = 'FWD'
                        else:
                            # 真正虧損 → 標示 EPS<0
                            eps_t = info.get('trailingEps')
                            if eps_t is not None and not pd.isna(eps_t) and eps_t < 0:
                                d['per_kind'] = 'LOSS'
                    eps_yf = info.get('trailingEps')
                    if eps_yf is not None and not pd.isna(eps_yf):
                        d['eps_ttm'] = round(float(eps_yf), 2)
                    pbr_yf = info.get('priceToBook')
                    if pbr_yf and not pd.isna(pbr_yf) and 0 < pbr_yf < 100:
                        d['pbr'] = round(float(pbr_yf), 2)
                    div_yf = info.get('trailingAnnualDividendYield') or info.get('dividendYield')
                    if div_yf and not pd.isna(div_yf):
                        dv = float(div_yf)
                        if dv < 1.0:
                            d['div_yield'] = round(dv * 100, 2)
                        else:
                            d['div_yield'] = round(dv, 2)
                except Exception:
                    pass
        except Exception:
            pass

        # 🆕 v9.9t：T3 信心度計算
        try:
            score, hits = _t3_confidence(
                d.get('close'), d.get('ema5'), d.get('ema20'),
                d.get('ema5_5d_ago'), d.get('ema20_5d_ago'))
            d['t3_confidence'] = score
            d['t3_confidence_hits'] = hits
        except Exception:
            d['t3_confidence'] = 0
            d['t3_confidence_hits'] = []

        # 🆕 v9.17.1：把波段診斷需要的「近 30 日 OHLC + 指標」存進 d
        # 🆕 v9.20：擴充到 252 日，並加 SEPA / RS 需要的 SMA + 52w 指標
        try:
            import numpy as _np
            tail = 252  # 一年 → SMA200 + 52w 高低足夠
            def _np_tail(s, n=tail):
                arr = s.values if hasattr(s, 'values') else _np.asarray(s)
                if len(arr) >= n:
                    return arr[-n:].tolist()
                return arr.tolist()
            o_s = get_col("open")
            ema10_s = ta.trend.EMAIndicator(c, 10).ema_indicator()
            sma50_s_ = ta.trend.SMAIndicator(c, 50).sma_indicator()
            sma150_s_ = ta.trend.SMAIndicator(c, 150).sma_indicator()
            sma200_s_ = ta.trend.SMAIndicator(c, 200).sma_indicator()
            # 🆕 v9.23: 加 dates 給 detail card ZigZag chart 用
            try:
                _idx_arr = df.index if hasattr(df, 'index') else None
                _dates_tail = ([str(x)[:10] for x in _idx_arr[-tail:]]
                                if _idx_arr is not None and len(_idx_arr) >= tail
                                else [str(x)[:10] for x in (_idx_arr or [])])
            except Exception:
                _dates_tail = []
            d['_swing_history'] = {
                'dates':  _dates_tail,
                'open':   _np_tail(o_s),
                'high':   _np_tail(h),
                'low':    _np_tail(l),
                'close':  _np_tail(c),
                'volume': _np_tail(v),
                'ema10':  _np_tail(ema10_s),
                'ema20':  _np_tail(ema20_s),
                'ema60':  _np_tail(ema60_s),
                'sma50':  _np_tail(sma50_s_),
                'sma150': _np_tail(sma150_s_),
                'sma200': _np_tail(sma200_s_),
                'adx':    _np_tail(adx_s),
                'atr':    _np_tail(atr_s),
                'rsi':    _np_tail(rsi_s),
            }

            # 🆕 v9.20：SEPA 指標直接算入 d
            try:
                from sepa_vcp import (compute_sma_helpers, compute_returns,
                                       check_sepa_trend_template, detect_vcp)
                _sma_helpers_local = compute_sma_helpers(pd.DataFrame({
                    'Close': c, 'High': h, 'Low': l
                }))
                _ret_local = compute_returns(pd.DataFrame({'Close': c}))
                d['sma150'] = _sma_helpers_local.get('sma150')
                d['sma200_real'] = _sma_helpers_local.get('sma200')
                d['high_52w'] = _sma_helpers_local.get('high_52w')
                d['low_52w'] = _sma_helpers_local.get('low_52w')
                d['from_52w_low'] = _sma_helpers_local.get('from_52w_low', 0)
                d['from_52w_high_pct'] = _sma_helpers_local.get('from_52w_high', 0)
                d['returns_13w'] = _ret_local.get('13w', 0)
                d['returns_26w'] = _ret_local.get('26w', 0)
                d['returns_39w'] = _ret_local.get('39w', 0)
                d['returns_52w'] = _ret_local.get('52w', 0)

                # SEPA Trend Template 判斷
                _sepa_pass, _sepa_n, _sepa_det = check_sepa_trend_template(
                    d.get('close'),
                    d.get('sma50'),
                    d.get('sma150'),
                    d.get('sma200_real'),
                    _sma_helpers_local.get('sma200_30d_ago'),
                    d.get('high_52w'),
                    d.get('low_52w'))
                d['sepa_passed'] = _sepa_pass
                d['sepa_n_met'] = _sepa_n
                d['sepa_details'] = _sepa_det

                # VCP 檢測
                _vcp_info = detect_vcp(pd.DataFrame({
                    'Open': o_s.values, 'High': h, 'Low': l,
                    'Close': c, 'Volume': v
                }))
                d['vcp_info'] = _vcp_info

                # 🆕 v9.23.3：雙底雙頂 + ZigZag VCP（統一 ATR×1.30 引擎）
                try:
                    from double_pattern import (detect_double_bottom, detect_double_top,
                                                  detect_vcp_zigzag)
                    _df_for_dbl = pd.DataFrame({
                        'Open': o_s.values, 'High': h, 'Low': l,
                        'Close': c, 'Volume': v
                    })
                    d['double_bottom_info'] = detect_double_bottom(_df_for_dbl)
                    d['double_top_info'] = detect_double_top(_df_for_dbl)
                    d['vcp_zigzag_info'] = detect_vcp_zigzag(_df_for_dbl)
                except Exception:
                    d['double_bottom_info'] = {'is_double_bottom': False, 'status': 'none'}
                    d['double_top_info'] = {'is_double_top': False, 'status': 'none'}
                    d['vcp_zigzag_info'] = {'is_vcp': False}

                # 🆕 v9.20.5：RS Rating — 優先讀 rs_ratings dict（所有 ticker 都有）
                # fallback：從 by_filter 結果反查（舊 JSON 沒 rs_ratings）
                try:
                    from pathlib import Path as _P
                    _sj = _P(__file__).parent / 'screener_results.json'
                    if _sj.exists():
                        _sd = json.load(open(_sj, encoding='utf-8'))
                        _tk_pure = ticker.replace('.TW', '')
                        # 路徑 A：直接從 rs_ratings 查（v9.20.5+）
                        _rs_dict = _sd.get('rs_ratings') or {}
                        if _rs_dict and _tk_pure in _rs_dict:
                            d['rs_rating'] = _rs_dict[_tk_pure]
                        else:
                            # 路徑 B：從 by_filter 反查（fallback）
                            _bf = _sd.get('by_filter', {})
                            for _items in _bf.values():
                                for _r in _items:
                                    if _r.get('ticker') == _tk_pure:
                                        if _r.get('rs_rating') is not None:
                                            d['rs_rating'] = _r.get('rs_rating')
                                            break
                                if d.get('rs_rating') is not None: break
                except Exception:
                    pass
            except Exception:
                pass
        except Exception:
            d['_swing_history'] = None

        return d
    except Exception as e:
        return {"_error": str(e)[:120]}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_indicators_range(ticker: str, market: str, start_date: str, end_date: str):
    """下載一次資料，計算全部 TA 序列後，回傳 [start_date, end_date] 內每個交易日的指標字典。
    返回：list of (date_str, d_dict)，依日期升冪排列。
    """
    import datetime as _dt
    symbol = get_yf_symbol(ticker)
    is_tw  = symbol.endswith(".TW") or symbol.endswith(".TWO")

    # 往前多拉 400 天讓 SMA200 等長周期指標穩定
    _end_dt   = _dt.date.fromisoformat(end_date)   + _dt.timedelta(days=1)
    _start_dt = _dt.date.fromisoformat(start_date) - _dt.timedelta(days=400)
    _end_str, _start_str = _end_dt.isoformat(), _start_dt.isoformat()

    df = None
    for attempt in range(3):
        try:
            if is_tw:
                raw = yf.download(symbol, start=_start_str, end=_end_str,
                                  interval="1d", progress=False,
                                  auto_adjust=False, multi_level_index=False)
                if raw is None or len(raw) < 30:
                    alt  = ticker + ".TWO"
                    raw2 = yf.download(alt, start=_start_str, end=_end_str,
                                       interval="1d", progress=False,
                                       auto_adjust=False, multi_level_index=False)
                    if raw2 is not None and len(raw2) > (len(raw) if raw is not None else 0):
                        raw, symbol = raw2, alt
                df = raw if raw is not None and len(raw) >= 30 else None
            else:
                df = yf.Ticker(symbol).history(start=_start_str, end=_end_str, interval="1d")
                if df is not None and len(df) < 30:
                    df = None
            if df is not None and len(df) >= 30:
                break
        except Exception:
            df = None
        if df is None and attempt < 2:
            time.sleep(3 + attempt * 3)

    if df is None or len(df) < 30:
        return []

    try:
        if hasattr(df.columns, 'levels'):
            df.columns = [str(col[0]).strip().capitalize()
                          if isinstance(col, tuple) else str(col).strip().capitalize()
                          for col in df.columns]
        else:
            df.columns = [str(col).strip().capitalize() for col in df.columns]
        col_map = {cn.lower(): cn for cn in df.columns}
        def _gc(n):
            return df[col_map[n]] if n in col_map else pd.Series(dtype=float, index=df.index)
        c  = _gc("close"); h = _gc("high"); l = _gc("low"); v = _gc("volume")
        if c.dropna().empty:
            return []

        stock_name = _get_stock_name(ticker, symbol)

        # ── 計算所有 TA 序列（一次性） ──────────────────────────────
        bb       = ta.volatility.BollingerBands(c, 20, 2)
        ema13_s  = ta.trend.EMAIndicator(c, 13).ema_indicator()
        ema10_s  = ta.trend.EMAIndicator(c, 10).ema_indicator()
        sma10_s  = ta.trend.SMAIndicator(c, 10).sma_indicator()
        ema20_s  = ta.trend.EMAIndicator(c, 20).ema_indicator()
        sma20_s  = ta.trend.SMAIndicator(c, 20).sma_indicator()
        ema30_s  = ta.trend.EMAIndicator(c, 30).ema_indicator()
        sma30_s  = ta.trend.SMAIndicator(c, 30).sma_indicator()
        ema50_s  = ta.trend.EMAIndicator(c, 50).ema_indicator()
        sma50_s  = ta.trend.SMAIndicator(c, 50).sma_indicator()
        ema60_s  = ta.trend.EMAIndicator(c, 60).ema_indicator()
        sma60_s  = ta.trend.SMAIndicator(c, 60).sma_indicator()
        ema100_s = ta.trend.EMAIndicator(c, 100).ema_indicator()
        sma100_s = ta.trend.SMAIndicator(c, 100).sma_indicator()
        ema200_s = ta.trend.EMAIndicator(c, 200).ema_indicator()
        sma200_s = ta.trend.SMAIndicator(c, 200).sma_indicator()
        adx_obj  = ta.trend.ADXIndicator(h, l, c, 14)
        adx_s    = adx_obj.adx()
        adxp_s   = adx_obj.adx_pos()
        adxn_s   = adx_obj.adx_neg()
        macd_obj     = ta.trend.MACD(c)
        macd_s       = macd_obj.macd()
        macd_hist_s  = macd_obj.macd_diff()
        rsi_s    = ta.momentum.RSIIndicator(c, 14).rsi()
        stoch_obj= ta.momentum.StochasticOscillator(h, l, c, 14, 3)
        stochk_s = stoch_obj.stoch()
        stochd_s = stoch_obj.stoch_signal()
        ao_s     = ta.momentum.AwesomeOscillatorIndicator(h, l).awesome_oscillator()
        mom_s    = c - c.shift(10)
        srsi_s   = ta.momentum.StochRSIIndicator(c, 14, 3, 3).stochrsi_d() * 100
        willr_s  = ta.momentum.WilliamsRIndicator(h, l, c, 14).williams_r()
        bbp_s    = c - ema13_s
        uo_s     = ta.momentum.UltimateOscillator(h, l, c, 7, 14, 28).ultimate_oscillator()
        cci_s    = ta.trend.CCIIndicator(h, l, c, 20).cci()
        ichi_s   = ta.trend.IchimokuIndicator(h, l, 9, 26, 52).ichimoku_base_line()
        bbu_s    = bb.bollinger_hband()
        bbl_s    = bb.bollinger_lband()
        # 🆕 v9.12
        bb_sma_s = bb.bollinger_mavg()
        bb_pctb_s = bb.bollinger_pband()
        bb_bw_s = bb.bollinger_wband()
        vol_ma20_s = (ta.trend.SMAIndicator(v, 20).sma_indicator()
                      if not v.dropna().empty else pd.Series(dtype=float, index=c.index))
        vwma_s   = (vwma(c, v, 20) if not v.dropna().empty
                    else pd.Series(dtype=float, index=c.index))
        hma_s    = hull_ma(c, 9)
        diff_s   = ema20_s - ema60_s     # EMA20/60 差值，用於交叉偵測

        # 週線序列（日線 resample）
        wc_ser = wm10_ser = wm20_ser = None
        try:
            c_tz = c.copy()
            if hasattr(c_tz.index, 'tz') and c_tz.index.tz is not None:
                c_tz.index = c_tz.index.tz_localize(None)
            wc_raw = c_tz.resample('W').last().dropna()
            if len(wc_raw) >= 20:
                wc_ser   = wc_raw
                wm10_ser = ta.trend.SMAIndicator(wc_raw, 10).sma_indicator()
                wm20_ser = ta.trend.SMAIndicator(wc_raw, 20).sma_indicator()
        except Exception:
            pass

        # 取第 i 行純量
        def at(s, i):
            if s is None or i < 0 or i >= len(s): return None
            val = s.iloc[i]
            return float(val) if pd.notna(val) else None

        start_dt = _dt.date.fromisoformat(start_date)
        end_dt_  = _dt.date.fromisoformat(end_date)

        out = []
        for idx in range(len(c)):
            ts_idx = c.index[idx]
            row_date = ts_idx.date() if hasattr(ts_idx, 'date') else pd.Timestamp(ts_idx).date()
            if row_date < start_dt or row_date > end_dt_:
                continue
            if idx < 2:      # ao_prev2 需要 idx-2
                continue
            close_v = at(c, idx)
            if close_v is None:
                continue

            prev_close = at(c, idx-1)
            change_pct = ((close_v - prev_close) / prev_close * 100
                          if close_v and prev_close and prev_close != 0 else None)
            change_amt = (close_v - prev_close
                          if close_v is not None and prev_close is not None else None)

            # EMA20/60 交叉距今天數
            cross_days = None
            for _k in range(1, min(idx, 120)):
                d1 = at(diff_s, idx - _k + 1)
                d0 = at(diff_s, idx - _k)
                if d1 is not None and d0 is not None:
                    if d0 < 0 and d1 >= 0:
                        cross_days = _k;  break
                    elif d0 > 0 and d1 <= 0:
                        cross_days = -_k; break

            # 🆕 v9.11：T1 觸發至今的漲跌（cross day → 此 idx）
            cross_day_close = None
            cross_change_pct = None
            if cross_days is not None:
                _ki = abs(cross_days)
                if idx - _ki >= 0:
                    cross_day_close = at(c, idx - _ki)
                    if cross_day_close and cross_day_close > 0 and close_v is not None:
                        cross_change_pct = (close_v - cross_day_close) / cross_day_close * 100

            # 週線對應值
            w_close_v = w_ma10_v = w_ma20_v = None
            if wc_ser is not None and wm10_ser is not None:
                try:
                    row_ts_naive = pd.Timestamp(row_date)
                    widx = int(wc_ser.index.searchsorted(row_ts_naive, side='right')) - 1
                    if widx >= 0:
                        w_close_v = float(wc_ser.iloc[widx]) if pd.notna(wc_ser.iloc[widx]) else None
                        w_ma10_v  = at(wm10_ser, widx)
                        w_ma20_v  = at(wm20_ser, widx)
                except Exception:
                    pass

            d_dict = {
                "name":         stock_name,
                "close":        close_v,
                "prev_close":   prev_close,
                "change_pct":   change_pct,
                "change_amt":   change_amt,
                "rsi":          at(rsi_s, idx),
                "stoch_k":      at(stochk_s, idx),
                "stoch_d":      at(stochd_s, idx),
                "stoch_d_prev": at(stochd_s, idx-1),
                "cci":          at(cci_s, idx),
                "adx":          at(adx_s, idx),
                "adx_prev":     at(adx_s, idx-1),
                "adx_pos":      at(adxp_s, idx),
                "adx_neg":      at(adxn_s, idx),
                "macd_hist":    at(macd_hist_s, idx),
                "macd_hist_prev": at(macd_hist_s, idx-1),
                "volume":       at(v, idx),
                "vol_ma20":     at(vol_ma20_s, idx),
                "ao":           at(ao_s, idx),
                "ao_prev":      at(ao_s, idx-1),
                "ao_prev2":     at(ao_s, idx-2),
                "mom":          at(mom_s, idx),
                "mom_prev":     at(mom_s, idx-1),
                "macd":         at(macd_s, idx),
                "stochrsi":     at(srsi_s, idx),
                "stochrsi_prev": at(srsi_s, idx-1),
                "willr":        at(willr_s, idx),
                "willr_prev":   at(willr_s, idx-1),
                "bbpower":      at(bbp_s, idx),
                "bbpower_prev": at(bbp_s, idx-1),
                "uo":           at(uo_s, idx),
                "bbu":          at(bbu_s, idx),
                "bbl":          at(bbl_s, idx),
                # 🆕 v9.12
                "bb_sma":       at(bb_sma_s, idx),
                "bb_pct_b":     at(bb_pctb_s, idx),
                "bb_bandwidth": at(bb_bw_s, idx),
                "bb_squeeze_pct": (
                    lambda: float(((bb_bw_s.iloc[max(0,idx-120):idx] <= bb_bw_s.iloc[idx]).mean() * 100))
                    if idx >= 60 and pd.notna(bb_bw_s.iloc[idx]) else None
                )(),
                "ema10":        at(ema10_s, idx),
                "sma10":        at(sma10_s, idx),
                "ema20":        at(ema20_s, idx),
                "ema20_prev":   at(ema20_s, idx-1),
                "sma20":        at(sma20_s, idx),
                "ema30":        at(ema30_s, idx),
                "sma30":        at(sma30_s, idx),
                "ema50":        at(ema50_s, idx),
                "sma50":        at(sma50_s, idx),
                "ema60":        at(ema60_s, idx),
                "ema60_prev":   at(ema60_s, idx-1),
                "ema20_cross_days": cross_days,
                "cross_day_close":  cross_day_close,    # 🆕 v9.11
                "cross_change_pct": cross_change_pct,   # 🆕 v9.11
                "w_close":      w_close_v,
                "w_ma10":       w_ma10_v,
                "w_ma20":       w_ma20_v,
                "w_dev":        ((w_close_v - w_ma10_v) / w_ma10_v * 100
                                 if w_close_v and w_ma10_v and w_ma10_v != 0 else None),
                "sma60":        at(sma60_s, idx),
                "ema100":       at(ema100_s, idx),
                "sma100":       at(sma100_s, idx),
                "ema200":       at(ema200_s, idx),
                "sma200":       at(sma200_s, idx),
                "sma200_prev":  at(sma200_s, idx-1),
                "ichimoku":     at(ichi_s, idx),
                "vwma":         at(vwma_s, idx),
                "hma":          at(hma_s, idx),
            }
            out.append((row_date.isoformat(), d_dict))

        return out
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────
# 🆕 v9.30：JUDGMENT LOGIC 已搬到 detail_card_render.py（檔案頂部 import）
# 移除函數：_j / _jz / _rec / fmt / _jadx / judge_oscillators / judge_mas
#           calc_summary / judge_trend / judge_position / judge_momentum
#           compute_momentum_grade / judge_aux / _calc_aux_summary / apply_cap
# 移除常數：TREND_W / POSITION_W / MOMENTUM_W / AUX_W
# 保留：_get_proximity_alerts（沒搬，下面繼續）
# ─────────────────────────────────────────────────────────────────




# ─────────────────────────────────────────────────────────────────
# 四群組判斷函數
# ─────────────────────────────────────────────────────────────────











# ─────────────────────────────────────────────────────────────────
# 接近條件預警 — 即使尚未觸發 T1/T3/停損，也預先提示可能即將發生
# ─────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────
# ⑦ 自適應趨勢策略 v7 + v8（2026-04-26 完整版）
#
# 【v7 base 基礎策略】1263 檔均值 +72.79%
# 進場兩觸發：T1 黃金交叉 | T3 多頭拉回 RSI<50
# 共同前提：EMA20>EMA60 + ADX≥22 + EMA120 60日跌幅<2%（防死亡迴圈）
# 出場 ATR/Price 自動分類：高波動 EMA死叉 / 穩健 EMA+RSI+ATR
# 長持鎖定：持倉>200天 + 浮動>50% + EMA120上升 → EMA60/120 慢出場
# 反向ETF：ATR×1.5 + RSI>70 出場（無 T4）
#
# 【v8 金字塔加碼】同股不限倉位（架構性突破）
# 模式：P{N}_{signals}+{filters}
#   P0_T1T3            +197.48% （最大獲利，最差 -290%，σ=12.7）
#   P0_T1T3+CB30       +134.07% （風控版，最差 -166%，風報比 0.81，σ=12.6）
# ★ P0_T1T3+POS        +141.68% （生產推薦，最差 -166%，風報比 0.85，σ=8.9）
#   P0_T1T3+PS         +110.26% （最保守，最差 -149%）
#
# 【POS 自適應規則】（深度分析後發現）
# 規則：累積已實現損益為正才允許加碼（自我驗證機制）
#   - 死亡迴圈股：累積永負 → 永不加碼 → 與 v7 相同
#   - 強趨勢股：累積快速為正 → 正常加碼 → 完整受益
#   - 震盪股：累積零附近 → 加碼受限 → 避免惡化
# 跨年度 σ=8.9（vs P0/CB30 σ=12.7），全部變體最穩定
#
# 【健壯性驗證】
# - Walk-forward EARLY (2020-2022) +54.6 / LATE (2023-2026) +54.4 → Edge 穩定
# - 扣 0.4275% 台股交易成本後 P0_T1T3 仍 +184.33%
# - Monte Carlo：EMA120/ATR 健壯、ADX 中度敏感、RSI 脆弱
# - Cohen's d：v7<0 → 退步機率 10×（POS 自動處理）
#
# 【實證測試結論】
# 進場過濾普遍誤殺強勢股；退場優化大多切碎飆股
# 加碼自適應控制（POS）才是改善風報比的關鍵
# ─────────────────────────────────────────────────────────────────

# 反向ETF：使用 ⑦反向ETF策略（T1/T3 based on own chart，無T4，ATR×1.5，RSI>70出場）
# 核心邏輯：反向ETF的EMA黃金交叉 = 大盤開始下跌 → 此時正是進場時機
# 回測驗證（2020~2026）：RSI>70出場（-9.49%）優於ADX<25→RSI>65（-19.57%）


# 一般特殊標的警告（非反向ETF，但操作需特別注意）



# ─────────────────────────────────────────────────────────────────
# INPUT PARSING  ← FIX: 修正 NameError（ticker 未定義的問題）
# ─────────────────────────────────────────────────────────────────
def parse_input(text: str) -> list:
    stocks = []
    seen   = set()
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts  = [p.strip() for p in line.split(",")]
        ticker = parts[0].upper()
        if ticker in seen:
            continue
        seen.add(ticker)
        if is_tw_stock(ticker):
            stocks.append((ticker, "台股"))
        elif ticker in SYMBOL_ALIASES:
            stocks.append((ticker, "指數"))
        else:
            # 🆕 v9.13 fix：移除 NASDAQ 預設（讓 TradingView 自動偵測交易所）
            # 原本 default 'NASDAQ' 對 NYSE/AMEX 股票（如 PL/JPM）會產生不存在的 URL
            market = parts[1].upper() if len(parts) > 1 else "US"
            stocks.append((ticker, market))
    return stocks

# ─────────────────────────────────────────────────────────────────
# HTML HELPERS
# ─────────────────────────────────────────────────────────────────

def jcell(val, judg):
    cls = {"買入":"j-buy","賣出":"j-sell"}.get(judg, "j-neutral")
    return f'<td class="{cls}">{val}</td>'

def classify_action(d: dict) -> str:
    """
    🆕 v8 操作分類（每檔股票歸類為四種狀態之一）
    回傳：'ENTRY' / 'EXIT' / 'HOLD' / 'WAIT'

    ENTRY (可進場): 多頭 + ADX≥22 + (T1 黃金交叉 ≤ 10 天 OR T3 RSI<50 拉回) OR T4 反彈條件達成
    EXIT  (應出倉): 多頭中但出現出場訊號（高 RSI、EMA 死叉迫近、深度乖離等）
    HOLD  (持倉中): 多頭 + ADX≥22 但無新進場訊號（安全持倉或注意觀察）
    WAIT  (觀望中): 空頭 / 假多頭 / 資料不足 / EMA 死叉等
    """
    ema20      = d.get("ema20")
    ema60      = d.get("ema60")
    adx        = d.get("adx")
    rsi        = d.get("rsi")
    rsi_prev   = d.get("rsi_prev")
    rsi_prev2  = d.get("rsi_prev2")
    atr14      = d.get("atr14")
    close      = d.get("close")
    cross_days = d.get("ema20_cross_days")

    if ema20 is None or ema60 is None:
        return 'WAIT'

    is_bull = ema20 > ema60
    adx_ok  = (adx is not None and adx >= 22)

    # 空頭：唯一進場 = T4 反彈
    if not is_bull:
        t4_rising = (rsi is not None and rsi < 32 and
                     rsi_prev is not None and rsi > rsi_prev and
                     rsi_prev2 is not None and rsi_prev > rsi_prev2)
        if t4_rising:
            return 'ENTRY'  # T4 條件達成
        return 'WAIT'

    # 多頭但 ADX 不足
    if not adx_ok:
        return 'WAIT'

    # 多頭 + ADX≥22：判斷進場 / 出倉 / 持倉
    # 出場訊號優先（保守起見）
    rel_atr = (atr14 / close * 100) if (atr14 and close and close > 0) else 0
    is_high_vol = rel_atr > 3.5
    ema_gap_pct = (ema20 - ema60) / ema60 * 100 if ema60 else None
    ema_danger  = ema_gap_pct is not None and ema_gap_pct < 1.0   # EMA 即將死叉

    if is_high_vol:
        # 飆股模式：只看 EMA 距離
        if ema_danger:
            return 'EXIT'
    else:
        # 穩健股：EMA 死叉迫近 OR (ADX<25 + RSI>75)
        rsi_triggered = (adx is not None and adx < 25 and rsi is not None and rsi > 75)
        if rsi_triggered or ema_danger:
            return 'EXIT'

    # 進場觸發
    t1_ok = (cross_days is not None and 0 < cross_days <= 10)
    t3_ok = (rsi is not None and rsi < 50)
    if t1_ok or t3_ok:
        return 'ENTRY'

    # 否則為持倉中（安全持倉 / 等待 T3 拉回）
    return 'HOLD'


def get_exit_signal(d: dict) -> tuple:
    """
    ④出場獲利 輕量版：回傳 (status_label, badge_inline_style) 供表格使用。
    v3：依 ATR/Price 自動分類出場規則
      高波動（ATR/P > 3.5%）→ 只看EMA死叉，不設RSI出場門檻
      穩健（ATR/P ≤ 3.5%）  → EMA死叉 + ADX<25時RSI>75
    """
    ema20     = d.get("ema20")
    ema60     = d.get("ema60")
    adx       = d.get("adx")
    rsi       = d.get("rsi")
    rsi_prev  = d.get("rsi_prev")
    rsi_prev2 = d.get("rsi_prev2")
    atr14     = d.get("atr14")
    close_v   = d.get("close")

    if ema20 is None or ema60 is None:
        return ("—", "background:#0a1020;color:#555e6a")

    is_bull = ema20 > ema60

    if not is_bull:
        _t4 = (rsi is not None and rsi < 32 and
               rsi_prev is not None and rsi > rsi_prev and
               rsi_prev2 is not None and rsi_prev > rsi_prev2)
        if _t4:
            return ("T4 RSI>55出場", "background:#2a1500;color:#ff9944;border:1px solid #ff994455")
        return ("空頭 — 不持倉", "background:#0a1020;color:#555e6a;border:1px solid #2a334455")

    # 計算 ATR/Price 判斷股性
    rel_atr = (atr14 / close_v * 100) if (atr14 and close_v and close_v > 0) else 0
    is_high_vol = rel_atr > 3.5   # 高波動飆股：只守EMA死叉

    ema_gap   = (ema20 - ema60) / ema60 * 100 if ema60 else None
    ema_danger  = ema_gap is not None and ema_gap < 1.0
    ema_warning = ema_gap is not None and 1.0 <= ema_gap < 3.0

    if is_high_vol:
        # 飆股模式：只看EMA距離，不設RSI出場
        if ema_danger:
            return ("⚠️ 出場訊號", "background:#2a0808;color:#ff5555;border:1px solid #ff555555")
        if ema_warning:
            return ("注意觀察",   "background:#1a1200;color:#e8a020;border:1px solid #e8a02044")
        return ("安全持倉",       "background:#0a1e10;color:#3dbb6a;border:1px solid #3dbb6a44")
    else:
        # 穩健股模式：EMA死叉 + ADX<25時RSI>75
        rsi_triggered = (adx is not None and adx < 25 and rsi is not None and rsi > 75)
        rsi_warning   = (adx is not None and adx < 25 and rsi is not None and rsi >= 70)
        if rsi_triggered or ema_danger:
            return ("⚠️ 出場訊號", "background:#2a0808;color:#ff5555;border:1px solid #ff555555")
        if rsi_warning or ema_warning:
            return ("注意觀察",   "background:#1a1200;color:#e8a020;border:1px solid #e8a02044")
        return ("安全持倉",       "background:#0a1e10;color:#3dbb6a;border:1px solid #3dbb6a44")






def render_table(results, platform_url_tpl: str = "https://www.perplexity.ai/search?q={prompt}",
                 selected_platform: str = "Perplexity") -> str:
    rows = ""
    for r in results:
        ticker, market, d, error = r[0], r[1], r[2], r[3]
        group_summs = r[5]   # (ts, ps, ms, xs)
        tsumm  = r[6]        # (tb, ts_, tn_, verdict)
        cap    = r[7]
        name = d.get("name", ticker) if d else ticker

        if error or not d:
            tv_url_err = get_tv_url(ticker, market)
            err_concepts = _get_concepts(ticker, max_n=4)
            err_concept_html = "".join(_concept_chip_html(c) for c in err_concepts) \
                       if err_concepts else '<span style="color:#334455;font-size:.7rem">—</span>'
            rows += (f'<tr>'
                     f'<td class="ticker-cell"><a href="{tv_url_err}" target="_blank" style="color:#e8f4fd;text-decoration:none;">{ticker}</a></td>'
                     f'<td style="color:#a8cce8;font-size:.78rem">—</td>'
                     f'<td class="j-na">—</td><td class="j-na">—</td>'
                     f'<td class="j-na">—</td>'
                     f'<td class="j-na">—</td>'  # 🆕 v9.11：T1累計欄
                     f'<td class="j-na">— 無資料 —</td><td class="j-na">—</td>'
                     f'<td>{err_concept_html}</td></tr>')
            continue

        close_val  = d.get("close")
        change_pct = d.get("change_pct")
        change_amt = d.get("change_amt")
        price_str  = fmt(close_val) if close_val is not None else "N/A"
        if change_pct is not None:
            chg_color = "#ff4444" if change_pct >= 0 else "#22cc88"
            chg_sign  = "▲" if change_pct >= 0 else "▼"
            amt_str   = f'{abs(change_amt):.2f}' if change_amt is not None else ""
            chg_str   = f'{amt_str} {chg_sign} {abs(change_pct):.2f}%'
        else:
            chg_color, chg_str = "#8899aa", "N/A"

        price_cell = (f'<td style="font-family:\'IBM Plex Mono\',monospace;font-size:.82rem;color:#e8f4fd">{price_str}</td>')
        chg_cell   = (f'<td style="font-family:\'IBM Plex Mono\',monospace;font-size:.82rem;color:{chg_color};font-weight:600">{chg_str}</td>')

        # 🆕 P/E 顯示（顏色判定 + 60 日動量箭頭 + 虧損標示）
        per_v = d.get('per')
        per_kind = d.get('per_kind', '')
        per_60d_chg = d.get('per_60d_chg_pct')
        if per_v is None:
            if per_kind == 'LOSS':
                pe_cell = ('<td style="color:#ff7777;font-size:.72rem;text-align:center;'
                           'font-weight:700" title="EPS ≤ 0 公司虧損中">虧損</td>')
            else:
                pe_cell = '<td style="color:#334455;font-size:.78rem;text-align:center">—</td>'
        else:
            if per_v <= 0 or per_v > 100:
                pe_color = '#ff5555'
            elif per_v < 10:
                pe_color = '#3dbb6a'
            elif per_v <= 20:
                pe_color = '#3dbb6a'
            elif per_v <= 30:
                pe_color = '#c8b87a'
            elif per_v <= 50:
                pe_color = '#e8a020'
            else:
                pe_color = '#ff5555'
            # 動量箭頭：60 日 PE 顯著降 → 🔻 (盈餘上修)；顯著升 → 🔺
            arrow = ''
            if per_60d_chg is not None:
                if per_60d_chg <= -15:
                    arrow = ' <span style="color:#3dbb6a;font-size:.7rem" title="60d PER 大降，盈餘上修">▼▼</span>'
                elif per_60d_chg <= -5:
                    arrow = ' <span style="color:#3dbb6a;font-size:.7rem" title="60d PER 微降">▼</span>'
                elif per_60d_chg >= 15:
                    arrow = ' <span style="color:#ff5555;font-size:.7rem" title="60d PER 大漲，盈餘下修風險">▲▲</span>'
                elif per_60d_chg >= 5:
                    arrow = ' <span style="color:#e8a020;font-size:.7rem" title="60d PER 微升">▲</span>'
            pe_cell = (f'<td style="font-family:\'IBM Plex Mono\',monospace;font-size:.82rem;'
                       f'color:{pe_color};font-weight:600;text-align:right">'
                       f'{per_v:.1f}{arrow}</td>')

        # 🆕 v9.11：T1 觸發至今累計 % 變化
        _cd = d.get('ema20_cross_days')
        _cp = d.get('cross_change_pct')
        if _cd is not None and _cd > 0 and _cp is not None:
            # 黃金交叉至今
            _ccolor = '#3dbb6a' if _cp >= 0 else '#ff5555'
            _csign = '+' if _cp >= 0 else ''
            cross_cell = (f'<td style="font-family:\'IBM Plex Mono\',monospace;'
                          f'font-size:.8rem;color:{_ccolor};font-weight:600;text-align:right" '
                          f'title="T1 黃金交叉至今 {_cd} 天 累計 {_csign}{_cp:.2f}%">'
                          f'{_csign}{_cp:.1f}%'
                          f'<span style="color:#5a7a99;font-size:.66rem"> /{_cd}d</span>'
                          f'</td>')
        elif _cd is not None and _cd < 0 and _cp is not None:
            # 死亡交叉至今（灰色）
            _csign = '+' if _cp >= 0 else ''
            cross_cell = (f'<td style="font-family:\'IBM Plex Mono\',monospace;'
                          f'font-size:.72rem;color:#7a8899;text-align:right" '
                          f'title="死亡交叉至今 {abs(_cd)} 天 累計 {_csign}{_cp:.2f}%">'
                          f'DC {_csign}{_cp:.1f}%'
                          f'<span style="color:#5a7a99;font-size:.66rem"> /{abs(_cd)}d</span>'
                          f'</td>')
        else:
            cross_cell = '<td style="color:#334455;font-size:.78rem;text-align:center">—</td>'

        _rlabel, _rbadge = get_rec_label(d, ticker)
        _xlabel, _xbadge = get_exit_signal(d)
        # 🆕 v9.9u：T3 信心度只顯示於 T3 相關標籤（T1/飆股/T4/觀望不顯示）
        _is_t3_label = 'T3' in _rlabel
        _conf_score = d.get('t3_confidence', 0) or 0
        _conf_html = (render_confidence_dots(_conf_score)
                      if _is_t3_label and _conf_score > 0 else '')
        tot_cell = (f'<td style="background:#060c18;font-size:.82rem;font-weight:700;line-height:1.6">'
                    f'<span style="display:inline-block;padding:2px 7px;border-radius:4px;'
                    f'font-size:.76rem;white-space:nowrap;{_rbadge}">{_rlabel}</span>'
                    f'{("&nbsp;" + _conf_html) if _conf_html else ""}'
                    f'</td>')
        exit_cell = (f'<td style="background:#060c18;font-size:.82rem;line-height:1.6">'
                     f'<span style="display:inline-block;padding:2px 7px;border-radius:4px;'
                     f'font-size:.76rem;white-space:nowrap;{_xbadge}">{_xlabel}</span></td>')

        tv_url = get_tv_url(ticker, market)
        is_perplexity = "{prompt}" in platform_url_tpl and "perplexity" in platform_url_tpl
        ai_url    = get_ai_url(ticker, name, d, platform_url_tpl)
        prompt_js = get_prompt_text(ticker, name, d).replace("\\n", "\\\\n").replace("'", "\\'")
        if is_perplexity:
            ticker_link = f'<a href="{ai_url}" target="_blank" title="Perplexity 公司資訊查詢（基本資料 + 業務 + 影響股價因素 + 概念股連動）" style="color:#e8f4fd;text-decoration:none;">{ticker}</a>'
        else:
            homepage = platform_url_tpl.split("?")[0].split("{")[0]
            ticker_link = (
                f'<a href="#" onclick="navigator.clipboard.writeText(\'{prompt_js}\').then(()=>{{'
                f'window.open(\'{homepage}\',\'_blank\');'
                f'alert(\'公司資訊查詢提示詞已複製！請在新視窗中貼上(Ctrl+V / Cmd+V)\');}});return false;"'
                f' title="複製公司資訊查詢提示詞並開啟{selected_platform}" style="color:#e8f4fd;text-decoration:none;">{ticker}</a>'
            )
        # 概念股標籤
        concepts = _get_concepts(ticker, max_n=4)
        concept_html = "".join(_concept_chip_html(c) for c in concepts) \
                       if concepts else '<span style="color:#334455;font-size:.7rem">—</span>'
        concept_cell = (f'<td style="max-width:220px;line-height:1.7">{concept_html}</td>')

        # 🆕 VWAPEXEC 適用分級徽章
        vw_tier = get_vwap_tier(ticker)
        tier_badge = ''
        _delta_v = vw_tier.get('delta', 0)
        if vw_tier.get('tier') == 'TOP':
            _title = f"VWAPEXEC TOP 200 — 歷史 Δ {_delta_v:+.0f}%"
            tier_badge = (
                f'<span title="{_title}" '
                f'style="display:inline-block;padding:1px 5px;background:#0a3a1f;color:#3dbb6a;'
                f'border-radius:3px;font-size:.65rem;font-weight:700;margin-right:4px">⭐</span>'
            )
        elif vw_tier.get('tier') == 'NA':
            _title = f"VWAPEXEC 不適用 — Δ {_delta_v:+.0f}%"
            tier_badge = (
                f'<span title="{_title}" '
                f'style="display:inline-block;padding:1px 5px;background:#3a0a0a;color:#ff8888;'
                f'border-radius:3px;font-size:.65rem;font-weight:700;margin-right:4px">⚠️</span>'
            )
        # OK tier 不顯示徽章（保持簡潔）

        rows += (f'<tr>'
                 f'<td class="ticker-cell">{tier_badge}{ticker_link}</td>'
                 f'<td style="color:#a8cce8;font-size:.78rem;white-space:nowrap;max-width:150px;overflow:hidden;text-overflow:ellipsis">'
                 f'<a href="{tv_url}" target="_blank" style="color:#a8cce8;text-decoration:none;">{name}</a></td>'
                 f'{price_cell}{chg_cell}{pe_cell}{cross_cell}{tot_cell}{exit_cell}{concept_cell}</tr>')

    return (f'<div style="background:#060c18;border-radius:12px;border:1px solid #1e3a5f;padding:4px">'
            f'<table class="res-table"><thead><tr>'
            f'<th>代號</th><th>名稱</th><th>現價</th><th>漲跌幅</th>'
            f'<th title="本益比（顏色：綠合理 / 黃合理偏高 / 橘偏高 / 紅虧損或過熱；▼=PER 60日下降=盈餘上修)">P/E</th>'
            f'<th title="T1 黃金交叉至今的累計漲跌（綠=漲 紅=跌；DC=死亡交叉至今）">T1累計</th>'
            f'<th style="background:#060c18;min-width:140px">操作建議</th>'
            f'<th style="background:#060c18;min-width:120px">④出場獲利</th>'
            f'<th style="background:#060c18;min-width:140px">概念股</th>'
            f'</tr></thead><tbody>{rows}</tbody></table></div>')


# ─────────────────────────────────────────────────────────────────
# EXCEL EXPORT  （四群組版）
# ─────────────────────────────────────────────────────────────────
def build_excel(results) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "指標報告"
    ws.sheet_view.showGridLines = False

    def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
    def fnt(c="000000", sz=9, bd=False): return Font(name="Arial", size=sz, bold=bd, color=c)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    mid  = Side(style="medium")
    thin = Side(style="thin")

    # ── 判斷顏色對照 ──────────────────────────────────────────────
    JUDG_FC = {"買入":"1565C0","強力買入":"0D47A1",
               "賣出":"B71C1C","強力賣出":"7B1616",
               "中立":"555555"}
    VERDICT_BG = {
        "強力買入":          "0D3B6E",
        "買入":              "0D2E50",
        "上限買入｜持有/短線": "3A2A00",
        "中立":              "1A2030",
        "賣出":              "3B0D0D",
        "強力賣出":          "4A0A0A",
        "過熱觀望｜禁止新倉": "3A1800",
        "空頭，不買":        "3A0808",
    }
    VERDICT_FC = {
        "強力買入":          "60CFFF",
        "買入":              "60B3FF",
        "上限買入｜持有/短線": "F0C030",
        "中立":              "9AAABB",
        "賣出":              "FF8080",
        "強力賣出":          "FF6B6B",
        "過熱觀望｜禁止新倉": "FF8830",
        "空頭，不買":        "FF5555",
    }

    # ── 欄位定義：(標題, 寬度, 背景hex, 群組key) ─────────────────
    # 群組 key: T=趨勢, P=位置, M=動能, X=輔助, O=總覽, Z=基本
    G_HDR = {"Z":"1E3A5F","T":"0D2040","P":"2D1A00","M":"1A0D30","X":"1A1A28","O":"2C1654"}
    G_DAT = {"Z":("F0F4FF","F8FAFF"),
             "T":("E8F0FB","F3F7FE"), "P":("FFF3E0","FFFBF0"),
             "M":("F3E8FF","FAF3FF"), "X":("ECEFF4","F5F7FA"),
             "O":("EDE7F6","F5F0FF")}

    col_defs = [
        # 基本
        ("代號",    9, "Z"), ("名稱",    18, "Z"),
        ("現價",    10, "Z"), ("漲跌%",  10, "Z"),
        # 趨勢結構 (5項 + 小結)
        ("趨勢方向", 22, "T"), ("趨勢強度", 20, "T"),
        ("多頭階段", 22, "T"), ("乖離風險", 18, "T"), ("週線結構", 20, "T"),
        ("趨勢小結(40%)", 26, "T"),
        # 位置風險 (3項 + 小結)
        ("EMA20乖離", 16, "P"), ("RSI(14)", 14, "P"), ("布林%B", 14, "P"),
        ("位置小結(30%)", 26, "P"),
        # 動能確認 (5項 + 小結)
        ("MACD零軸", 14, "M"), ("MACD柱體", 16, "M"),
        ("動量(10)", 14, "M"), ("量能比率", 14, "M"), ("MA10位置", 14, "M"),
        ("動能小結(20%)", 26, "M"),
        # 輔助指標 (10項 + 小結)
        ("隨機%K",  12, "X"), ("CCI(20)",  12, "X"), ("StochRSI", 12, "X"),
        ("威廉%R",  12, "X"), ("牛熊力度", 12, "X"), ("終極震盪", 12, "X"),
        ("EMA(10)", 12, "X"), ("EMA(60)",  12, "X"), ("SMA(200)", 12, "X"),
        ("Hull MA", 12, "X"),
        ("輔助小結(10%)", 26, "X"),
        # 整體
        ("操作建議", 28, "O"), ("Cap說明",  36, "O"),
    ]

    # ── 寫標題行 ──────────────────────────────────────────────────
    for ci, (label, width, gkey) in enumerate(col_defs, 1):
        c = ws.cell(1, ci, label)
        c.font = fnt("FFFFFF", 9, True)
        c.fill = fill(G_HDR[gkey])
        c.alignment = ctr
        c.border = Border(bottom=mid, right=thin)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "E2"

    def summ_str(b, s, n, v): return f"買:{b} 賣:{s} 中:{n} → {v}"

    def ind_val(item):
        """從 (label, val, judg) 或 (val, judg) 中取出 val 和 judg"""
        if len(item) == 3:
            return item[1], item[2]
        return item[0], item[1]

    for ri, item in enumerate(results, 2):
        ticker, market, d, error = item[0], item[1], item[2], item[3]
        groups      = item[4]   # (g_trend, g_position, g_momentum, g_aux)
        group_summs = item[5]   # (ts, ps, ms, xs)
        tsumm       = item[6]   # (tb, ts_, tn_, verdict)
        cap         = item[7]

        ws.row_dimensions[ri].height = 20
        is_even = (ri % 2 == 0)

        def cell(col, val, gkey, fc="000000", sz=9, bd=False, align=ctr):
            bg = G_DAT[gkey][0] if is_even else G_DAT[gkey][1]
            c = ws.cell(ri, col, val)
            c.font = fnt(fc, sz, bd)
            c.fill = fill(bg)
            c.alignment = align
            c.border = Border(right=thin)
            return c

        def verdict_cell(col, verdict, gkey):
            bg  = VERDICT_BG.get(verdict, "626567")
            fc  = VERDICT_FC.get(verdict, "FFFFFF")
            c = ws.cell(ri, col, verdict)
            c.font = fnt(fc, 9, True)
            c.fill = fill(bg)
            c.alignment = ctr
            c.border = Border(right=Side(style="medium"))
            return c

        # 基本資訊
        name = d.get("name", ticker) if d else ticker
        cell(1, ticker, "Z", "1565C0", 10, True)
        cell(2, name,   "Z", "333333", 9, False, left)

        if error or not d:
            for ci in range(3, len(col_defs)+1):
                cell(ci, "無資料", "Z", "AAAAAA", 9)
            continue

        close_val  = d.get("close")
        change_pct = d.get("change_pct")
        cell(3, f"{close_val:.2f}" if close_val else "N/A", "Z", "222222", 9)
        if change_pct is not None:
            chg_fc = "B71C1C" if change_pct >= 0 else "1B5E20"
            cell(4, f"{'▲' if change_pct>=0 else '▼'}{abs(change_pct):.2f}%", "Z", chg_fc, 9, True)
        else:
            cell(4, "N/A", "Z", "888888", 9)

        g_trend, g_position, g_momentum, g_aux = groups
        ts_s, ps_s, ms_s, xs_s = group_summs
        tb, ts_, tn_, tr_ = tsumm

        # ── 趨勢結構 (5項 + 小結) ──────────────────────────────
        ci = 5
        for it in g_trend[:5]:
            v, j = ind_val(it)
            cell(ci, f"{v} / {j}", "T", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci += 1
        verdict_cell(ci, ts_s[3], "T")
        ws.cell(ri, ci).value = summ_str(*ts_s); ci += 1

        # ── 位置風險 (3項 + 小結) ──────────────────────────────
        for it in g_position[:3]:
            v, j = ind_val(it)
            cell(ci, f"{v} / {j}", "P", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci += 1
        verdict_cell(ci, ps_s[3], "P")
        ws.cell(ri, ci).value = summ_str(*ps_s); ci += 1

        # ── 動能確認 (5項 + 小結) ──────────────────────────────
        for it in g_momentum[:5]:
            v, j = ind_val(it)
            cell(ci, f"{v} / {j}", "M", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci += 1
        verdict_cell(ci, ms_s[3], "M")
        ws.cell(ri, ci).value = summ_str(*ms_s); ci += 1

        # ── 輔助指標 (10項 + 小結) ─────────────────────────────
        aux_start = ci
        for it in g_aux[:10]:
            v, j = ind_val(it)
            cell(ci, f"{v} / {j}", "X", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci += 1
        # pad if fewer than 10 items（與 build_stock_range_excel 保持一致）
        while ci < aux_start + 10:
            cell(ci, "", "X"); ci += 1
        verdict_cell(ci, xs_s[3], "X")
        ws.cell(ri, ci).value = summ_str(*xs_s); ci += 1

        # ── 操作建議 + Cap ──────────────────────────────────────
        verdict_cell(ci, tr_, "O")
        ws.cell(ri, ci).value = tr_; ci += 1
        cap_txt = cap if cap else "—"
        cell(ci, cap_txt, "O", "884400" if cap else "888888", 9, False, left)

    # 時間戳記
    ws.cell(len(results) + 3, 1,
            f"產出時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = fnt("999999", 8)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─── 🆕 v9.10p：PDF 完整指標報告 ──────────────────────────────
def build_pdf(results) -> bytes:
    """產生個股完整指標分析 PDF（圖文並茂）
    包含：封面 + 每股一頁完整指標 + 操作建議"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    import io as _io
    from datetime import datetime as _dt

    # 註冊中文字體（CID 內建，雲端可用）
    try:
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        CN_FONT = 'STSong-Light'
    except Exception:
        CN_FONT = 'Helvetica'

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=1.5*cm, rightMargin=1.5*cm)
    story = []
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle('Title', parent=styles['Title'],
                               fontName=CN_FONT, fontSize=20,
                               textColor=colors.HexColor('#3b9eff'),
                               alignment=1, spaceAfter=10)
    h1_st = ParagraphStyle('H1', parent=styles['Heading1'],
                            fontName=CN_FONT, fontSize=14,
                            textColor=colors.HexColor('#7abadd'),
                            spaceBefore=8, spaceAfter=4)
    h2_st = ParagraphStyle('H2', parent=styles['Heading2'],
                            fontName=CN_FONT, fontSize=11,
                            textColor=colors.HexColor('#3dbb6a'),
                            spaceBefore=6, spaceAfter=3)
    body_st = ParagraphStyle('Body', parent=styles['Normal'],
                              fontName=CN_FONT, fontSize=9,
                              textColor=colors.HexColor('#333333'),
                              leading=13)
    small_st = ParagraphStyle('Small', parent=styles['Normal'],
                               fontName=CN_FONT, fontSize=7,
                               textColor=colors.HexColor('#666666'))

    # ── 封面 ──
    story.append(Paragraph("📊 Stock001 個股指標分析報告", title_st))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"報告生成: {_dt.now().strftime('%Y-%m-%d %H:%M:%S')}", body_st))
    story.append(Paragraph(
        f"涵蓋股票: {len(results)} 支", body_st))
    story.append(Spacer(1, 0.5*cm))

    # 摘要表
    summary_rows = [['#', 'Ticker', '名稱', '收盤', 'RSI', 'ADX',
                      'EMA20/60', '建議']]
    for i, item in enumerate(results, 1):
        try:
            ticker = item[0] if len(item) > 0 else ''
            d = item[2] if len(item) > 2 and item[2] else None
            if d is None:
                summary_rows.append([str(i), ticker, '—', '—', '—', '—', '—', '無資料'])
                continue
            name = (d.get('name') or '')[:12]
            close = d.get('close', 0) or 0
            rsi = d.get('rsi', 0) or 0
            adx = d.get('adx', 0) or 0
            e20 = d.get('ema20', 0) or 0
            e60 = d.get('ema60', 0) or 0
            bull = '多頭' if (e20 > e60) else '空頭'
            try:
                rec_label, _ = get_rec_label(d, ticker)
            except Exception:
                rec_label = '—'
            summary_rows.append([
                str(i), ticker, name, f'{close:.2f}',
                f'{rsi:.1f}', f'{adx:.1f}', bull, rec_label[:14]
            ])
        except Exception:
            continue

    summary_tbl = Table(summary_rows, colWidths=[
        0.8*cm, 2.2*cm, 4*cm, 1.8*cm, 1.5*cm, 1.5*cm, 1.5*cm, 4*cm])
    summary_tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), CN_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0a1830')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f5f8fa')]),
    ]))
    story.append(Paragraph("📋 個股摘要", h1_st))
    story.append(summary_tbl)
    story.append(PageBreak())

    # ── 每股一頁詳細 ──
    for idx, item in enumerate(results, 1):
        try:
            ticker = item[0] if len(item) > 0 else ''
            market = item[1] if len(item) > 1 else 'TW'
            d = item[2] if len(item) > 2 else None
            if d is None: continue

            name = d.get('name', '')
            close = d.get('close', 0) or 0
            rsi = d.get('rsi', 0) or 0
            rsi_prev = d.get('rsi_prev', 0) or 0
            adx = d.get('adx', 0) or 0
            ema20 = d.get('ema20', 0) or 0
            ema60 = d.get('ema60', 0) or 0
            ema5 = d.get('ema5', 0) or 0
            ema120 = d.get('ema120', 0) or 0
            atr14 = d.get('atr14', 0) or 0
            sma200 = d.get('sma200', 0) or 0
            cd = d.get('ema20_cross_days', 0) or 0
            per = d.get('per')
            pbr = d.get('pbr')
            div = d.get('div_yield')
            eps = d.get('eps_ttm')
            change_pct = d.get('change_pct', 0) or 0

            is_bull = ema20 > ema60
            adx_th = 18 if (ticker.upper().replace('-USD','').isalpha()
                             and ticker.upper().isupper()) else 22
            mkt_tag = ('🇺🇸 US' if (ticker.upper().isalpha()
                                      and ticker.upper().isupper()
                                      and not ticker.endswith('-USD'))
                       else ('🪙 Crypto' if ticker.endswith('-USD')
                             else '🇹🇼 TW'))

            # 標題
            story.append(Paragraph(
                f"#{idx} {ticker} {name}", h1_st))
            story.append(Paragraph(
                f"市場: {mkt_tag} ｜ 收盤: <b>{close:.2f}</b> "
                f"({change_pct:+.2f}%)", body_st))
            story.append(Spacer(1, 0.2*cm))

            # ① 市場環境
            story.append(Paragraph("① 市場環境", h2_st))
            adx_ok = adx >= adx_th
            env_text = (
                f"狀態: <b>{'✅ 多頭市場' if is_bull and adx_ok else ('⚠️ 假多頭' if is_bull else '🚫 空頭')}</b><br/>"
                f"EMA20: {ema20:.2f} ｜ EMA60: {ema60:.2f} "
                f"({'多' if is_bull else '空'}, 差距 {(ema20-ema60)/ema60*100:+.1f}%)<br/>"
                f"ADX: <b>{adx:.1f}</b> "
                f"({'≥' if adx_ok else '<'}{adx_th} {'達標' if adx_ok else '不達標'})<br/>"
                f"黃金交叉: {cd if cd > 0 else '空頭'} 天前"
            )
            if cd and 1 <= cd <= 10:
                if (ticker.upper().isalpha() and ticker.upper().isupper()):
                    if 1 <= cd <= 5:
                        env_text += " ⚡ 早鳥期"
                    else:
                        env_text += " ⚠️ 已過早鳥"
                else:
                    if 5 <= cd <= 7:
                        env_text += " ⭐ Sweet Spot"
            story.append(Paragraph(env_text, body_st))
            story.append(Spacer(1, 0.2*cm))

            # ② 進場判斷
            story.append(Paragraph("② 進場判斷", h2_st))
            t1_ok = is_bull and adx_ok and 0 < cd <= 10
            t3_ok = is_bull and adx_ok and rsi < 50
            t4_ok = (not is_bull) and rsi < 32 and rsi > rsi_prev
            entry_lines = [
                f"T1 黃金交叉 (≤10d): {'✅' if t1_ok else '☐'} (cross={cd}d)",
                f"T3 多頭拉回 (RSI<50): {'✅' if t3_ok else '☐'} (RSI={rsi:.1f})",
                f"T4 空頭反彈 (RSI<32+上升): {'✅' if t4_ok else '☐'}",
            ]
            t3_conf = d.get('t3_confidence', 0) or 0
            if t3_conf > 0:
                entry_lines.append(
                    f"T3 信心度: {'●'*t3_conf}{'○'*(5-t3_conf)} {t3_conf}/5")
            story.append(Paragraph("<br/>".join(entry_lines), body_st))
            story.append(Spacer(1, 0.2*cm))

            # ③ 出場停損
            story.append(Paragraph("③ 出場停損", h2_st))
            atr_mult = 3.0 if adx >= 30 else (
                2.0 if (not is_bull and t4_ok) else 2.5)
            stop_p = close - atr14 * atr_mult
            stop_pct = -atr14 * atr_mult / close * 100 if close > 0 else 0
            story.append(Paragraph(
                f"停損價: <b>{stop_p:.2f}</b> "
                f"(收盤 {close:.2f} − ATR×{atr_mult} {atr14*atr_mult:.2f} "
                f"= {stop_pct:.1f}%)<br/>"
                f"ATR 14: {atr14:.2f} ｜ ATR/Price: {atr14/close*100:.2f}%",
                body_st))
            story.append(Spacer(1, 0.2*cm))

            # ④ 出場獲利
            story.append(Paragraph("④ 出場獲利", h2_st))
            ema_gap = (ema20 - ema60) / ema60 * 100 if ema60 > 0 else 0
            is_high_vol = atr14 / close * 100 > 3.5 if close > 0 else False
            if is_high_vol:
                exit_text = (
                    "🚀 飆股模式 (ATR/P > 3.5%)<br/>"
                    f"EMA死叉差距: {ema_gap:.1f}%<br/>"
                    "RSI 出場: 停用 (飆股 RSI 出場會砍主升段)<br/>"
                    "持倉到 EMA 死叉為止"
                )
            else:
                exit_text = (
                    "🛡 穩健股模式 (ATR/P ≤ 3.5%)<br/>"
                    f"EMA死叉差距: {ema_gap:.1f}%<br/>"
                    f"RSI 出場: ADX<25 + RSI>75 (現 RSI {rsi:.1f})<br/>"
                    "持到 EMA 死叉或停損觸發"
                )
            story.append(Paragraph(exit_text, body_st))
            story.append(Spacer(1, 0.2*cm))

            # 估值參考
            if any(v is not None for v in [eps, per, pbr, div]):
                story.append(Paragraph("💰 估值參考", h2_st))
                val_text = []
                if eps is not None: val_text.append(f"EPS(TTM): {eps:.2f}")
                if per is not None:
                    pe_tag = '虧損' if per <= 0 else (
                        '便宜' if per < 20 else (
                            '合理' if per <= 30 else (
                                '偏高' if per <= 50 else '過熱')))
                    val_text.append(f"PER: {per:.1f} ({pe_tag})")
                if pbr is not None: val_text.append(f"PBR: {pbr:.2f}")
                if div is not None: val_text.append(f"殖利率: {div:.2f}%")
                story.append(Paragraph(" ｜ ".join(val_text), body_st))
                story.append(Spacer(1, 0.2*cm))

            # ⑤ 推薦策略
            story.append(Paragraph("⑤ 推薦策略", h2_st))
            try:
                rec_label, _ = get_rec_label(d, ticker)
            except Exception:
                rec_label = '—'
            story.append(Paragraph(f"<b>{rec_label}</b>", body_st))

            # 完整指標表
            story.append(Paragraph("📐 完整指標數值", h2_st))
            ind_data = [
                ['指標', '值', '指標', '值'],
                ['EMA5', f'{ema5:.2f}' if ema5 else '—',
                 'EMA20', f'{ema20:.2f}'],
                ['EMA60', f'{ema60:.2f}',
                 'EMA120', f'{ema120:.2f}' if ema120 else '—'],
                ['SMA200', f'{sma200:.2f}' if sma200 else '—',
                 'ATR 14', f'{atr14:.2f}'],
                ['RSI 14', f'{rsi:.1f}',
                 'ADX 14', f'{adx:.1f}'],
                ['cross_days', f'{cd}',
                 'ATR/Price%', f'{atr14/close*100:.2f}%' if close else '—'],
            ]
            ind_tbl = Table(ind_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm])
            ind_tbl.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), CN_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7abadd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
            ]))
            story.append(ind_tbl)

            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(
                "⚠️ 本報告僅供參考，投資需自負風險。"
                "詳細策略邏輯請見 INDICATORS.md / MANUAL.md",
                small_st))
            if idx < len(results):
                story.append(PageBreak())
        except Exception as _e:
            story.append(Paragraph(
                f"#{idx} {ticker} 產生失敗: {str(_e)[:80]}", small_st))
            story.append(PageBreak())
            continue

    doc.build(story)
    return buf.getvalue()


def build_stock_range_excel(ticker: str, market: str,
                             start_date: str, end_date: str) -> bytes:
    """產生單一股票指定時間範圍的 Excel 報告（每個交易日一列）"""
    date_rows = fetch_indicators_range(ticker, market, start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = (ticker[:20] + " 歷史")
    ws.sheet_view.showGridLines = False

    def _fl(hx): return PatternFill("solid", start_color=hx, fgColor=hx)
    def _fn(fc="000000", sz=9, bd=False): return Font(name="Arial", size=sz, bold=bd, color=fc)
    _ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _left = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    _mid  = Side(style="medium")
    _thn  = Side(style="thin")

    JUDG_FC = {"買入":"1565C0","強力買入":"0D47A1",
               "賣出":"B71C1C","強力賣出":"7B1616","中立":"555555"}
    VERDICT_BG = {
        "強力買入":"0D3B6E","買入":"0D2E50",
        "上限買入｜持有/短線":"3A2A00","中立":"1A2030",
        "賣出":"3B0D0D","強力賣出":"4A0A0A",
        "過熱觀望｜禁止新倉":"3A1800","空頭，不買":"3A0808",
    }
    VERDICT_FC = {
        "強力買入":"60CFFF","買入":"60B3FF",
        "上限買入｜持有/短線":"F0C030","中立":"9AAABB",
        "賣出":"FF8080","強力賣出":"FF6B6B",
        "過熱觀望｜禁止新倉":"FF8830","空頭，不買":"FF5555",
    }
    G_HDR = {"Z":"1E3A5F","T":"0D2040","P":"2D1A00","M":"1A0D30","X":"1A1A28","O":"2C1654"}
    G_DAT = {"Z":("F0F4FF","F8FAFF"),
             "T":("E8F0FB","F3F7FE"),"P":("FFF3E0","FFFBF0"),
             "M":("F3E8FF","FAF3FF"),"X":("ECEFF4","F5F7FA"),
             "O":("EDE7F6","F5F0FF")}

    # 欄位定義：日期 / 現價 / 漲跌% / 趨勢(5+小結) / 位置(3+小結) / 動能(5+小結) / 輔助(10+小結) / 整體+Cap
    col_defs = [
        ("日期",    12, "Z"), ("現價",   10, "Z"), ("漲跌%",  10, "Z"),
        # 趨勢
        ("趨勢方向", 22, "T"), ("趨勢強度", 20, "T"), ("多頭階段", 22, "T"),
        ("乖離風險", 18, "T"), ("週線結構", 20, "T"), ("趨勢小結(40%)", 26, "T"),
        # 位置
        ("EMA20乖離", 16, "P"), ("RSI(14)", 14, "P"), ("布林%B", 14, "P"),
        ("位置小結(30%)", 26, "P"),
        # 動能
        ("MACD零軸", 14, "M"), ("MACD柱體", 16, "M"), ("動量(10)", 14, "M"),
        ("量能比率", 14, "M"), ("MA10位置", 14, "M"), ("動能小結(20%)", 26, "M"),
        # 輔助
        ("隨機%K",  12, "X"), ("CCI(20)",  12, "X"), ("StochRSI", 12, "X"),
        ("威廉%R",  12, "X"), ("牛熊力度", 12, "X"), ("終極震盪", 12, "X"),
        ("EMA(10)", 12, "X"), ("EMA(60)",  12, "X"), ("SMA(200)", 12, "X"),
        ("Hull MA", 12, "X"), ("輔助小結(10%)", 26, "X"),
        # 整體
        ("操作建議", 28, "O"), ("Cap說明",  36, "O"),
    ]   # 共 32 欄

    for ci, (label, width, gkey) in enumerate(col_defs, 1):
        cl = ws.cell(1, ci, label)
        cl.font      = _fn("FFFFFF", 9, True)
        cl.fill      = _fl(G_HDR[gkey])
        cl.alignment = _ctr
        cl.border    = Border(bottom=_mid, right=_thn)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "D2"   # 凍結前三欄（日期/現價/漲跌%）

    if not date_rows:
        cl = ws.cell(2, 1, "無資料：請確認代號或日期範圍")
        cl.font = _fn("CC4400", 10, True)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def _summ(b, s, n, v): return f"買:{b} 賣:{s} 中:{n} → {v}"
    def _ival(item): return (item[1], item[2]) if len(item) == 3 else (item[0], item[1])

    for ri, (date_str, d_dict) in enumerate(date_rows, 2):
        ws.row_dimensions[ri].height = 18
        is_even = (ri % 2 == 0)

        def mk(col, val, gkey, fc="000000", sz=9, bd=False, al=_ctr):
            bg = G_DAT[gkey][0] if is_even else G_DAT[gkey][1]
            cl = ws.cell(ri, col, val)
            cl.font = _fn(fc, sz, bd); cl.fill = _fl(bg)
            cl.alignment = al; cl.border = Border(right=_thn)
            return cl

        def mv(col, verdict, gkey):
            cl = ws.cell(ri, col, verdict)
            cl.font  = _fn(VERDICT_FC.get(verdict, "FFFFFF"), 9, True)
            cl.fill  = _fl(VERDICT_BG.get(verdict, "626567"))
            cl.alignment = _ctr
            cl.border = Border(right=Side(style="medium"))
            return cl

        if not d_dict or not d_dict.get("close"):
            mk(1, date_str, "Z", "444444", 9)
            for col in range(2, len(col_defs)+1):
                mk(col, "無資料", "Z", "AAAAAA", 9)
            continue

        # ── 計算指標 ─────────────────────────────────────────────
        g_trend    = judge_trend(d_dict)
        g_position = judge_position(d_dict)
        g_momentum = judge_momentum(d_dict)
        g_aux      = judge_aux(d_dict)
        ts = calc_summary(g_trend,    TREND_W)
        ps = calc_summary(g_position, POSITION_W)
        mom_grade  = compute_momentum_grade(d_dict)
        ms_b, ms_s, ms_n, _ = calc_summary(g_momentum, MOMENTUM_W)
        ms  = (ms_b, ms_s, ms_n, mom_grade)
        xs  = _calc_aux_summary(g_aux, AUX_W)
        tb  = round(ts[0]+ps[0]+ms[0]+xs[0], 1)
        ts_ = round(ts[1]+ps[1]+ms[1]+xs[1], 1)
        tn_ = round(ts[2]+ps[2]+ms[2]+xs[2], 1)
        verdict, cap = apply_cap(_rec(tb, ts_), d_dict, mom_grade)

        close_v    = d_dict.get("close")
        change_pct = d_dict.get("change_pct")
        mk(1, date_str, "Z", "1565C0", 9, True)
        mk(2, f"{close_v:.2f}" if close_v else "N/A", "Z", "222222", 9)
        if change_pct is not None:
            mk(3, f"{'▲' if change_pct>=0 else '▼'}{abs(change_pct):.2f}%",
               "Z", "B71C1C" if change_pct >= 0 else "1B5E20", 9, True)
        else:
            mk(3, "N/A", "Z", "888888", 9)

        ci = 4
        # ── 趨勢 5+小結 ────────────────────────────────────────
        for it in g_trend[:5]:
            v2, j = _ival(it)
            mk(ci, f"{v2} / {j}", "T", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci+=1
        mv(ci, ts[3], "T"); ws.cell(ri, ci).value = _summ(*ts); ci+=1

        # ── 位置 3+小結 ────────────────────────────────────────
        for it in g_position[:3]:
            v2, j = _ival(it)
            mk(ci, f"{v2} / {j}", "P", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci+=1
        mv(ci, ps[3], "P"); ws.cell(ri, ci).value = _summ(*ps); ci+=1

        # ── 動能 5+小結 ────────────────────────────────────────
        for it in g_momentum[:5]:
            v2, j = _ival(it)
            mk(ci, f"{v2} / {j}", "M", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci+=1
        mv(ci, ms[3], "M"); ws.cell(ri, ci).value = _summ(*ms); ci+=1

        # ── 輔助 10+小結 ───────────────────────────────────────
        aux_start = ci
        for it in g_aux[:10]:
            v2, j = _ival(it)
            mk(ci, f"{v2} / {j}", "X", JUDG_FC.get(j,"444444"), 9, j!="中立"); ci+=1
        while ci < aux_start + 10:     # 補空格（項目不足 10 時）
            mk(ci, "", "X"); ci+=1
        mv(ci, xs[3], "X"); ws.cell(ri, ci).value = _summ(*xs); ci+=1

        # ── 整體 ───────────────────────────────────────────────
        mv(ci, verdict, "O"); ws.cell(ri, ci).value = verdict; ci+=1
        mk(ci, cap if cap else "—", "O", "884400" if cap else "888888", 9, False, _left)

    # 時間戳記
    ws.cell(len(date_rows)+3, 1,
            f"產出：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
            f"{ticker} | {start_date}～{end_date}"
            ).font = _fn("999999", 8)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="tv-header" style="justify-content:space-between">
  <div style="display:flex;align-items:center;gap:14px">
    <div style="font-size:1.8rem">📊</div>
    <div><h1>Indicator Scanner</h1>
    <div class="sub">四群組加權評分 · 趨勢40% / 位置30% / 動能20% / 輔助10% · Hard Limits · Excel 匯出</div></div>
  </div>
  <div style="text-align:right;font-family:'IBM Plex Mono',monospace;line-height:1.4">
    <div style="color:#3b9eff;font-size:.95rem;font-weight:700;letter-spacing:.04em">{APP_VERSION}</div>
    <div style="color:#5a8ab0;font-size:.7rem">更新：{APP_UPDATED}</div>
  </div>
</div>
<div style="background:#0d1b2e;border:1px solid #1e3a5f;border-radius:6px;padding:8px 14px;margin-bottom:.5rem;font-size:.72rem;color:#7aaac8">
  📌 <b style="color:#8ab8d8">{APP_VERSION} 更新重點</b>：{APP_NOTES}
</div>
<div style="background:#0a1628;border:1px solid #1a2f48;border-radius:6px;padding:6px 14px;margin-bottom:1rem;font-size:.68rem;color:#5a8ab0">
  🔬 <b style="color:#7aaac8">健壯性驗證</b>：{APP_VALIDATIONS}
</div>""", unsafe_allow_html=True)


# 🆕 v9.9h：TOP 200 即時掃描（永遠顯示在頂部，不受 results 影響）
# v9.10：加入更新按鈕 + 過期警示（fix 「不會自更新」）
def _signal_age_days(updated_str, market='tw'):
    """計算 updated_at 距今幾天，回傳 (days, is_stale)
    🆕 v9.10g：美股過期門檻放寬到 >2 天（時差 + 假日因素）
       台股 >1 天即視為過期"""
    import datetime as _dt
    threshold = 2.0 if market == 'us' else 1.0
    try:
        dt = _dt.datetime.strptime(updated_str.strip()[:10], '%Y-%m-%d')
        diff = (_dt.datetime.now() - dt).total_seconds() / 86400
        return diff, diff > threshold
    except Exception:
        return 99, True


def _is_cloud_env():
    """偵測是否在雲端環境（沒 data_cache）"""
    from pathlib import Path as _P
    dc = _P(__file__).parent / 'data_cache'
    if not dc.exists(): return True
    # 抽樣：data_cache 有 < 100 個 parquet → 視為雲端
    n = sum(1 for _ in dc.glob('*.parquet'))
    return n < 100


def _show_update_result_if_any():
    """如有上次更新結果，顯示在頁面頂部（不會被 rerun 清掉）
    在 panel 開頭呼叫"""
    res = st.session_state.get('_last_update_result')
    if not res: return
    market, ok, msg, ts = res
    flag = '🇹🇼' if market == 'tw' else '🇺🇸'
    if ok:
        st.success(f"✅ {flag} 更新完成（{ts}）— 訊號已刷新")
    else:
        st.error(f"❌ {flag} 更新失敗（{ts}）")
    with st.expander("📋 完整 log（點擊展開）", expanded=not ok):
        st.code(msg or "(無輸出)", language='text')
    if st.button("✕ 清除此訊息", key=f"clr_update_msg_{market}"):
        del st.session_state['_last_update_result']
        st.rerun()


def _trigger_update_signals(script_name, market='tw'):
    """跑 update，回傳 (ok, msg)
    🆕 v9.10d：雲端直接 import update_signals_cloud 模組呼叫（避開 subprocess 環境問題）
    🏠 本機跑完整 update_*_signals.py（含 backtest，subprocess）"""
    from pathlib import Path as _P
    import traceback as _tb
    proj = _P(__file__).parent

    # ☁️ 雲端 → 直接 in-process import 呼叫（最可靠）
    if _is_cloud_env():
        try:
            # 動態 import（每次都 reload 確保最新）
            import importlib, sys as _sys
            spec_path = proj / 'update_signals_cloud.py'
            if not spec_path.exists():
                return False, "update_signals_cloud.py 不存在"
            # 把 proj 加 sys.path 確保 import 到對的版本
            if str(proj) not in _sys.path:
                _sys.path.insert(0, str(proj))
            if 'update_signals_cloud' in _sys.modules:
                cloud = importlib.reload(_sys.modules['update_signals_cloud'])
            else:
                cloud = importlib.import_module('update_signals_cloud')
            # 直接呼叫對應函式
            import io as _io, contextlib as _ct
            buf = _io.StringIO()
            with _ct.redirect_stdout(buf), _ct.redirect_stderr(buf):
                if market == 'tw':
                    ok = cloud.update_tw()
                elif market == 'us':
                    ok = cloud.update_us()
                else:
                    return False, f"unknown market: {market}"
            log = buf.getvalue()
            return bool(ok), log[-1500:] if log else "(無輸出)"
        except Exception as e:
            err = f"Exception: {type(e).__name__}: {e}\n\n{_tb.format_exc()[-1000:]}"
            return False, err

    # 🏠 本機 → 跑完整 subprocess
    import subprocess as _sp
    import sys as _sys
    try:
        result = _sp.run([_sys.executable, str(proj / script_name)],
                         cwd=str(proj), capture_output=True, text=True,
                         timeout=600, encoding='utf-8', errors='replace')
        return result.returncode == 0, result.stdout[-1000:] + result.stderr[-500:]
    except Exception as e:
        return False, str(e)


def _render_top200_panel():
    # （v9.10e bug 修復：result 顯示移到全頁面頂端，避免兩個 panel 重複 button key）
    try:
        from pathlib import Path as _P
        import json as _json
        sig_path = _P(__file__).parent / 'top200_signals.json'
        if not sig_path.exists():
            st.warning("⚠️ top200_signals.json 不存在 — 請先跑 `python update_daily_signals.py`")
            return
        sig_data = _json.load(open(sig_path, encoding='utf-8'))
        _e_raw = sig_data.get('entry', [])
        _x_raw = sig_data.get('exit', [])
        _h_raw = sig_data.get('hold', [])
        _updated = sig_data.get('updated_at', '?')
        age_days, is_stale = _signal_age_days(_updated)

        # 🆕 過期警示 + 一鍵更新（雲端 yfinance 即時抓取版）
        is_cloud = _is_cloud_env()
        if is_stale or _updated == 'unknown':
            col1, col2 = st.columns([5, 1])
            with col1:
                env_str = '☁️ 雲端 yfinance 即時抓取' if is_cloud else '本機 data_cache'
                st.warning(
                    f"⚠️ TOP 200 訊號已過期 **{age_days:.1f} 天**（最後更新 {_updated}）— "
                    f"{env_str}，點右側按鈕立即更新")
            with col2:
                btn_label = "🔄 雲端更新" if is_cloud else "🔄 立即更新"
                btn_help = ("yfinance 抓最新一日 ~10 秒" if is_cloud
                            else "完整 update 約 2-5 分鐘")
                if st.button(btn_label, key="refresh_top200",
                             help=btn_help, use_container_width=True):
                    spinner_msg = "雲端 yfinance 抓取中…（~10 秒）" if is_cloud \
                                  else "正在更新 TOP 200 訊號…（約 2-5 分鐘）"
                    import datetime as _dt
                    import traceback as _tb
                    ts = _dt.datetime.now().strftime('%H:%M:%S')
                    # 🛡 全外層 try/except 防止 exception 吞掉 result
                    try:
                        with st.spinner(spinner_msg):
                            ok, msg = _trigger_update_signals(
                                'update_daily_signals.py', market='tw')
                        st.session_state['_last_update_result'] = (
                            'tw', ok, msg, ts)
                    except Exception as _e:
                        st.session_state['_last_update_result'] = (
                            'tw', False,
                            f"Outer exception: {type(_e).__name__}: {_e}\n\n"
                            f"{_tb.format_exc()}",
                            ts)
                    st.cache_data.clear()
                    st.rerun()

        if not (_e_raw or _x_raw or _h_raw):
            return

        # 主面板（過期時邊框轉橙色）
        border_color = '#ff9944' if is_stale else '#3dbb6a55'
        title_color = '#ff9944' if is_stale else '#3dbb6a'
        age_str = f"⚠️ {age_days:.1f} 天前" if is_stale else "今日"
        st.markdown(
            f'<div style="background:#0a1a2a;border:1px solid {border_color};border-radius:10px;'
            f'padding:10px 14px;margin:8px 0;display:flex;gap:14px;align-items:center;flex-wrap:wrap">'
            f'<div style="font-size:1.05rem;font-weight:700;color:{title_color}">📊 {age_str} TOP 200 即時掃描</div>'
            f'<div style="color:#7abadd;font-size:.85rem">'
            f'🚀 進場 <b>{len(_e_raw)}</b>　│　🚪 出倉 <b>{len(_x_raw)}</b>　│　📌 持倉 <b>{len(_h_raw)}</b></div>'
            f'<div style="color:#7a8899;font-size:.72rem;flex:1">'
            f'⭐ TOP 200 適用清單　│　資料 {_updated}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        def _row_html(rows, color, label, max_n=999):
            """max_n=999 預設顯示全部（先前限 15 檔遭使用者反映）
            label='' 時不顯示頂部標題（用於 expander 內，標題已在 expander label）"""
            if not rows:
                return f'<div style="color:#3a5a7a;font-size:.72rem;padding:6px">— 暫無 —</div>'
            out = ''
            if label:
                out = (f'<div style="color:{color};font-size:.85rem;font-weight:700;'
                       f'padding:4px 8px 6px">{label} ({len(rows)})</div>')
            for r in rows[:max_n]:
                sig = r.get('sig', '')
                # 🆕 T3 信心度 ●○ — 只在 T3 訊號顯示（T1/飆股不顯示）
                conf = r.get('t3_confidence', 0) or 0
                is_t3 = sig.startswith('T3') or 'T3' in sig
                if conf > 0 and is_t3:
                    conf_html = (
                        f'<span style="font-size:.65rem;letter-spacing:1px;'
                        f'font-family:monospace" title="T3 信心度 {conf}/5">'
                        f'<span style="color:#3dbb6a">{"●"*conf}</span>'
                        f'<span style="color:#3a5a7a">{"○"*(5-conf)}</span>'
                        f'</span>')
                else:
                    conf_html = ''
                # 🆕 P/E（顏色判定）
                pe_v = r.get('pe')
                if pe_v is None:
                    pe_html = '<span style="color:#3a5a7a;font-size:.7rem">PE —</span>'
                else:
                    if pe_v <= 0 or pe_v > 100:
                        pe_color = '#ff5555'
                    elif pe_v < 20:
                        pe_color = '#3dbb6a'
                    elif pe_v <= 30:
                        pe_color = '#c8b87a'
                    elif pe_v <= 50:
                        pe_color = '#e8a020'
                    else:
                        pe_color = '#ff5555'
                    pe_html = f'<span style="color:{pe_color};font-size:.7rem">PE {pe_v:.1f}</span>'
                # 🆕 v9.10g：safe formatting，避免 None 拋 exception
                _close_v = r.get("close", 0) or 0
                _delta_v = r.get("delta", 0) or 0
                # 🆕 v9.10m：delta = 60 日動量，用顏色區分漲跌
                if _delta_v >= 5:
                    _delta_color, _delta_bg = '#3dbb6a', '#0a3a1f'  # 強綠
                elif _delta_v >= 0:
                    _delta_color, _delta_bg = '#7abadd', '#0a2030'  # 弱藍
                elif _delta_v >= -10:
                    _delta_color, _delta_bg = '#e8a020', '#1a1408'  # 橙
                else:
                    _delta_color, _delta_bg = '#ff5555', '#2a0808'  # 紅
                _delta_str = f'{_delta_v:+.0f}%' if abs(_delta_v) < 1000 else f'{_delta_v:+.0f}%'
                out += (
                    f'<div style="display:flex;gap:6px;padding:3px 8px;font-size:.78rem;'
                    f'border-bottom:1px solid #1a2a3f;align-items:baseline">'
                    f'<span style="color:{color};font-weight:700;font-family:monospace;'
                    f'min-width:48px">{r.get("ticker", "?")}</span>'
                    f'<span style="color:#a8cce8;flex:1;overflow:hidden;text-overflow:ellipsis;'
                    f'white-space:nowrap;max-width:90px">{r.get("name","")}</span>'
                    f'<span style="color:#e8f4fd;font-family:monospace">{_close_v:.2f}</span>'
                    f'{pe_html}'
                    f'<span style="color:#7a8899;font-size:.7rem">{sig}</span>'
                    f'{conf_html}'
                    f'<span style="color:{_delta_color};font-size:.7rem;background:{_delta_bg};'
                    f'padding:1px 5px;border-radius:3px" title="60日動量">{_delta_str}</span>'
                    f'</div>'
                )
            if len(rows) > max_n:
                out += (f'<div style="text-align:center;padding:4px;color:#5a8ab0;font-size:.7rem">'
                        f'＋{len(rows)-max_n} 檔</div>')
            return out

        cols = st.columns(3)
        with cols[0]:
            with st.expander(f"🚀 可進場 ({len(_e_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#0a1e10;border:1px solid #3dbb6a55;border-radius:8px;'
                    f'padding:6px 4px">{_row_html(_e_raw, "#3dbb6a", "")}</div>',
                    unsafe_allow_html=True)
            # 🆕 v9.9l：存成自選股清單（按鈕在進場欄下方）
            if _e_raw:
                _save_cols = st.columns([3, 1])
                with _save_cols[0]:
                    _wl_name = st.text_input(
                        "自選股名稱",
                        value=f"TOP200進場_{_updated}",
                        key="top200_save_name",
                        label_visibility="collapsed",
                        placeholder="自選股名稱",
                    )
                with _save_cols[1]:
                    if st.button("💾 存", key="top200_save_btn",
                                 use_container_width=True,
                                 help=f"把這 {len(_e_raw)} 檔可進場股票存成自選股清單"):
                        # 載入現有 watchlists（雙軌：localStorage + 檔案）
                        try:
                            from streamlit_local_storage import LocalStorage as _LS
                            _ls = _LS()
                        except Exception:
                            _ls = None
                        _wl_path = _P(__file__).parent / 'watchlists.json'

                        wls = {}
                        if _ls:
                            try:
                                v = _ls.getItem("stock001_watchlists")
                                if v:
                                    wls = _json.loads(v) if isinstance(v, str) else v
                            except Exception:
                                pass
                        if not wls and _wl_path.exists():
                            try:
                                wls = _json.loads(_wl_path.read_text(encoding='utf-8'))
                            except Exception:
                                pass

                        # 新增清單
                        tickers_str = "\n".join(r['ticker'] for r in _e_raw)
                        wls[_wl_name] = tickers_str
                        # 儲存
                        text = _json.dumps(wls, ensure_ascii=False, indent=2)
                        if _ls:
                            try: _ls.setItem("stock001_watchlists", text)
                            except: pass
                        try:
                            _wl_path.write_text(text, encoding='utf-8')
                        except Exception:
                            pass
                        st.success(f"✅ 已存「{_wl_name}」（{len(_e_raw)} 檔）")
                        st.rerun()
        with cols[1]:
            with st.expander(f"🚪 應出倉 ({len(_x_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#1a0808;border:1px solid #ff555555;border-radius:8px;'
                    f'padding:6px 4px">{_row_html(_x_raw, "#ff7777", "")}</div>',
                    unsafe_allow_html=True)
        with cols[2]:
            # 持倉中 預設收合（通常較多，避免畫面太長）
            with st.expander(f"📌 持倉中 ({len(_h_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#0a1830;border:1px solid #5a8ab055;border-radius:8px;'
                    f'padding:6px 4px">{_row_html(_h_raw, "#7abadd", "")}</div>',
                    unsafe_allow_html=True)
    except Exception as _panel_err:
        # 🆕 v9.10g：不再吞 exception — 顯示錯誤給用戶看
        import traceback as _tb
        st.error(f"❌ TOP 200 panel render 失敗：{type(_panel_err).__name__}: {_panel_err}")
        with st.expander("traceback"):
            st.code(_tb.format_exc())


# 🆕 v9.10e：在所有 panel 之前統一顯示更新結果（只呼叫一次避免 duplicate key）
_show_update_result_if_any()


# 🆕 v9.10z：上次更新時間橫幅（讓用戶清楚資料新鮮度）
def _render_update_status_bar():
    """顯示 alerts JSON 的時間戳，讓用戶確認資料新鮮度"""
    try:
        from pathlib import Path as _P
        import json as _json
        import datetime as _dt
        proj = _P(__file__).parent
        rows = []
        for fname, label, emoji in [
            ('top200_signals.json', 'TW TOP 200', '🇹🇼'),
            ('us_top200_signals.json', 'US TOP 100', '🇺🇸'),
        ]:
            p = proj / fname
            if not p.exists(): continue
            try:
                d = _json.loads(p.read_text(encoding='utf-8'))
            except: continue
            updated = d.get('updated_at', '?')
            computed = d.get('computed_at', '')
            n_alerts = len(d.get('alerts', []))
            n_entry = len(d.get('entry', []))
            source = d.get('source', '?')
            # 計算過期天數
            try:
                dt = _dt.datetime.strptime(updated[:10], '%Y-%m-%d')
                age = (_dt.datetime.now() - dt).total_seconds() / 86400
                age_str = (f'<span style="color:#3dbb6a">{age:.1f}d 前</span>'
                           if age < 1.5 else
                           f'<span style="color:#e8a020">⚠️ {age:.1f}d 前</span>'
                           if age < 3 else
                           f'<span style="color:#ff5555">🚨 {age:.1f}d 前過期</span>')
            except:
                age_str = '<span style="color:#7a8899">unknown</span>'
            rows.append({
                'emoji': emoji, 'label': label,
                'updated': updated, 'computed': computed,
                'n_alerts': n_alerts, 'n_entry': n_entry,
                'age_str': age_str, 'source': source,
            })
        if not rows: return

        cells = []
        for r in rows:
            cells.append(
                f'<div style="flex:1;min-width:220px;padding:6px 10px;'
                f'background:#0a1628;border-radius:6px;'
                f'border:1px solid #5a8ab055">'
                f'<div style="display:flex;gap:8px;align-items:center">'
                f'<span style="font-size:1rem">{r["emoji"]}</span>'
                f'<span style="color:#7abadd;font-weight:700;font-size:.82rem">'
                f'{r["label"]}</span>'
                f'<span style="color:#7a8899;font-size:.7rem">資料 {r["updated"]}</span>'
                f'<span style="font-size:.7rem;margin-left:auto">{r["age_str"]}</span>'
                f'</div>'
                f'<div style="color:#a8cce8;font-size:.72rem;margin-top:2px">'
                f'進場 <b>{r["n_entry"]}</b>檔｜警報 <b>{r["n_alerts"]}</b>檔｜'
                f'計算 {r["computed"][-8:] if r["computed"] else "?"}'
                f'</div>'
                f'</div>'
            )
        st.markdown(
            f'<div style="display:flex;gap:8px;margin:6px 0;flex-wrap:wrap">'
            + ''.join(cells) +
            f'</div>'
            f'<div style="text-align:right;color:#5a7a90;font-size:.65rem;'
            f'margin-top:-2px">⏰ 自動更新: 台股收盤後 13:30 / 美股盤後 06:00 (台北)'
            f'｜手動觸發: 各 panel 內按「🔄 更新」按鈕</div>',
            unsafe_allow_html=True
        )
    except Exception:
        pass


_render_update_status_bar()
# 🐛 v9.22.2：用戶要求移除 TW TOP 200 panel + US TOP panel
# _render_top200_panel()


# 🆕 v9.9o：美股 TOP 100 panel
# v9.10：加入更新按鈕 + 過期警示
def _render_us_top_panel():
    try:
        from pathlib import Path as _P
        import json as _json
        sig_path = _P(__file__).parent / 'us_top200_signals.json'
        if not sig_path.exists():
            st.warning("⚠️ us_top200_signals.json 不存在 — 請跑 `python update_us_signals.py`")
            return
        sig_data = _json.load(open(sig_path, encoding='utf-8'))
        _e_raw = sig_data.get('entry', [])
        _x_raw = sig_data.get('exit', [])
        _h_raw = sig_data.get('hold', [])
        _updated = sig_data.get('updated_at', '?')
        _tot = sig_data.get('top_total', 100)
        age_days, is_stale = _signal_age_days(_updated, market='us')

        # 🆕 過期警示 + 一鍵更新（雲端 yfinance 即時抓取版）
        is_cloud = _is_cloud_env()
        if is_stale or _updated == 'unknown':
            col1, col2 = st.columns([5, 1])
            with col1:
                env_str = '☁️ 雲端 yfinance 即時抓取' if is_cloud else '本機 data_cache'
                st.warning(
                    f"⚠️ US TOP 訊號已過期 **{age_days:.1f} 天**（最後更新 {_updated}）— "
                    f"{env_str}，點右側按鈕立即更新")
            with col2:
                btn_label = "🔄 雲端更新 US" if is_cloud else "🔄 立即更新 US"
                btn_help = ("yfinance 抓最新一日 ~10 秒" if is_cloud
                            else "完整 update 約 3-8 分鐘")
                if st.button(btn_label, key="refresh_us_top",
                             help=btn_help, use_container_width=True):
                    spinner_msg = "雲端 yfinance 抓取中…（~10 秒）" if is_cloud \
                                  else "正在更新 US 訊號…（約 3-8 分鐘）"
                    import datetime as _dt
                    import traceback as _tb
                    ts = _dt.datetime.now().strftime('%H:%M:%S')
                    try:
                        with st.spinner(spinner_msg):
                            ok, msg = _trigger_update_signals(
                                'update_us_signals.py', market='us')
                        st.session_state['_last_update_result'] = (
                            'us', ok, msg, ts)
                    except Exception as _e:
                        st.session_state['_last_update_result'] = (
                            'us', False,
                            f"Outer exception: {type(_e).__name__}: {_e}\n\n"
                            f"{_tb.format_exc()}",
                            ts)
                    st.cache_data.clear()
                    st.rerun()

        if not (_e_raw or _x_raw or _h_raw):
            return

        border_color = '#ff9944' if is_stale else '#5a8ab055'
        title_color = '#ff9944' if is_stale else '#7abadd'
        age_str = f"⚠️ {age_days:.1f} 天前" if is_stale else "今日"
        st.markdown(
            f'<div style="background:#0a1a2a;border:1px solid {border_color};border-radius:10px;'
            f'padding:10px 14px;margin:8px 0;display:flex;gap:14px;align-items:center;flex-wrap:wrap">'
            f'<div style="font-size:1.05rem;font-weight:700;color:{title_color}">🇺🇸 {age_str} US TOP {_tot} 即時掃描</div>'
            f'<div style="color:#a8cce8;font-size:.85rem">'
            f'🚀 進場 <b>{len(_e_raw)}</b>　│　🚪 出倉 <b>{len(_x_raw)}</b>　│　📌 持倉 <b>{len(_h_raw)}</b></div>'
            f'<div style="color:#7a8899;font-size:.72rem;flex:1">'
            f'P10_T1T3+POS+ADX18 高流動 ADV≥$104M（v9.10 美股最佳 RR 0.496）　│　資料 {_updated}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        def _us_row_html(rows):
            if not rows:
                return f'<div style="color:#3a5a7a;font-size:.72rem;padding:6px">— 暫無 —</div>'
            out = ''
            for r in rows:
                sig = r.get('sig', '')
                _close_v = r.get("close", 0) or 0
                _delta_v = r.get("delta", 0) or 0
                _name = r.get("name", "") or ""
                # 🆕 v9.10m：delta = 60 日動量，用顏色區分
                if _delta_v >= 5:
                    _dc, _dbg = '#3dbb6a', '#0a3a1f'
                elif _delta_v >= 0:
                    _dc, _dbg = '#7abadd', '#0a2030'
                elif _delta_v >= -10:
                    _dc, _dbg = '#e8a020', '#1a1408'
                else:
                    _dc, _dbg = '#ff5555', '#2a0808'
                _dstr = f'{_delta_v:+.0f}%'
                out += (
                    f'<div style="display:flex;gap:6px;padding:3px 8px;font-size:.78rem;'
                    f'border-bottom:1px solid #1a2a3f;align-items:baseline">'
                    f'<span style="font-weight:700;font-family:monospace;'
                    f'min-width:60px;color:#e8f4fd">{r.get("ticker", "?")}</span>'
                    f'<span style="color:#a8cce8;flex:1;overflow:hidden;text-overflow:ellipsis;'
                    f'white-space:nowrap;max-width:130px;font-size:.72rem">{_name}</span>'
                    f'<span style="color:#a8cce8;font-family:monospace">{_close_v:.2f}</span>'
                    f'<span style="color:#7a8899;font-size:.7rem">{sig}</span>'
                    f'<span style="font-size:.7rem;background:{_dbg};color:{_dc};'
                    f'padding:1px 5px;border-radius:3px" title="60日動量">{_dstr}</span>'
                    f'</div>'
                )
            return out

        cols = st.columns(3)
        with cols[0]:
            with st.expander(f"🚀 可進場 ({len(_e_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#0a1e10;border:1px solid #3dbb6a55;border-radius:8px;'
                    f'padding:6px 4px">{_us_row_html(_e_raw)}</div>',
                    unsafe_allow_html=True)
            if _e_raw:
                _us_save = st.columns([3, 1])
                with _us_save[0]:
                    _us_name = st.text_input(
                        "美股自選名稱",
                        value=f"US_進場_{_updated}",
                        key="us_save_name",
                        label_visibility="collapsed",
                    )
                with _us_save[1]:
                    if st.button("💾 存", key="us_save_btn", use_container_width=True):
                        try:
                            from streamlit_local_storage import LocalStorage as _LS
                            _ls = _LS()
                        except Exception:
                            _ls = None
                        _wl_path = _P(__file__).parent / 'watchlists.json'
                        wls = {}
                        if _ls:
                            try:
                                v = _ls.getItem("stock001_watchlists")
                                if v: wls = _json.loads(v) if isinstance(v, str) else v
                            except Exception: pass
                        if not wls and _wl_path.exists():
                            try: wls = _json.loads(_wl_path.read_text(encoding='utf-8'))
                            except Exception: pass
                        wls[_us_name] = "\n".join(r['ticker'] for r in _e_raw)
                        text = _json.dumps(wls, ensure_ascii=False, indent=2)
                        if _ls:
                            try: _ls.setItem("stock001_watchlists", text)
                            except: pass
                        try: _wl_path.write_text(text, encoding='utf-8')
                        except: pass
                        st.success(f"✅ 已存「{_us_name}」({len(_e_raw)} 檔)")
                        st.rerun()
        with cols[1]:
            with st.expander(f"🚪 應出倉 ({len(_x_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#1a0808;border:1px solid #ff555555;border-radius:8px;'
                    f'padding:6px 4px">{_us_row_html(_x_raw)}</div>',
                    unsafe_allow_html=True)
        with cols[2]:
            with st.expander(f"📌 持倉中 ({len(_h_raw)})", expanded=False):
                st.markdown(
                    f'<div style="background:#0a1830;border:1px solid #5a8ab055;border-radius:8px;'
                    f'padding:6px 4px">{_us_row_html(_h_raw)}</div>',
                    unsafe_allow_html=True)
    except Exception as _us_err:
        # 🆕 v9.10g：不再吞 exception
        import traceback as _tb
        st.error(f"❌ US TOP panel render 失敗：{type(_us_err).__name__}: {_us_err}")
        with st.expander("traceback"):
            st.code(_tb.format_exc())


# 🐛 v9.22.2：用戶要求移除 US TOP panel
# _render_us_top_panel()


# 🆕 v9.10u：產業輪動 Top 5 / Bottom 5 panel
def _render_sector_rotation_panel():
    """顯示過去 1 個月最強 5 + 最弱 5 個產業（動量延續策略）"""
    try:
        rankings = _get_sector_ranking_recent()
        if not rankings:
            return
        top5 = rankings[:5]
        bot5 = rankings[-5:][::-1]  # 由最弱開始顯示
        # Top 5 panel
        st.markdown(
            f'<div style="background:#0a1628;border:1px solid #ffd70055;'
            f'border-radius:10px;padding:10px 14px;margin:8px 0">'
            f'<div style="display:flex;gap:14px;align-items:center;'
            f'justify-content:space-between;flex-wrap:wrap;margin-bottom:8px">'
            f'<div style="font-size:1.05rem;font-weight:700;color:#ffd700">'
            f'🎯 產業輪動 Top 5 強勢（動量延續策略）</div>'
            f'<div style="color:#7a8899;font-size:.72rem">'
            f'過去 1 月平均月報酬｜回測 6 年總 +400%（vs 等權 +138%）/ Sharpe 1.24</div>'
            f'</div>'
            f'<div style="display:flex;gap:8px;flex-wrap:wrap">'
            + ''.join([
                f'<div style="background:#1a2030;border:1px solid {("#3dbb6a" if r > 0 else "#ff5555")}55;'
                f'border-radius:6px;padding:6px 10px;flex:1;min-width:150px">'
                f'<div style="color:#ffd700;font-weight:700;font-size:.85rem">'
                f'#{i+1} {sec}</div>'
                f'<div style="color:{("#3dbb6a" if r > 0 else "#ff5555")};font-size:1.1rem;font-weight:700">'
                f'{r:+.2f}%</div>'
                f'<div style="color:#7a8899;font-size:.7rem">{n} 檔平均</div>'
                f'</div>'
                for i, (sec, r, n) in enumerate(top5)
            ])
            + f'</div></div>',
            unsafe_allow_html=True
        )
        # 🆕 v9.10v：Bottom 5 弱勢產業 panel
        st.markdown(
            f'<div style="background:#1a0a0a;border:1px solid #ff555555;'
            f'border-radius:10px;padding:10px 14px;margin:8px 0">'
            f'<div style="display:flex;gap:14px;align-items:center;'
            f'justify-content:space-between;flex-wrap:wrap;margin-bottom:8px">'
            f'<div style="font-size:1.05rem;font-weight:700;color:#ff7755">'
            f'⚠️ 產業輪動 Bottom 5 弱勢（避開或反指標）</div>'
            f'<div style="color:#7a8899;font-size:.72rem">'
            f'過去 1 月最弱｜動量延續策略避開這些產業</div>'
            f'</div>'
            f'<div style="display:flex;gap:8px;flex-wrap:wrap">'
            + ''.join([
                f'<div style="background:#1a1010;border:1px solid {("#3dbb6a" if r > 0 else "#ff5555")}55;'
                f'border-radius:6px;padding:6px 10px;flex:1;min-width:150px">'
                f'<div style="color:#ff7755;font-weight:700;font-size:.85rem">'
                f'弱 {i+1} {sec}</div>'
                f'<div style="color:{("#3dbb6a" if r > 0 else "#ff5555")};font-size:1.1rem;font-weight:700">'
                f'{r:+.2f}%</div>'
                f'<div style="color:#7a8899;font-size:.7rem">{n} 檔平均</div>'
                f'</div>'
                for i, (sec, r, n) in enumerate(bot5)
            ])
            + f'</div></div>',
            unsafe_allow_html=True
        )
    except Exception:
        pass


# 🐛 v9.20.9：用戶要求移除產業輪動 panel（Top 5 / Bottom 5）
# _render_sector_rotation_panel()


# 🆕 v9.10y：警報中 panel（強訊號 + 即將觸發）
def _render_alerts_panel():
    """從 top200_signals.json + us_top200_signals.json 讀 alerts 列表並顯示"""
    try:
        from pathlib import Path as _P
        import json as _json
        all_alerts = []
        for fname, market_tag in [
            ('top200_signals.json', '🇹🇼'),
            ('us_top200_signals.json', '🇺🇸'),
        ]:
            p = _P(__file__).parent / fname
            if not p.exists(): continue
            d = _json.loads(p.read_text(encoding='utf-8'))
            for a in d.get('alerts', []):
                a = {**a, 'market': market_tag}
                all_alerts.append(a)
        if not all_alerts:
            return

        bull5 = [a for a in all_alerts if a.get('level') == 5]
        bull4 = [a for a in all_alerts if a.get('level') == 4 and a.get('side') == 'bull']
        bull3 = [a for a in all_alerts if a.get('level') == 3 and a.get('side') == 'bull']
        bear4 = [a for a in all_alerts if a.get('level') == 4 and a.get('side') == 'bear']
        bear3 = [a for a in all_alerts if a.get('level') == 3 and a.get('side') == 'bear']
        bear2 = [a for a in all_alerts if a.get('level') == 2 and a.get('side') == 'bear']
        # 🆕 v9.12：T1 即將上穿（level=2 + 'T1 即將上穿' tag）獨立分組
        bull2_t1 = [a for a in all_alerts if a.get('level') == 2 and a.get('side') == 'bull'
                    and 'T1 即將上穿' in a.get('tag', '')]
        bull2_other = [a for a in all_alerts if a.get('level') == 2 and a.get('side') == 'bull'
                       and 'T1 即將上穿' not in a.get('tag', '')]
        imm_bull = [a for a in all_alerts if a.get('level') == 'imm_bull']
        imm_bear = [a for a in all_alerts if a.get('level') == 'imm_bear']

        total_strong = len(bull5) + len(bull4) + len(bull3) + len(bear4) + len(bear3) + len(bear2) + len(bull2_other)
        total_imm = len(imm_bull) + len(imm_bear) + len(bull2_t1)
        if total_strong == 0 and total_imm == 0:
            return

        # 🐛 v9.20.7：用戶要求移除「警報中 — 強訊號 N 檔｜即將觸發 M 檔」總覽 banner
        # 但保留下方分組詳細列表（強看多/強看空/即將觸發）

        def _row_html(a):
            color = '#3dbb6a' if a.get('side') == 'bull' else '#ff5555'
            stars = ''
            lv = a.get('level')
            if lv == 5: stars = '★★★★★'
            elif lv == 4: stars = '★★★★'
            elif lv == 3: stars = '★★★'
            elif lv == 2: stars = '★★'
            else: stars = '⏰'
            close_v = a.get('close', 0) or 0
            # 🆕 v9.11：顯示 quality_score（priority sweep 證實能 +84% CAGR）
            qs = a.get('quality_score')
            qs_html = (f'<span style="color:#e8a020;font-family:monospace;'
                       f'min-width:55px;font-size:.7rem" title="品質分（高=優先）">'
                       f'Q{qs:+.1f}</span>'
                       if qs is not None else
                       f'<span style="min-width:55px"></span>')
            return (
                f'<div style="display:flex;gap:8px;align-items:center;'
                f'padding:4px 8px;border-bottom:1px solid #2a1830;font-size:.78rem">'
                f'<span style="font-weight:700;font-family:monospace;'
                f'min-width:60px;color:{color}">{a.get("market", "")} {a.get("ticker", "?")}</span>'
                f'<span style="color:#a8cce8;flex:1;overflow:hidden;'
                f'text-overflow:ellipsis;max-width:140px">{a.get("name", "")[:14]}</span>'
                f'<span style="color:#e8f4fd;font-family:monospace;min-width:60px">{close_v:.2f}</span>'
                f'<span style="color:#ffd700;font-family:monospace;min-width:60px">{stars}</span>'
                + qs_html +
                f'<span style="color:{color};flex:1.5;font-size:.72rem">{a.get("tag", "")}</span>'
                f'<span style="color:#7a8899;font-size:.7rem;flex:1">{a.get("expect", "")}</span>'
                f'</div>'
            )

        cols = st.columns(2)
        strong_groups = [
            ('🚀🚀 強看多 ★★★★★', bull5, '#3dbb6a'),
            ('🚀 強看多 ★★★★', bull4, '#3dbb6a'),
            ('⚡ 中強看多 ★★★', bull3, '#7abadd'),
            ('🚨🚨 強看空 ★★★★', bear4, '#ff5555'),
            ('🚨 強看空 ★★★', bear3, '#ff5555'),
            ('⚠️ 中度看空 ★★', bear2, '#e8a020'),
        ]
        with cols[0]:
            st.markdown(
                f'<div style="font-weight:700;color:#ffd700;margin-bottom:4px;'
                f'font-size:.9rem">🎯 強警報觸發中</div>',
                unsafe_allow_html=True)
            has_strong = False
            for label, items, color in strong_groups:
                if not items: continue
                has_strong = True
                st.markdown(
                    f'<div style="color:{color};font-size:.8rem;font-weight:700;'
                    f'margin-top:6px">{label} ({len(items)})</div>'
                    + ''.join(_row_html(a) for a in items[:8]),
                    unsafe_allow_html=True
                )
            if not has_strong:
                st.markdown(
                    f'<div style="color:#7a8899;font-size:.78rem;'
                    f'padding:10px">— 暫無強警報 —</div>',
                    unsafe_allow_html=True
                )

        with cols[1]:
            st.markdown(
                f'<div style="font-weight:700;color:#e8a020;margin-bottom:4px;'
                f'font-size:.9rem">⏰ 即將觸發（離條件 1 步）</div>',
                unsafe_allow_html=True)
            # 🆕 v9.12：T1 即將上穿獨立分組（置頂、藍色標示）
            if bull2_t1:
                st.markdown(
                    f'<div style="color:#5a9acf;font-size:.8rem;font-weight:700;'
                    f'margin-top:6px;background:#0a1828;padding:3px 6px;'
                    f'border-left:3px solid #5a9acf;border-radius:3px">'
                    f'🎯 T1 即將上穿 watchlist ({len(bull2_t1)})'
                    f'<span style="color:#7a8899;font-weight:400;font-size:.7rem">'
                    f'  距 EMA20 ≤ 1% + 連 2 漲 + ADX≥22 + 多頭</span></div>'
                    + ''.join(_row_html(a) for a in bull2_t1[:15]),
                    unsafe_allow_html=True
                )
            if imm_bull:
                st.markdown(
                    f'<div style="color:#3dbb6a;font-size:.8rem;font-weight:700;'
                    f'margin-top:6px">🌱 K 線即將看多 ({len(imm_bull)})</div>'
                    + ''.join(_row_html(a) for a in imm_bull[:10]),
                    unsafe_allow_html=True
                )
            if imm_bear:
                st.markdown(
                    f'<div style="color:#ff7755;font-size:.8rem;font-weight:700;'
                    f'margin-top:6px">⚠️ K 線即將看空 ({len(imm_bear)})</div>'
                    + ''.join(_row_html(a) for a in imm_bear[:10]),
                    unsafe_allow_html=True
                )
            if not imm_bull and not imm_bear and not bull2_t1:
                st.markdown(
                    f'<div style="color:#7a8899;font-size:.78rem;'
                    f'padding:10px">— 暫無即將觸發 —</div>',
                    unsafe_allow_html=True
                )
    except Exception:
        pass


# 🐛 v9.20.9：用戶要求移除警報中 panel（強警報觸發中 + 即將觸發）
# _render_alerts_panel()


def _render_screener_panel():
    """🆕 v9.13：全市場篩選器（下拉式選單，37 種驗證過的指標組合）"""
    try:
        from pathlib import Path as _P
        import json as _json
        from screener_filters import FILTERS
        import datetime as _dt

        # 🆕 v9.13：偵測資料新鮮度（修 timezone bug — JSON 是台北時間，server 可能 UTC）
        screener_json = _P(__file__).parent / 'screener_results.json'
        data_freshness = ''
        data_color = '#7a8899'
        data_warning = ''
        if screener_json.exists():
            try:
                d_freshness = _json.loads(screener_json.read_text(encoding='utf-8'))
                computed_at_str = d_freshness.get('computed_at', '?')
                # 計算距今多久（用台北時區一致比對）
                try:
                    computed_dt = _dt.datetime.strptime(computed_at_str, '%Y-%m-%d %H:%M:%S')
                    # 🆕 Streamlit Cloud server 多為 UTC，JSON 的 computed_at 多為台北時間
                    # 統一用台北時間比對：UTC + 8h = 台北
                    now_taipei = _dt.datetime.utcnow() + _dt.timedelta(hours=8)
                    delta = now_taipei - computed_dt
                    hours = delta.total_seconds() / 3600
                    # 若是負值（server 已是台北時區）→ 用 datetime.now()
                    if hours < -1:
                        delta = _dt.datetime.now() - computed_dt
                        hours = delta.total_seconds() / 3600
                    # 取絕對值避免顯示負值
                    hours = abs(hours)
                    if hours < 6:
                        data_color = '#3dbb6a'  # 綠 — 新鮮
                        freshness_label = f'✓ 新鮮（{hours:.1f}h 前）'
                    elif hours < 24:
                        data_color = '#7abadd'  # 藍 — 尚可
                        freshness_label = f'△ 尚可（{hours:.1f}h 前）'
                    elif hours < 48:
                        data_color = '#e8a020'  # 橘 — 偏舊
                        freshness_label = f'⚠️ 偏舊（{hours/24:.1f}d 前）'
                        data_warning = ' — 建議手動觸發 cron 取得新資料'
                    else:
                        data_color = '#ff5555'  # 紅 — 過時
                        freshness_label = f'🚨 過時（{hours/24:.1f}d 前）'
                        data_warning = ' — 資料嚴重過時，請立即手動觸發 cron'
                    data_freshness = (
                        f'<span style="color:{data_color};font-weight:700;font-size:.78rem">{freshness_label}</span>'
                        f'<span style="color:#7a8899;font-size:.72rem"> ｜ '
                        f'資料時間：{computed_at_str}{data_warning}</span>'
                    )
                except Exception:
                    data_freshness = (
                        f'<span style="color:#7a8899;font-size:.72rem">資料時間：{computed_at_str}</span>'
                    )
            except Exception:
                pass

        st.markdown(
            f'<div style="background:#0f1830;border:2px solid #5a8ab0;'
            f'border-radius:10px;padding:10px 14px;margin:12px 0">'
            f'<div style="font-size:1.05rem;font-weight:700;color:#7abadd;'
            f'margin-bottom:4px">'
            f'🔍 全市場篩選器 — {len(FILTERS)} 種驗證過的指標組合'
            f'</div>'
            f'<div style="font-size:.7rem;color:#a8c8d8;line-height:1.5;margin-bottom:6px">'
            f'掃描全部 1925 TW + 2254 US，用驗證過的 alpha 訊號篩選'
            f'</div>'
            # 🆕 資料新鮮度 banner
            f'<div style="background:#0a1422;border-left:3px solid {data_color};'
            f'padding:5px 10px;border-radius:4px;margin-top:6px">'
            f'📅 {data_freshness}'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        # UI — 🆕 v9.13：多條件 multiselect + 排除條件
        st.markdown(
            '<div style="font-size:.75rem;color:#3dbb6a;margin:6px 0 2px;font-weight:700">'
            '✅ 包含條件（多選 = 全部都要符合 / 改 OR 則任一即可）</div>',
            unsafe_allow_html=True)
        filter_names = st.multiselect(
            '包含條件',
            options=list(FILTERS.keys()),
            key='screener_filter',
            label_visibility='collapsed',
            help='多選 = AND 邏輯（要全部符合）。例：倒鎚 + 多頭排列 + 強趨勢 → 三個都符合的股票才出現',
            placeholder='選擇至少一個條件...'
        )
        st.markdown(
            '<div style="font-size:.75rem;color:#ff7755;margin:8px 0 2px;font-weight:700">'
            '❌ 排除條件（選了的條件不能符合 — 任一符合就剔除）</div>',
            unsafe_allow_html=True)
        exclude_names = st.multiselect(
            '排除條件',
            options=list(FILTERS.keys()),
            key='screener_exclude',
            label_visibility='collapsed',
            help='選擇要排除的條件。例：找看多但「不要即將死叉」的，把「即將死叉警告」加到這裡',
            placeholder='（可選 — 選擇要排除的條件）'
        )
        # 邏輯選擇
        cols = st.columns([1, 1, 1, 1])
        with cols[0]:
            logic_mode = st.radio(
                '邏輯', options=['AND（全部符合）', 'OR（任一符合）'],
                key='screener_logic', horizontal=True,
                label_visibility='collapsed'
            )
            logic = 'AND' if 'AND' in logic_mode else 'OR'
        with cols[1]:
            scan_market = st.selectbox(
                '市場',
                options=['全部 (TW+US)', '只 TW 台股', '只 US 美股'],
                key='screener_market',
                label_visibility='collapsed'
            )
        with cols[2]:
            run_scan = st.button(
                '🚀 開始篩選',
                key='screener_run',
                use_container_width=True
            )
        with cols[3]:
            limit_n = st.number_input(
                '上限', min_value=10, max_value=200, value=50, step=10,
                key='screener_limit',
                label_visibility='collapsed'
            )

        # 執行篩選
        if run_scan:
            if not filter_names:
                st.warning('⚠️ 請至少選擇一個篩選條件')
                return

            from screener_filters import filter_universe, intersect_from_json
            DATA = _P(__file__).parent / 'data_cache'
            screener_json = _P(__file__).parent / 'screener_results.json'

            # 🆕 v9.13：明確的市場篩選邏輯
            want_tw = ('全部' in scan_market) or ('只 TW' in scan_market) or ('TW only' in scan_market)
            want_us = ('全部' in scan_market) or ('只 US' in scan_market) or ('US only' in scan_market)

            tw_uni, us_uni = [], []
            if DATA.exists():
                if want_tw:
                    tw_uni = sorted([p.stem for p in DATA.glob('*.parquet')
                                      if p.stem and p.stem[0].isdigit() and len(p.stem) == 4
                                      and not p.stem.startswith('00')])
                if want_us:
                    US_ETF_EX = {'SPY','QQQ','IWM','DIA','VOO','VTI','VEA','VWO','BND','TLT','EFA','AGG',
                                 'LQD','HYG','GLD','SLV','USO','UNG','UCO','SCO','EEM','EWJ','EWZ','EWY',
                                 'FXI','MCHI','XLK','XLF','XLV','XLE','XLY','XLP','XLI','XLU','XLB','XLC',
                                 'SMH','SOXX','IBB','TQQQ','SQQQ','SOXL','SOXS','UPRO','SPXU','VXX','UVXY',
                                 'ARKK','ARKG','ARKF','ARKW','ARKQ'}
                    us_uni = sorted([p.stem for p in DATA.glob('*.parquet')
                                      if p.stem and p.stem.isalpha() and p.stem.isupper()
                                      and 1 <= len(p.stem) <= 5 and p.stem not in US_ETF_EX])

            total_uni = len(tw_uni) + len(us_uni)
            # 顯示用標籤
            inc_label = (filter_names[0] if len(filter_names) == 1
                         else f'{len(filter_names)} 個包含條件 ({logic})')
            exc_label = (f' / ❌ 排除 {len(exclude_names)} 條件'
                          if exclude_names else '')
            display_filter = inc_label + exc_label

            # 🆕 雲端模式：data_cache 沒檔 → 讀 JSON + 多 filter 交集 + 排除
            if total_uni == 0:
                if screener_json.exists():
                    try:
                        d = _json.loads(screener_json.read_text(encoding='utf-8'))
                        all_results = intersect_from_json(d.get('by_filter', {}),
                                                            filter_names, logic=logic,
                                                            exclude_names=exclude_names)
                        # 篩選市場
                        if want_tw and not want_us:
                            all_results = [r for r in all_results if r.get('market') == 'tw']
                        elif want_us and not want_tw:
                            all_results = [r for r in all_results if r.get('market') == 'us']
                        tw_n = sum(1 for r in all_results if r.get('market') == 'tw')
                        us_n = sum(1 for r in all_results if r.get('market') == 'us')
                        excl_msg = f'｜❌排除 {len(exclude_names)} 條件' if exclude_names else ''
                        st.success(f'✅ 雲端模式 — {len(filter_names)} 包含 {logic}{excl_msg} → '
                                    f'**{len(all_results)} 檔**'
                                    f'（🇹🇼 {tw_n} / 🇺🇸 {us_n}）｜資料時間：{d.get("computed_at", "?")}')
                        st.session_state['screener_results'] = all_results
                        st.session_state['screener_last_filter'] = display_filter
                        st.rerun()
                    except Exception as e:
                        st.error(f'❌ 讀 screener_results.json 失敗：{type(e).__name__}: {e}')
                        return
                else:
                    st.error(
                        f'❌ data_cache 為空 + screener_results.json 不存在。\n\n'
                        f'解法：\n'
                        f'• 等下一次 cron 自動跑（每天 09:00 台北）\n'
                        f'• 手動觸發 weekly_full_scan workflow\n'
                        f'• 或本地跑：`python screener_full_local.py`'
                    )
                    return

            # Local 模式：即時跑 filter（多條件）
            universes = []
            if tw_uni: universes.append(('tw', tw_uni))
            if us_uni: universes.append(('us', us_uni))
            st.info(f'📊 即時掃描: {", ".join(f"{m}: {len(u)} 檔" for m, u in universes)}'
                    f' ｜ {len(filter_names)} 條件 {logic}')

            with st.spinner(f'掃描 {total_uni} 檔... ({len(filter_names)} 包含 {logic} + {len(exclude_names)} 排除)'):
                import time as _t
                t0 = _t.time()
                all_results = []
                for market, uni in universes:
                    res = filter_universe(uni, market, filter_names, logic=logic,
                                            exclude_names=exclude_names)
                    all_results.extend(res)
                elapsed = _t.time() - t0
            tw_n = sum(1 for r in all_results if r.get('market') == 'tw')
            us_n = sum(1 for r in all_results if r.get('market') == 'us')
            excl_msg = f'｜❌排除 {len(exclude_names)} 條件' if exclude_names else ''
            st.success(f'✅ 完成 ({elapsed:.1f}s) — 找到 **{len(all_results)} 檔**'
                        f'（🇹🇼 {tw_n} / 🇺🇸 {us_n}）{excl_msg}')

            # 載入 name map
            name_map = {}
            if (_P(__file__).parent / 'tw_stock_list.json').exists():
                try:
                    d = _json.loads((_P(__file__).parent / 'tw_stock_list.json').read_text(encoding='utf-8'))
                    for t, info in d.items():
                        if isinstance(info, dict):
                            name_map[t] = info.get('name', '')
                except Exception: pass
            if (_P(__file__).parent / 'us_full_tickers.json').exists():
                try:
                    full = _json.loads((_P(__file__).parent / 'us_full_tickers.json').read_text(encoding='utf-8'))
                    for x in full.get('detail', []):
                        sym = x.get('symbol', '')
                        nm = x.get('name', '')
                        for sep in [' - ', ' Common ', ' Class ', ' Ordinary ']:
                            if sep in nm:
                                nm = nm.split(sep)[0]; break
                        if sym: name_map[sym] = nm[:40]
                except Exception: pass
            for r in all_results:
                r['name'] = name_map.get(r['ticker'], '')

            # 存進 session state
            st.session_state['screener_results'] = all_results
            st.session_state['screener_last_filter'] = filter_name

        # 顯示既有結果（即使沒按按鈕，也顯示上次結果）
        results = st.session_state.get('screener_results', [])
        last_filter = st.session_state.get('screener_last_filter', '')
        if results:
            # 🆕 v9.15：算 JSON 距今經過幾個交易日（用於 cross_days 推算）
            # 修正：JSON 早上 07:03 跑（用昨日收盤）+ 用戶下午看（今日已收盤）= 1 個交易日已過
            #       但 calendar 不到 24h，舊邏輯 int(10h/24)=0 會誤判 0 天
            json_age_days = 0
            try:
                if screener_json.exists():
                    d_age = _json.loads(screener_json.read_text(encoding='utf-8'))
                    cat_str = d_age.get('computed_at', '')
                    if cat_str:
                        cat_dt = _dt.datetime.strptime(cat_str, '%Y-%m-%d %H:%M:%S')
                        now_taipei = _dt.datetime.utcnow() + _dt.timedelta(hours=8)
                        # 🆕 計算經過的「TW 交易 session 結束」次數
                        # TW 收盤 13:30 Taipei，週末跳過
                        sessions = 0
                        check = cat_dt
                        # 跳到 cat_dt 之後最近的 13:30
                        while check < now_taipei:
                            next_close = check.replace(hour=13, minute=30, second=0, microsecond=0)
                            if next_close <= check:
                                next_close += _dt.timedelta(days=1)
                            # 跳過週末
                            while next_close.weekday() >= 5:
                                next_close += _dt.timedelta(days=1)
                            if next_close <= now_taipei:
                                sessions += 1
                                check = next_close
                            else:
                                break
                        json_age_days = sessions
            except Exception: pass

            # 排序：imminent_dc 後排（潛在風險）
            def _sort_key(r):
                return (1 if r.get('imminent_dc') else 0, r.get('rsi') or 99)
            results = sorted(results, key=_sort_key)

            stale_note = (f' ⚠️ 資料 {json_age_days} 個交易日前（今日 cd 比 snapshot 多 {json_age_days}）'
                           if json_age_days >= 2 else
                           f' 🔄 JSON 早於今日收盤，今日 cd 比 snapshot 多 1（看「→今 +Xd」欄）'
                           if json_age_days == 1 else
                           f' ✨ JSON 今日剛跑（snapshot = 今日 cd）'
                           if json_age_days == 0 else '')
            # 🆕 v9.20.3：SEPA / VCP 相關 filter 顯示 RS vs RSI 提示
            _is_sepa_filter = last_filter and any(
                k in last_filter for k in ['SEPA', 'VCP', 'Minervini', 'RS Rating', 'Pivot'])
            _rs_help = ('<div style="font-size:.65rem;color:#5a8aa0;margin-bottom:4px;'
                         'background:#0a1422;padding:3px 6px;border-radius:3px;'
                         'border-left:2px solid #5dccdd">'
                         '💡 <b>RSI ≠ RS Rating</b>：'
                         'RSI = 相對強弱指標（個股價格動能 0-100，30 超賣 / 70 過熱）；'
                         '<b>RS</b> = 相對強度評分（vs universe 同期百分位，Minervini 建議 ≥70）。'
                         'SEPA / VCP / Minervini 系列濾條看 <b>RS Rating</b>，不是 RSI。'
                         '</div>') if _is_sepa_filter else ''
            st.markdown(
                _rs_help +
                f'<div style="font-weight:700;color:#7abadd;margin-bottom:6px;'
                f'font-size:.9rem">📋 結果（{last_filter}） — {len(results)} 檔'
                f'<span style="color:#e8a020;font-size:.7rem;font-weight:400">{stale_note}</span>'
                f'</div>',
                unsafe_allow_html=True
            )

            # 🆕 v9.25：移除勾選功能，改用純展示表格
            _display_rows = []
            for r in results[:limit_n]:
                flag = '🇹🇼' if r['market'] == 'tw' else '🇺🇸'
                cd = r.get('cross_days')
                if cd is not None and cd > 0:
                    cd_str = (f'+{cd}d→今+{cd+json_age_days}d'
                                if json_age_days >= 1 else f'+{cd}d')
                elif cd:
                    cd_str = f'{cd}d'
                else:
                    cd_str = '-'
                rs_v = r.get('rs_rating')
                pctb = r.get('pct_b')
                _beta = r.get('beta_60d')
                _display_rows.append({
                    '市場': flag,
                    '代號': r['ticker'],
                    '名稱': (r.get('name', '') or '')[:14],
                    '現價': r['close'],
                    '多空': '🟢' if r.get('is_bull') else '🔴',
                    'RSI':  round(r['rsi'], 1) if r.get('rsi') is not None else None,
                    'RS':   round(rs_v, 1) if rs_v is not None else None,
                    'β':    round(_beta, 2) if _beta is not None else None,
                    'ADX':  round(r['adx'], 1) if r.get('adx') is not None else None,
                    'cross': cd_str,
                    '%B':   round(pctb, 2) if pctb is not None else None,
                    '距高': r.get('from_high', 0),
                    'DC警示': '⚠️' if r.get('imminent_dc') else '',
                })
            _df_results = pd.DataFrame(_display_rows)

            st.dataframe(
                _df_results,
                hide_index=True,
                use_container_width=True,
                column_config={
                    '代號': st.column_config.TextColumn('代號', width='small'),
                    '名稱': st.column_config.TextColumn('名稱', width='medium'),
                    '現價': st.column_config.NumberColumn('現價', format='%.2f', width='small'),
                    'RSI': st.column_config.NumberColumn('RSI', help='相對強弱指標（個股動能）', width='small'),
                    'RS':  st.column_config.NumberColumn('RS', help='RS Rating（vs universe 百分位 — Minervini ≥70）', width='small'),
                    'β':   st.column_config.NumberColumn('β', help='60 日 Beta（vs ^GSPC / ^TWII）— >1.5 高 Beta，<0.7 防禦', width='small', format='%.2f'),
                    'ADX': st.column_config.NumberColumn('ADX', width='small'),
                    '%B':  st.column_config.NumberColumn('%B', format='%.2f', width='small'),
                    '距高': st.column_config.NumberColumn('距高%', format='%.0f', width='small'),
                },
                height=min(600, max(150, len(_display_rows) * 35 + 50)),
            )

            # 🆕 v9.25.1：儲存「整個篩選結果」為自選股（不需勾選 — 直接全部）
            _all_tickers = [r['代號'] for r in _display_rows]
            n_all = len(_all_tickers)
            st.markdown(
                f'<div style="font-size:.78rem;color:#7abadd;margin-top:6px">'
                f'💾 把這 <b style="color:#3dbb6a">{n_all}</b> 檔結果全部存成自選股清單'
                f'</div>',
                unsafe_allow_html=True
            )
            _scols = st.columns([3, 1, 1])
            today_str = pd.Timestamp.now().strftime('%Y%m%d')
            _filter_label = (filter_names[0] if len(filter_names) == 1
                              else f'{len(filter_names)}條件')
            default_name = f'篩選_{_filter_label[:12].strip()}_{today_str}'.replace(' ', '_')
            with _scols[0]:
                save_name = st.text_input(
                    '清單名稱', value=default_name,
                    label_visibility='collapsed', key='screener_save_name',
                    placeholder='清單名稱'
                )
            with _scols[1]:
                save_btn = st.button(f'💾 存 ({n_all})',
                                       key='screener_save_btn',
                                       use_container_width=True,
                                       disabled=(n_all == 0),
                                       help='覆寫此清單名稱')
            with _scols[2]:
                add_btn = st.button(f'➕ 加入 ({n_all})',
                                      key='screener_add_btn',
                                      use_container_width=True,
                                      disabled=(n_all == 0),
                                      help='合併進已存在的清單')

            if save_btn or add_btn:
                if not save_name.strip():
                    st.warning('請輸入清單名稱')
                else:
                    tickers_text = '\n'.join(_all_tickers)
                    _wl_path = _P(__file__).parent / 'watchlists.json'
                    try:
                        from streamlit_local_storage import LocalStorage
                        _ls = LocalStorage()
                    except Exception:
                        _ls = None
                    wls = {}
                    if _ls is not None:
                        try:
                            v = _ls.getItem('stock001_watchlists')
                            if v:
                                wls = _json.loads(v) if isinstance(v, str) else v
                        except Exception: pass
                    if not wls and _wl_path.exists():
                        try: wls = _json.loads(_wl_path.read_text(encoding='utf-8'))
                        except Exception: wls = {}

                    name = save_name.strip()
                    if add_btn and name in wls:
                        existing = (set(wls[name].split('\n'))
                                     if isinstance(wls[name], str) else set())
                        new_set = existing | set(_all_tickers)
                        wls[name] = '\n'.join(sorted(new_set))
                        msg = f'➕ 已合併進「{name}」（共 {len(new_set)} 檔）'
                    else:
                        wls[name] = tickers_text
                        msg = f'💾 已儲存「{name}」（{n_all} 檔）'

                    text = _json.dumps(wls, ensure_ascii=False, indent=2)
                    if _ls:
                        try: _ls.setItem('stock001_watchlists', text)
                        except: pass
                    try: _wl_path.write_text(text, encoding='utf-8')
                    except: pass
                    st.success(msg)
                    st.rerun()
    except Exception as e:
        try:
            st.markdown(
                f'<div style="color:#ff7755;font-size:.7rem">'
                f'篩選器 panel 載入失敗：{type(e).__name__}: {str(e)[:100]}</div>',
                unsafe_allow_html=True)
            import traceback as _tb
            with st.expander("traceback"):
                st.code(_tb.format_exc())
        except: pass


_render_screener_panel()


def _render_rs_leading_high_panel():
    """🆕 v9.24：紫色點訊號專屬 panel（RS 領先創新高）

    讀 screener_results.json 的 by_filter，列出通過 RS Leading High 的標的
    """
    try:
        from pathlib import Path as _P
        import json as _json
        sj = _P(__file__).parent / 'screener_results.json'
        if not sj.exists():
            return

        d = _json.loads(sj.read_text(encoding='utf-8'))
        bf = d.get('by_filter', {})
        # 從所有 RS Leading High 變體 filter 中蒐集 ticker 集合
        rs_filter_keys = [k for k in bf.keys() if k.startswith('🟣 RS 領先創新高')]
        if not rs_filter_keys:
            return

        # 用「任何分數」filter 做總清單（最寬鬆的）
        main_key = next((k for k in rs_filter_keys if '任何分數' in k), rs_filter_keys[0])
        items = bf.get(main_key, [])
        if not items:
            return

        # 按 rs_leading_high_score DESC 排序
        items_sorted = sorted(items,
                                key=lambda x: -(x.get('rs_leading_high_score') or 0))

        st.markdown(
            f'<div style="margin-top:18px;padding:12px 14px;'
            f'background:linear-gradient(90deg,#2a1042,#1a0828);'
            f'border-left:4px solid #b266ff;border-radius:8px">'
            f'<div style="font-size:1rem;font-weight:700;color:#d99fff">'
            f'🟣 RS 領先創新高訊號（紫色點）'
            f'</div>'
            f'<div style="font-size:.75rem;color:#a87acc;margin-top:2px">'
            f'TraderLion / William O\'Neil / Mark Minervini 視為機構累積足跡 —'
            f' RS 線創新高但股價未創新高 = 機構安靜累積中'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        # Summary
        tw_n = sum(1 for i in items_sorted if i.get('market') == 'tw')
        us_n = sum(1 for i in items_sorted if i.get('market') == 'us')
        themed = [i for i in items_sorted if i.get('rs_leading_high_theme')]
        high_quality = [i for i in items_sorted
                         if (i.get('rs_leading_high_score') or 0) >= 60]
        dense_purple = [i for i in items_sorted
                         if (i.get('rs_leading_high_purple_dots') or 0) >= 5]

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric('總訊號', f'{len(items_sorted)}',
                       help=f'TW {tw_n} / US {us_n}')
        with c2:
            st.metric('高品質 (≥60 分)', f'{len(high_quality)}')
        with c3:
            st.metric('紫色點密集 (≥5)', f'{len(dense_purple)}')
        with c4:
            st.metric('Eddy 主題', f'{len(themed)}')

        # 主表格
        import pandas as _pd
        rows = []
        for i, item in enumerate(items_sorted, 1):
            rows.append({
                '#': i,
                'Ticker': item.get('ticker'),
                'Name': item.get('name', '')[:20],
                'Market': '🇺🇸' if item.get('market') == 'us' else '🇹🇼',
                'Score': item.get('rs_leading_high_score'),
                'Close': item.get('close'),
                'DistHi%': item.get('rs_leading_high_distance'),
                'PurpD': item.get('rs_leading_high_purple_dots'),
                'RS Rating': item.get('rs_rating'),
                'Theme': item.get('rs_leading_high_theme', ''),
            })
        df_rs = _pd.DataFrame(rows)

        st.dataframe(
            df_rs,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Score': st.column_config.ProgressColumn(
                    'Score', min_value=0, max_value=100, format='%.1f',
                    help='品質分數 0-100，>60 = 高品質'),
                'PurpD': st.column_config.NumberColumn(
                    'PurpD', help='近 20 日 RS 創新高次數（紫色點密度）'),
                'DistHi%': st.column_config.NumberColumn(
                    'DistHi%', format='%.1f%%',
                    help='股價距 63 日高點百分比；5-15% 區間最佳'),
                'Theme': st.column_config.TextColumn(
                    'Theme', help='Eddy 關注主題（AI 儲存 / AI 能源）')
            }
        )

        # 操作提示
        st.markdown(
            f'<div style="font-size:.72rem;color:#7a8899;margin-top:4px">'
            f'💡 點 Ticker 進入個股 detail card 查看完整 ZigZag 對照圖、'
            f'動能衰減、3 段建倉等資訊。OOS 回測 21d 勝率 78%、Sharpe 1.71'
            f'（樣本 SP500 50 檔 × 2023-2024，p=0.20 統計顯著性未達；'
            f'建議作為 T3 子集輸入而非獨立策略）'
            f'</div>',
            unsafe_allow_html=True
        )
    except Exception as e:
        try:
            st.markdown(f'RS Leading High panel 載入失敗: {type(e).__name__}')
        except Exception:
            pass


_render_rs_leading_high_panel()


def _render_sympathy_panel():
    """🆕 v9.25.3：補漲候選股 panel（Sympathy Play）

    讀 sympathy_latest.json 顯示當日 leader + 補漲候選清單
    """
    try:
        from pathlib import Path as _P
        import json as _json
        sp = _P(__file__).parent / 'sympathy_latest.json'
        if not sp.exists():
            return

        d = _json.loads(sp.read_text(encoding='utf-8'))
        leaders = d.get('leaders', [])
        candidates = d.get('candidates', [])
        date_str = d.get('date', '?')

        if not leaders and not candidates:
            return

        # Header
        st.markdown(
            f'<div style="margin-top:18px;padding:12px 14px;'
            f'background:linear-gradient(90deg,#1a2810,#0a2010);'
            f'border-left:4px solid #66ff99;border-radius:8px">'
            f'<div style="font-size:1rem;font-weight:700;color:#9fff9f">'
            f'🚀 補漲候選股 — Sympathy Play ({date_str})'
            f'</div>'
            f'<div style="font-size:.75rem;color:#7acc7a;margin-top:2px">'
            f'同族群中相關性高、漲幅落後的個股 — 機構買盤 leader 大漲後，'
            f'1-5 個交易日內常出現補漲。OOS 12 月回測：勝率 72.7%、Sharpe 1.7（score≥0.75 勝率 85.7%）'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        # Leaders 區塊
        if leaders:
            st.markdown(f'**今日 Leaders（{len(leaders)} 檔）**')
            import pandas as _pd
            ld_rows = []
            for ld in leaders:
                flag = '🇹🇼' if '.TW' in ld['ticker'] else '🇺🇸'
                ld_rows.append({
                    'Market': flag,
                    'Ticker': ld['ticker'],
                    'Group': ld['group'],
                    'Ret%': round(ld['return_pct'] * 100, 2),
                    'VolR': ld['volume_ratio'],
                    'Close': ld['close'],
                })
            st.dataframe(_pd.DataFrame(ld_rows), hide_index=True,
                          use_container_width=True,
                          column_config={
                              'Ret%': st.column_config.NumberColumn('Ret%', format='%+.2f%%'),
                              'VolR': st.column_config.NumberColumn('VolR', format='%.2fx'),
                              'Close': st.column_config.NumberColumn('Close', format='%.2f'),
                          })

        # Candidates 區塊
        if candidates:
            high_n = sum(1 for c in candidates if c['score'] >= 0.75)
            theme_n = sum(1 for c in candidates if 'AI_' in (c.get('group') or ''))
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric('總候選', f'{len(candidates)}')
            with c2:
                st.metric('高品質 (score≥0.75)', f'{high_n}',
                          help='OOS 該區間勝率 85.7%')
            with c3:
                st.metric('AI 主題', f'{theme_n}')

            cand_rows = []
            for i, c in enumerate(candidates, 1):
                flag = '🇹🇼' if '.TW' in c['ticker'] else '🇺🇸'
                cand_rows.append({
                    '#': i,
                    'Market': flag,
                    'Ticker': c['ticker'],
                    'Group': c.get('group', ''),
                    'Leader': c['leader'],
                    'Score': c['score'],
                    'Corr60d': c['corr_60d'],
                    'SprdPct': c['spread_pctile'],
                    'Lag%': round(c['lag_today'] * 100, 2),
                })
            df_cand = _pd.DataFrame(cand_rows)
            st.dataframe(
                df_cand,
                hide_index=True,
                use_container_width=True,
                column_config={
                    'Score': st.column_config.ProgressColumn(
                        'Score', min_value=0, max_value=1,
                        format='%.3f',
                        help='≥0.75 強訊號，≥0.6 中，≥0.45 弱'),
                    'Corr60d': st.column_config.NumberColumn(
                        'Corr60d', format='%.3f', help='與 leader 60 日相關性'),
                    'SprdPct': st.column_config.NumberColumn(
                        'SprdPct', format='%.3f',
                        help='peer/leader 價格比百分位（越低越落後）'),
                    'Lag%': st.column_config.NumberColumn(
                        'Lag%', format='%+.2f%%', help='今日落後 leader 幅度'),
                }
            )

            st.markdown(
                f'<div style="font-size:.72rem;color:#7a8899;margin-top:4px">'
                f'💡 訊號有效期 5 個交易日；持有規則：+8% TP / -4% SL / 5d 收盤出'
                f'</div>',
                unsafe_allow_html=True
            )
    except Exception:
        pass


_render_sympathy_panel()


def _render_hit_rate_panel():
    """🆕 v9.11：Live 命中率追蹤 — 從 alert_history.json 顯示 5/15/30d 命中率"""
    try:
        from pathlib import Path as _P
        import json as _json
        p = _P(__file__).parent / 'alert_history.json'
        if not p.exists():
            return  # 還沒有歷史資料，不顯示

        h = _json.loads(p.read_text(encoding='utf-8'))
        alerts = h.get('alerts', [])
        stats = h.get('stats', {})
        total = len(alerts)
        if total == 0:
            return

        last_update = h.get('last_outcomes_update', h.get('last_updated', ''))

        # group key → 中文標籤
        label_map = {
            'tw_5_bull':            '🇹🇼 ★★★★★ 看多 (倒鎚+RSI≤25+ADX↑)',
            'tw_4_bull':            '🇹🇼 ★★★★ 看多 (倒鎚+跌深)',
            'tw_3_bull':            '🇹🇼 ★★★ 看多 (底十字+RSI≤25)',
            'tw_4_bear':            '🇹🇼 ★★★★ 看空 (三烏+距高<5%+量縮)',
            'tw_3_bear':            '🇹🇼 ★★★ 看空 (空頭吞噬+RSI≥75)',
            'tw_2_bear':            '🇹🇼 ★★ 看空 (黃昏星+RSI≥75)',
            'tw_imm_bull_bull':     '🇹🇼 ⏰ 即將看多',
            'tw_imm_bear_bear':     '🇹🇼 ⏰ 即將看空',
            'us_5_bull':            '🇺🇸 ★★★★★ 看多',
            'us_4_bull':            '🇺🇸 ★★★★ 看多',
            'us_3_bull':            '🇺🇸 ★★★ 看多',
            'us_4_bear':            '🇺🇸 ★★★★ 看空',
            'us_3_bear':            '🇺🇸 ★★★ 看空',
            'us_2_bear':            '🇺🇸 ★★ 看空',
            'us_imm_bull_bull':     '🇺🇸 ⏰ 即將看多',
            'us_imm_bear_bear':     '🇺🇸 ⏰ 即將看空',
        }

        # 已回算的筆數（任一 horizon 有 ret_pct 就算）
        total_with_outcome = sum(
            1 for a in alerts
            if any(a.get('outcomes', {}).get(f'{n}d', {}).get('ret_pct') is not None
                   for n in [5, 15, 30])
        )

        st.markdown(
            f'<div style="background:#0d1c2e;border:2px solid #2a4a70;'
            f'border-radius:10px;padding:10px 14px;margin:12px 0">'
            f'<div style="font-size:1.05rem;font-weight:700;color:#7abadd;'
            f'margin-bottom:4px">'
            f'📈 Live 命中率追蹤 — 累積 {total} 筆 / 已回算 {total_with_outcome} 筆'
            f'</div>'
            f'<div style="font-size:.7rem;color:#7a8899">'
            f'最後回算：{last_update or "（尚未回算）"}'
            f'</div></div>',
            unsafe_allow_html=True
        )

        if not stats:
            st.markdown(
                f'<div style="color:#7a8899;font-size:.78rem;padding:8px 14px">'
                f'仍在累積資料（5/15/30 天後才有結果）</div>',
                unsafe_allow_html=True)
            return

        # 排序：看多由強→弱，看空由強→弱，imm 最後
        order = [
            'tw_5_bull', 'tw_4_bull', 'tw_3_bull',
            'us_5_bull', 'us_4_bull', 'us_3_bull',
            'tw_4_bear', 'tw_3_bear', 'tw_2_bear',
            'us_4_bear', 'us_3_bear', 'us_2_bear',
            'tw_imm_bull_bull', 'tw_imm_bear_bear',
            'us_imm_bull_bull', 'us_imm_bear_bear',
        ]
        sorted_keys = ([k for k in order if k in stats]
                       + [k for k in stats if k not in order])

        rows_html = []
        for key in sorted_keys:
            s = stats[key]
            label = label_map.get(key, key)
            is_bear = ('bear' in key)
            cells = []
            for hd in ['5d', '15d', '30d']:
                v = s.get(hd)
                if v:
                    n_, hr, mr = v['n'], v['hit_rate'], v['mean_ret']
                    # 看多: ret_pct > 0 = 命中（綠）；看空：ret_pct < 0 = 命中（綠）
                    expected_pos = (mr > 0) if not is_bear else (mr < 0)
                    color = '#3dbb6a' if expected_pos else '#ff5555' if mr != 0 else '#7a8899'
                    cells.append(
                        f'<span style="color:{color};font-family:monospace;'
                        f'font-size:.78rem;min-width:130px;display:inline-block">'
                        f'{hd}: <b>{hr:.0f}%</b> n={n_} ({mr:+.1f}%)</span>'
                    )
                else:
                    cells.append(
                        f'<span style="color:#445566;font-family:monospace;'
                        f'font-size:.78rem;min-width:130px;display:inline-block">{hd}: —</span>'
                    )
            rows_html.append(
                f'<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;'
                f'padding:5px 8px;border-bottom:1px solid #1a2a3a">'
                f'<span style="color:#aac4e0;flex:1;min-width:240px;'
                f'font-size:.82rem">{label}</span>'
                + ''.join(cells)
                + f'</div>'
            )

        st.markdown(
            f'<div style="background:#0a1428;border:1px solid #1f3550;'
            f'border-radius:8px;padding:6px 10px;margin:8px 0">'
            + ''.join(rows_html)
            + '</div>',
            unsafe_allow_html=True
        )

        # 警示文字：樣本太少時提醒
        if total_with_outcome < 10:
            st.markdown(
                f'<div style="color:#e8a020;font-size:.7rem;padding:4px 14px">'
                f'⚠️ 樣本數還小（{total_with_outcome} 筆），建議累積到 ≥30 筆再做決策'
                f'</div>',
                unsafe_allow_html=True
            )
    except Exception as e:
        try:
            st.markdown(
                f'<div style="color:#ff7755;font-size:.7rem;padding:4px 14px">'
                f'命中率 panel 載入失敗：{type(e).__name__}</div>',
                unsafe_allow_html=True)
        except:
            pass


_render_hit_rate_panel()


# 🆕 市場廣度警報（D 路徑）
_breadth = _get_market_breadth()
if _breadth.get('has_data'):
    _level_style = {
        'red':      ('#2a0a00', '#ff5555', '#ffb090'),
        'red_bear': ('#1a0a05', '#ff5555', '#ffa090'),
        'yellow':   ('#1a1200', '#f0a030', '#f0c890'),
        'green':    ('#0a1e10', '#3dbb6a', '#a8e0b8'),
    }.get(_breadth['level'], ('#0a1628', '#7aaac8', '#a0b8c8'))
    _bg, _bord, _fg = _level_style
    st.markdown(
        f'<div style="background:{_bg};border:1px solid {_bord}55;'
        f'border-left:3px solid {_bord};border-radius:6px;'
        f'padding:8px 14px;margin-bottom:1rem;font-size:.74rem;'
        f'color:{_fg}"><b style="color:{_bord}">市場廣度</b> '
        f'{_breadth["msg"]}</div>',
        unsafe_allow_html=True
    )

with st.sidebar:
    # ═══════════════════════════════════════════════════════════════
    # § 1. 股票清單（含自選股管理）
    # ═══════════════════════════════════════════════════════════════
    st.markdown(
        '<div style="font-size:.95rem;font-weight:700;color:#e8f4fd;'
        'margin:4px 0 8px;letter-spacing:.04em">📋 股票清單</div>',
        unsafe_allow_html=True
    )

    # 🆕 v9.15.3：持久化訊息（解決 st.rerun() 立刻清掉 st.success() 的問題）
    # 在 rerun 後第一次 render 時顯示上次的訊息
    _pending = st.session_state.pop('_wl_pending_msg', None)
    if _pending:
        _msg, _level = _pending
        if _level == 'success':
            st.success(_msg)
        elif _level == 'error':
            st.error(_msg)
        elif _level == 'warning':
            st.warning(_msg)
        else:
            st.info(_msg)

    # ── 自選股持久化：localStorage（雲端用） + 本地 JSON 雙軌 ──
    import json as _json
    from pathlib import Path as _Path
    _WATCHLIST_FILE = _Path(__file__).parent / 'watchlists.json'

    # localStorage（雲端唯一可持久化選項）
    try:
        from streamlit_local_storage import LocalStorage
        _localS = LocalStorage()
    except Exception:
        _localS = None

    def _load_watchlists() -> dict:
        # ① localStorage 優先（瀏覽器端持久，雲端最可靠）
        if _localS is not None:
            try:
                v = _localS.getItem("stock001_watchlists")
                if v:
                    if isinstance(v, str):
                        return _json.loads(v)
                    if isinstance(v, dict):
                        return v
            except Exception:
                pass
        # ② 檔案 fallback（本地有效）
        if _WATCHLIST_FILE.exists():
            try:
                return _json.loads(_WATCHLIST_FILE.read_text(encoding='utf-8'))
            except Exception:
                return {}
        return {}

    def _save_watchlists(d: dict, push_github: bool = False):
        text = _json.dumps(d, ensure_ascii=False, indent=2)
        # ① 寫 localStorage（雲端持久）
        if _localS is not None:
            try:
                _localS.setItem("stock001_watchlists", text)
            except Exception:
                pass
        # ② 寫檔（本地有效；雲端 ephemeral 但不傷）
        try:
            _WATCHLIST_FILE.write_text(text, encoding='utf-8')
        except Exception:
            pass
        # 🆕 v9.14 ③ 推送到 GitHub（需 secrets["GITHUB_TOKEN"]）
        if push_github:
            return _push_watchlists_to_github(d)
        return None

    def _push_watchlists_to_github(d: dict, repo: str = 'zeushuan/Stock001',
                                    file_path: str = 'watchlists_user.json'):
        """🆕 v9.14：透過 GitHub API 推送 watchlists 到 repo
        需要 Streamlit secrets 設 GITHUB_TOKEN（PAT，repo 權限）
        回傳 (success, message)"""
        try:
            token = st.secrets.get('GITHUB_TOKEN', '') if hasattr(st, 'secrets') else ''
            if not token:
                return (False, '⚠️ 未設定 GITHUB_TOKEN secret')
            import base64
            import datetime as _dt_local  # 🆕 v9.15.4：函式內 local 匯入避免 scope 問題
            content_str = _json.dumps(d, ensure_ascii=False, indent=2)
            content_b64 = base64.b64encode(content_str.encode()).decode()
            api_url = f'https://api.github.com/repos/{repo}/contents/{file_path}'
            headers = {
                'Authorization': f'token {token}',
                'Accept': 'application/vnd.github.v3+json',
            }
            # Get existing file SHA (if any)
            r = requests.get(api_url, headers=headers, timeout=10)
            sha = r.json().get('sha') if r.status_code == 200 else None

            payload = {
                'message': f'auto: 更新 watchlists ({_dt_local.datetime.now().strftime("%Y-%m-%d %H:%M")})',
                'content': content_b64,
                'branch': 'main',
            }
            if sha:
                payload['sha'] = sha
            r = requests.put(api_url, headers=headers, json=payload, timeout=15)
            if r.status_code in (200, 201):
                return (True, f'✅ 已推送到 GitHub ({len(d)} 個清單)')
            else:
                return (False, f'❌ GitHub 推送失敗 ({r.status_code}): {r.text[:200]}')
        except Exception as e:
            return (False, f'❌ 推送錯誤：{type(e).__name__}: {e}')

    def _pull_watchlists_from_github(repo: str = 'zeushuan/Stock001',
                                      file_path: str = 'watchlists_user.json'):
        """🆕 v9.14：從 GitHub repo 拉 watchlists（同步至此裝置）"""
        try:
            token = st.secrets.get('GITHUB_TOKEN', '') if hasattr(st, 'secrets') else ''
            url = f'https://raw.githubusercontent.com/{repo}/main/{file_path}'
            headers = {'Authorization': f'token {token}'} if token else {}
            r = requests.get(url, headers=headers, timeout=10)
            if r.status_code == 200 and r.text.strip():
                return _json.loads(r.text)
        except Exception:
            pass
        return None

    # 從 GitHub 讀取預設清單
    GITHUB_LIST_URL = "https://raw.githubusercontent.com/zeushuan/stock001/main/stocks.txt"
    @st.cache_data(ttl=None, show_spinner=False)
    def load_default_stocks() -> str:
        try:
            r = requests.get(GITHUB_LIST_URL, timeout=6)
            if r.status_code == 200 and r.text.strip():
                lines = [l.strip() for l in r.text.splitlines() if l.strip()]
                return "\n".join(lines)
        except Exception:
            pass
        return "DJI\nSPX\n0050\n2330\n00632R\n00737\nBOTZ"

    default_stocks = load_default_stocks()

    # 🆕 v9.14：載入 3 個預設範本清單（持倉 / 半導體 / 高股息）
    _PRESETS_PATH = _Path(__file__).parent / 'watchlists_presets.json'
    @st.cache_data(ttl=None, show_spinner=False)
    def load_presets() -> dict:
        try:
            if _PRESETS_PATH.exists():
                d = _json.loads(_PRESETS_PATH.read_text(encoding='utf-8'))
                return d.get('presets', {})
        except Exception: pass
        return {}
    _presets = load_presets()
    # 預設清單轉 ticker text 格式
    _preset_texts = {
        name: '\n'.join(info.get('tickers', []))
        for name, info in _presets.items()
    }

    # 載入自選股（每次 rerun 都從持久層重新讀，避免 session_state 過時）
    _wls = _load_watchlists()
    st.session_state['watchlists'] = _wls

    # 🆕 自選股下拉：預設清單 + 預設範本（read-only，名稱前加 🔒）+ 使用者自存
    _preset_options = [f'🔒 {name}（範本）' for name in _preset_texts.keys()]
    _user_options = sorted(_wls.keys())
    _wl_options = ["（預設清單）"] + _preset_options + _user_options

    _selected_wl = st.selectbox(
        "自選股清單", options=_wl_options,
        index=0, key="watchlist_select",
        help="🔒 = 範本清單（可載入但無法直接覆寫，請另存新名稱）"
    )

    # 依選擇載入清單內容
    if _selected_wl == "（預設清單）":
        _initial_text = default_stocks
    elif _selected_wl.startswith('🔒 '):
        # 範本清單：去掉 🔒 + （範本）後綴
        _preset_name = _selected_wl[2:].rsplit('（範本）', 1)[0].strip()
        _initial_text = _preset_texts.get(_preset_name, default_stocks)
    else:
        _initial_text = _wls.get(_selected_wl, default_stocks)

    stock_input = st.text_area(
        "代號（每行一檔）",
        value=_initial_text, height=180,
        key=f"stock_input_{_selected_wl}",
        help="台股：純數字或含字母 ETF（2330/00878/00632R）；美股：英文代號（NVDA/SPY）；# 開頭為註解",
    )

    # 儲存 / 刪除（更緊湊的 2 欄佈局）
    _c1, _c2 = st.columns([2, 1])
    with _c1:
        _save_name = st.text_input(
            "新清單名稱",
            placeholder="輸入名稱…",
            label_visibility="collapsed",
            key="wl_save_name"
        )
    with _c2:
        if st.button("💾 存", use_container_width=True, key="wl_save_btn"):
            n = _save_name.strip()
            if not n:
                st.session_state['_wl_pending_msg'] = ("請先輸入清單名稱", 'warning')
                st.rerun()
            elif n.startswith('🔒') or '（範本）' in n:
                st.session_state['_wl_pending_msg'] = ("不能用範本名稱，請改用其他名稱", 'warning')
                st.rerun()
            else:
                _wls[n] = stock_input
                # 🆕 v9.14：勾「同步到 GitHub」就推
                push_ok = st.session_state.get('wl_push_github', False)
                result = _save_watchlists(_wls, push_github=push_ok)
                msg = f"✓ 已存「{n}」"
                if result is not None:
                    success, push_msg = result
                    msg += f" ｜ {push_msg}"
                # 🆕 v9.15.3：用 session_state 保留訊息跨 rerun
                st.session_state['_wl_pending_msg'] = (msg, 'success' if (result is None or result[0]) else 'error')
                st.rerun()

    # 🆕 v9.14：GitHub 同步 + 刪除按鈕
    _wl_user_only = _selected_wl != "（預設清單）" and not _selected_wl.startswith('🔒 ')
    _c3, _c4 = st.columns([2, 1])
    with _c3:
        st.checkbox("🔄 同步到 GitHub (需設 GITHUB_TOKEN secret)",
                     key='wl_push_github',
                     help='勾選後存/刪會透過 GitHub API 自動推送 watchlists_user.json')
    with _c4:
        if st.button("📥 從 GitHub 拉",
                      use_container_width=True, key='wl_pull_btn',
                      help='從 GitHub repo 拉最新 watchlists（覆蓋本地）'):
            remote = _pull_watchlists_from_github()
            if remote:
                _save_watchlists(remote, push_github=False)
                st.session_state['_wl_pending_msg'] = (f"✓ 已拉取 {len(remote)} 個清單", 'success')
            else:
                st.session_state['_wl_pending_msg'] = (
                    "拉取失敗（檢查 GITHUB_TOKEN 或 watchlists_user.json 是否存在）", 'error')
            st.rerun()

    if _wl_user_only:
        if st.button(f"🗑 刪除「{_selected_wl}」",
                     use_container_width=True, key="wl_del_btn"):
            _wls.pop(_selected_wl, None)
            push_ok = st.session_state.get('wl_push_github', False)
            result = _save_watchlists(_wls, push_github=push_ok)
            msg = f"✓ 已刪除「{_selected_wl}」"
            if result is not None:
                success, push_msg = result
                msg += f" ｜ {push_msg}"
            st.session_state['_wl_pending_msg'] = (msg, 'success' if (result is None or result[0]) else 'error')
            st.rerun()

    # ── 🆕 v9.28：個股備註區 ─────────────────────────────────
    with st.expander("📝 個股備註（自行寫筆記）", expanded=False):
        st.caption(
            '為個股寫備註（如開倉成本、停損價、催化劑等），detail card 會自動顯示。'
            '格式：每行一筆「TICKER: 備註內容」'
        )
        # 取當前 ticker_notes
        _notes_dict = (_wls.get('_ticker_notes', {})
                        if isinstance(_wls.get('_ticker_notes'), dict) else {})

        # 把當前清單中的 ticker 預先帶出來
        _current_tickers = []
        try:
            for line in (stock_input or '').split('\n'):
                line = line.split('#', 1)[0].strip()
                if line: _current_tickers.append(line.upper())
        except Exception: pass

        # Compose default text — 既有的 notes 全部 + 當前清單中沒備註的 ticker
        _lines = []
        for tk, note in sorted(_notes_dict.items()):
            _lines.append(f'{tk}: {note}')
        seen = set(_notes_dict.keys())
        for tk in _current_tickers:
            if tk not in seen and tk not in [l.split(':')[0].strip() for l in _lines]:
                _lines.append(f'{tk}: ')   # 空備註 placeholder
        _notes_default = '\n'.join(_lines)

        _notes_input = st.text_area(
            '備註內容',
            value=_notes_default, height=200,
            key='ticker_notes_input',
            help='每行格式：TICKER: 備註內容\n空白備註會被忽略不存',
            label_visibility='collapsed',
        )

        if st.button('💾 存備註', use_container_width=True, key='wl_notes_save'):
            new_notes = {}
            for line in (_notes_input or '').split('\n'):
                if ':' not in line: continue
                tk, _, note = line.partition(':')
                tk = tk.strip().upper()
                note = note.strip()
                if tk and note:   # 兩者都要有
                    new_notes[tk] = note
            _wls['_ticker_notes'] = new_notes
            push_ok = st.session_state.get('wl_push_github', False)
            result = _save_watchlists(_wls, push_github=push_ok)
            msg = f'✓ 已存 {len(new_notes)} 個備註'
            if result is not None:
                msg += f' ｜ {result[1]}'
            st.session_state['_wl_pending_msg'] = (
                msg, 'success' if (result is None or result[0]) else 'error')
            st.rerun()

    # ── 🆕 匯出 / 匯入 JSON 永久備份 ────────────────────────────
    with st.expander("💾 備份 / 還原 自選股清單", expanded=False):
        st.caption(
            "瀏覽器 localStorage 可能因清快取/隱私模式/閒置等原因消失。"
            "建議定期匯出 JSON 永久保存。"
        )
        # 匯出
        if _wls:
            _export_text = _json.dumps(_wls, ensure_ascii=False, indent=2)
            st.download_button(
                "📥 匯出全部清單（JSON）",
                data=_export_text,
                file_name="watchlists_backup.json",
                mime="application/json",
                use_container_width=True,
                key="wl_export_btn",
            )
        else:
            st.caption("（目前無清單可匯出）")
        # 匯入
        _uploaded = st.file_uploader(
            "📤 匯入 JSON 檔還原清單",
            type=['json'], key="wl_import_uploader"
        )
        if _uploaded is not None:
            try:
                _imported = _json.loads(_uploaded.read().decode('utf-8'))
                if isinstance(_imported, dict):
                    _merge_mode = st.radio(
                        "匯入方式",
                        ["合併（保留現有）", "覆蓋（清空再匯入）"],
                        horizontal=True, key="wl_import_mode"
                    )
                    if st.button("✅ 確認匯入", type="primary",
                                 use_container_width=True,
                                 key="wl_import_confirm"):
                        if _merge_mode.startswith("覆蓋"):
                            _wls = _imported
                        else:
                            _wls = {**_imported, **_wls}
                        _save_watchlists(_wls)
                        st.success(f"✓ 已匯入 {len(_imported)} 個清單"
                                   f"（合併後總共 {len(_wls)}）")
                        st.rerun()
                else:
                    st.error("檔案格式錯誤（不是字典結構）")
            except Exception as e:
                st.error(f"解析失敗：{e}")

    if _localS is not None:
        st.caption(
            "💾 localStorage 儲存於瀏覽器（建議定期匯出備份）"
        )

    # ── 🆕 v9.15.2：GitHub 同步診斷工具 ────────────────────────
    with st.expander("🔧 GitHub 同步診斷", expanded=False):
        st.caption("一次檢查 token / repo / 推送狀態，找出為何同步沒成功")

        # 🆕 v9.15.3：診斷訊息持久化（跨 rerun 保留）
        _diag_msg = st.session_state.pop('_gh_diag_msg', None)
        if _diag_msg:
            _msg_t, _level_t = _diag_msg
            if _level_t == 'success':
                st.success(_msg_t)
            elif _level_t == 'error':
                st.error(_msg_t)
            else:
                st.info(_msg_t)

        # ① 檢查 secret
        try:
            _tok = st.secrets.get('GITHUB_TOKEN', '') if hasattr(st, 'secrets') else ''
        except Exception as e:
            _tok = ''
            st.error(f"讀 secrets 失敗：{type(e).__name__}: {e}")

        if _tok:
            _tok_disp = _tok[:7] + '...' + _tok[-4:] if len(_tok) > 11 else 'too short'
            st.markdown(f"**① GITHUB_TOKEN secret**: ✅ 有設定（{_tok_disp}，長度 {len(_tok)}）")
        else:
            st.markdown("**① GITHUB_TOKEN secret**: ❌ **未設定** ← 主要問題！")
            st.markdown(
                "📋 修正步驟：\n"
                "1. GitHub → Settings → Developer settings → Personal access tokens (classic)\n"
                "2. Generate new token，**勾 `repo` scope**\n"
                "3. 複製 `ghp_xxxxx...`\n"
                "4. Streamlit Cloud → app 右上 ⋮ → Settings → **Secrets**\n"
                "5. 加入 `GITHUB_TOKEN = \"ghp_xxxxx...\"`（注意要有引號）\n"
                "6. Save → app 自動重啟（約 30 秒）"
            )

        # ② checkbox 狀態
        _chk_state = st.session_state.get('wl_push_github', False)
        if _chk_state:
            st.markdown("**② 同步 checkbox**: ✅ 已勾選（按存/刪會推 GitHub）")
        else:
            st.markdown(
                "**② 同步 checkbox**: ⚠️ **未勾選** ← 即使有 token 也不會推！"
                "<br>請勾上方「🔄 同步到 GitHub」再按存",
                unsafe_allow_html=True
            )

        # ③ 測試 GitHub API 連線
        if _tok:
            if st.button("🧪 測試 GitHub API 連線", key="gh_test_btn"):
                try:
                    _api = 'https://api.github.com/repos/zeushuan/Stock001'
                    _h = {'Authorization': f'token {_tok}',
                          'Accept': 'application/vnd.github.v3+json'}
                    _r = requests.get(_api, headers=_h, timeout=10)
                    if _r.status_code == 200:
                        _info = _r.json()
                        _diag_text = (
                            f"✅ Repo 可存取：{_info.get('full_name')} "
                            f"（permissions: push={_info.get('permissions', {}).get('push', False)}, "
                            f"admin={_info.get('permissions', {}).get('admin', False)}）"
                        )
                        st.session_state['_gh_diag_msg'] = (_diag_text, 'success')
                    elif _r.status_code == 401:
                        st.session_state['_gh_diag_msg'] = (
                            "❌ 401 Unauthorized — token 無效或過期", 'error')
                    elif _r.status_code == 403:
                        st.session_state['_gh_diag_msg'] = (
                            "❌ 403 Forbidden — token 沒給 `repo` scope", 'error')
                    elif _r.status_code == 404:
                        st.session_state['_gh_diag_msg'] = (
                            "❌ 404 Not Found — repo 名錯了，或 token 無權限看", 'error')
                    else:
                        st.session_state['_gh_diag_msg'] = (
                            f"❌ {_r.status_code}: {_r.text[:300]}", 'error')
                except Exception as e:
                    st.session_state['_gh_diag_msg'] = (
                        f"❌ 連線錯誤：{type(e).__name__}: {e}", 'error')
                st.rerun()

        # ④ 手動推送測試
        if _tok:
            if st.button("🚀 立即推送目前 watchlists 到 GitHub", key="gh_push_now_btn",
                          help="不論 checkbox，立即觸發推送（測試用）"):
                _result = _push_watchlists_to_github(_wls)
                _suc, _msg = _result
                _confirm_text = (
                    "\n\n🔍 確認步驟：\n"
                    "1. 開 https://github.com/zeushuan/Stock001/commits/main\n"
                    "2. 應該有 1 分鐘內的新 commit `auto: 更新 watchlists ...`\n"
                    "3. 點進去看 `watchlists_user.json` 變動"
                ) if _suc else ""
                st.session_state['_gh_diag_msg'] = (
                    _msg + _confirm_text, 'success' if _suc else 'error')
                st.rerun()

        # ⑤ 檢查 GitHub 上目前狀態
        st.markdown("---")
        st.markdown("**③ GitHub 上目前狀態**：")
        _check_url = "https://api.github.com/repos/zeushuan/Stock001/contents/watchlists_user.json"
        try:
            _r2 = requests.get(_check_url, timeout=8)
            if _r2.status_code == 200:
                import base64 as _b64
                _content = _b64.b64decode(_r2.json().get('content', '')).decode('utf-8')
                try:
                    _d_remote = _json.loads(_content)
                    st.success(f"✅ watchlists_user.json 已存在 GitHub（{len(_d_remote)} 個清單）")
                    _names = list(_d_remote.keys())
                    st.caption(f"清單：{', '.join(_names[:8])}{'...' if len(_names) > 8 else ''}")
                    _commit_url = "https://github.com/zeushuan/Stock001/commits/main/watchlists_user.json"
                    st.markdown(f"[📋 看歷次 commit]({_commit_url})")
                except Exception:
                    st.warning("檔案存在但解析失敗")
            elif _r2.status_code == 404:
                st.warning(
                    "⚠️ **GitHub 上尚無 watchlists_user.json**"
                    "（從未推送成功 → 第一次按「🚀 立即推送」會建立）"
                )
            else:
                st.info(f"GitHub 查詢回應 {_r2.status_code}")
        except Exception as e:
            st.info(f"無法查詢 GitHub：{type(e).__name__}")

    st.markdown("---")

    # ═══════════════════════════════════════════════════════════════
    # § 2. 分析設定
    # ═══════════════════════════════════════════════════════════════
    st.markdown(
        '<div style="font-size:.95rem;font-weight:700;color:#e8f4fd;'
        'margin:4px 0 8px;letter-spacing:.04em">⚙️ 分析設定</div>',
        unsafe_allow_html=True
    )

    # ── 日期 ──
    import datetime as _dt
    use_hist = st.checkbox("📅 指定歷史日期", value=False, key="use_hist_date")
    if use_hist:
        hist_date = st.date_input(
            "選擇日期",
            value=_dt.date.today() - _dt.timedelta(days=1),
            min_value=_dt.date(2010, 1, 1),
            max_value=_dt.date.today() - _dt.timedelta(days=1),
            label_visibility="collapsed",
        )
        selected_end_date = hist_date.isoformat()
        st.markdown(
            f'<div style="font-size:.68rem;color:#f0a030;background:#1a1200;'
            f'border:1px solid #6a4a00;border-radius:6px;padding:4px 8px;'
            f'margin:2px 0 6px">⏱ 歷史模式：{hist_date.strftime("%Y-%m-%d")}</div>',
            unsafe_allow_html=True)
    else:
        selected_end_date = ""

    # ── 策略風格 ──
    # 🆕 v9.10n：加入 TW/US 真正最佳選項，並提示哪個對應哪個市場
    _STRATEGY_OPTIONS = [
        "⭐ TW 最佳 (P5+VWAPEXEC)",      # 🆕 TW 真正最佳
        "⭐ US 最佳 (P10+POS+ADX18)",    # 🆕 US 真正最佳
        "🛡️ 極致風控 (IND+DXY)",
        "🛟 超低風險 (五重保護)",
        "🌊 保守 (POS+DXY)",
        "⚖️ 平衡 (POS)",
        "🤖 RL 智能加碼",
        "🚀 進攻 (P0_T1T3)",
    ]
    strategy_style = st.selectbox(
        "🎯 策略風格", options=_STRATEGY_OPTIONS, index=0,
        key="strategy_style",
        help=("⭐ TW 最佳：搜台股 (4 位數) 用此｜"
              "⭐ US 最佳：搜美股 (英文 ticker) 用此｜"
              "其他 6 個是早期風格（TEST 期 RR 都不到 0.25）")
    )

    st.markdown("<div style='margin:10px 0'></div>", unsafe_allow_html=True)
    fetch_btn = st.button("🔍  開始抓取資料", type="primary",
                          use_container_width=True)
    st.markdown("---")

    # ── 🔎 市場掃描器（v9.0 新增）──────────────────────────────
    with st.expander("🔎 市場掃描器（找進場候選）", expanded=False):
        scan_market = st.radio(
            "市場", options=["🇹🇼 台股", "🇺🇸 美股"],
            horizontal=True, key="scan_market_choice"
        )
        scan_signal_filter = st.multiselect(
            "信號類型篩選",
            options=["T1 黃金交叉", "T3 多頭拉回", "🟢 多頭觀察"],
            default=["T1 黃金交叉", "T3 多頭拉回"],
            key="scan_signal_filter",
        )

        # 產業別篩選：依市場列出可選類別
        @st.cache_data(ttl=86400, show_spinner=False)
        def _industry_options(market_choice: str) -> list:
            """依市場回傳該市場存在的產業類別清單（去重排序）"""
            from pathlib import Path as _P
            base = _P(__file__).parent
            out = set()
            if "台股" in market_choice or "TW" in market_choice:
                p = base / 'tw_universe.txt'
                if p.exists():
                    for line in p.read_text(encoding='utf-8').splitlines():
                        if not line or line.startswith('#'): continue
                        parts = line.split('|')
                        if len(parts) >= 5 and parts[4]:
                            out.add(parts[4])
                        elif len(parts) >= 3 and parts[2] in (
                                'ETF','ETN','特別股'):
                            out.add(parts[2])
            else:
                p = base / 'us_sectors.txt'
                if p.exists():
                    for line in p.read_text(encoding='utf-8').splitlines():
                        if not line or line.startswith('#'): continue
                        parts = line.split('|')
                        if len(parts) >= 3:
                            sector = parts[2]
                            if sector and sector != 'NO_SECTOR':
                                out.add(_US_SECTOR_ZH.get(sector, sector))
            return sorted(out)

        scan_industry_filter = st.multiselect(
            "產業別篩選（不選=全部）",
            options=_industry_options(scan_market),
            default=[],
            key="scan_industry_filter",
            help="只篩選選中的產業；不選代表掃描全市場"
        )

        scan_btn = st.button("🚀 開始掃描", use_container_width=True,
                             key="scan_btn")

    st.markdown("---")

    # 各風格對應的描述（v9.10n 更新：加 TW/US 最佳 + TEST 期實測數字）
    # 數字來源：test_strategy_styles_compare.py 全市場 6 年回測（2026-04-29）
    # mean = TEST 22 月平均報酬 / sharpe = TEST RR
    _style_meta = {
        # 🆕 真正最佳（TEST out-of-sample 實測）
        "⭐ TW 最佳 (P5+VWAPEXEC)": dict(
            mode="P5_T1T3+POS+IND+DXY+VWAPEXEC",
            color="#ffd700",
            icon="⭐",
            mean=39.3, low=-64.4, sharpe=0.611, sigma=8.0,
            note="🇹🇼 全 1042 檔 TEST 22 月 RR 0.611（勝率 56% / 中位 +2.3%）｜真正最佳｜配 TOP 200 tier 年化 +110%｜需 vwap_cache (5-min bar)",
        ),
        "⭐ US 最佳 (P10+POS+ADX18)": dict(
            mode="P10_T1T3+POS+ADX18",
            color="#ffd700",
            icon="⭐",
            mean=49.5, low=-99.9, sharpe=0.496, sigma=10.0,
            note="🇺🇸 高流動 555 檔 TEST 22 月 RR 0.496（勝率 55% / 中位 +3.2%）｜vs SPY +6pp 真實 alpha｜需 ADV ≥ $104M 篩選",
        ),
        # 早期風格（保留，TEST RR 標記顯示衰退）
        "🛟 超低風險 (五重保護)": dict(
            mode="P0_T1T3+POS+IND+DXY+WRSI+WADX",
            color="#ff6dc8",
            icon="🛟",
            mean=27.0, low=-111.9, sharpe=0.241, sigma=8.50,
            note="🇹🇼 only｜TEST RR 0.241｜五重保護：POS+產業跨市場+DXY+週RSI+週ADX。最低風險僅 -88（史上最低）",
        ),
        "🤖 RL 智能加碼": dict(
            mode="P0_T1T3+RL",
            color="#10c0c0",
            icon="🤖",
            mean=29.2, low=-152.5, sharpe=0.191, sigma=11.5,
            note="🇹🇼 TEST RR 0.191｜🇺🇸 TEST RR 0.324｜Q-Learning 自動加碼。US 表現勝 TW 1.7×",
        ),
        "🛡️ 極致風控 (IND+DXY)": dict(
            mode="P0_T1T3+POS+IND+DXY",
            color="#9d6dff",
            icon="🛡️",
            mean=31.2, low=-139.6, sharpe=0.223, sigma=7.89,
            note="🇹🇼 only｜TEST RR 0.223｜產業 specific 跨市場 + DXY。FULL 6 年 RR 1.20 但 TEST 大幅衰退",
        ),
        "🌊 保守 (POS+DXY)":  dict(
            mode="P0_T1T3+POS+DXY",
            color="#3dbb6a",
            icon="🌊",
            mean=27.6, low=-139.6, sharpe=0.198, sigma=7.67,
            note="🇹🇼 only｜TEST RR 0.198｜弱美元 + 累積為正加碼。2022 熊市保護有效",
        ),
        "⚖️ 平衡 (POS)": dict(
            mode="P0_T1T3+POS",
            color="#3b9eff",
            icon="⚖️",
            mean=15.8, low=-176.0, sharpe=0.090, sigma=8.89,
            note="🇹🇼 TEST RR 0.090（最差）｜🇺🇸 TEST RR 0.341｜美股表現勝台股 3.8×（無 VWAPEXEC 時 TW 表現差）",
        ),
        "🚀 進攻 (P0_T1T3)": dict(
            mode="P0_T1T3",
            color="#f0c030",
            icon="🚀",
            mean=42.0, low=-187.7, sharpe=0.224, sigma=12.7,
            note="🇹🇼 TEST RR 0.224｜🇺🇸 TEST RR 0.324｜不限門檻最大化獲利但變動大",
        ),
    }
    style_info = _style_meta[strategy_style]
    # 🆕 v9.10n：sharpe 重新定義為 TEST RR（更有意義的指標）
    st.markdown(
        f'<div style="background:#0a1628;border:1px solid {style_info["color"]}55;'
        f'border-radius:6px;padding:8px 12px;margin-top:6px;font-size:.68rem;line-height:1.5">'
        f'<div style="color:{style_info["color"]};font-weight:700;margin-bottom:3px">'
        f'{style_info["icon"]} {style_info["mode"]}</div>'
        f'<div style="color:#7aaac8">'
        f'TEST 均報 <b style="color:{style_info["color"]}">+{style_info["mean"]:.1f}%</b> ｜ '
        f'最差 <b style="color:#ff5555">{style_info["low"]:+.1f}%</b><br>'
        f'TEST RR <b style="color:#f0c030">{style_info["sharpe"]:.3f}</b> ｜ '
        f'跨年σ <b style="color:#7abadd">{style_info["sigma"]:.1f}</b><br>'
        f'<span style="color:#5a8ab0;font-size:.67rem">{style_info["note"]}</span>'
        f'</div></div>',
        unsafe_allow_html=True
    )
    st.session_state['active_strategy'] = style_info

    st.markdown("---")
    st.markdown("""
<div style="font-size:.75rem;color:#8ab8d8;font-weight:700;letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px">
  📖 使用說明
</div>
<div style="font-size:.72rem;color:#5a8ab0;line-height:2;background:#0a1220;border-radius:8px;padding:10px 12px;border:1px solid #1a2f48">
  <span style="color:#3b9eff;font-weight:700">①</span> 在上方輸入欲掃描的股票代號<br>
  <span style="color:#3b9eff;font-weight:700">②</span> 點擊「開始抓取資料」按鈕<br>
  <span style="color:#3b9eff;font-weight:700">③</span> 查看右側指標總表與個股詳情<br>
  <span style="color:#3b9eff;font-weight:700">④</span> 點擊代號連結 → AI 技術分析<br>
  <span style="color:#3b9eff;font-weight:700">⑤</span> 點擊名稱連結 → TradingView 圖表<br>
  <span style="color:#3b9eff;font-weight:700">⑥</span> 底部可下載 Excel 完整報告
</div>
<div style="font-size:.68rem;color:#334455;line-height:1.8;margin-top:10px">
  <b style="color:#3a5a7a">判斷說明</b><br>
  <span style="color:#3b9eff">■</span> 買入 / 強力買入：多頭訊號<br>
  <span style="color:#ff5555">■</span> 賣出 / 強力賣出：空頭訊號<br>
  <span style="color:#556677">■</span> 中立：無明確方向<br><br>
  <b style="color:#3a5a7a">資料來源</b>：Yahoo Finance<br>
  台股自動加 .TW（含反/槓桿 ETF）<br><br>
  <b style="color:#3a5a7a">震盪指標（12）</b><br>
  RSI · 隨機%K · CCI · ADX · AO<br>動量 · MACD · StochRSI · 威廉%R<br>牛熊力度 · 終極震盪 · 布林%B<br><br>
  <b style="color:#3a5a7a">移動均線（15）</b><br>
  EMA/SMA 10/20/30/50/100/200<br>一目均衡 · VWMA · Hull MA
</div>""", unsafe_allow_html=True)

# ── 版本標記：格式變更時自動清除舊快取 ──────────────────────────
_RESULTS_VERSION = 158  # v9.22.4：修「全選儲存失敗」bug（data_editor cached state 用 ver key 重置）2026-05-11
if st.session_state.get("results_version") != _RESULTS_VERSION:
    for _k in ["results", "debug_msgs"]:
        st.session_state.pop(_k, None)
    # ⭐ 同步清 @st.cache_data 函式快取，讓資料結構變更立即生效（VWAP 等）
    try:
        fetch_indicators.clear()
    except Exception:
        pass
    try:
        fetch_indicators_range.clear()
    except Exception:
        pass
    # 🆕 v9.9c：清 Portfolio TOP 200 推薦快取（NLP 改進等格式變更需重算）
    try:
        _scan_top200_signals.clear()
        _load_vwap_applicable.clear()
        _load_otc_tickers.clear()
        get_news_sentiment.clear()
    except Exception:
        pass
    st.session_state["results_version"] = _RESULTS_VERSION

# ── 🔎 市場掃描器處理（v9.0 新增）─────────────────────────────
if scan_btn:
    is_tw = (scan_market == "🇹🇼 台股")
    st.markdown(f"""
<div style="background:#0d1b2e;border:1px solid #1e3a5f;border-radius:10px;
            padding:16px 20px;margin-bottom:1rem">
  <div style="font-size:1.05rem;font-weight:700;color:#8ab8d8">
    🔎 市場掃描中 {'🇹🇼 台股' if is_tw else '🇺🇸 美股'}
  </div>
  <div style="font-size:.75rem;color:#5a8ab0;margin-top:4px">
    策略：{style_info['icon']} {style_info['mode']}
  </div>
</div>""", unsafe_allow_html=True)

    progress = st.progress(0.0, text="準備中...")
    scan_results = []

    # 通用 yfinance 即時掃描函數（雲端 fallback / 美股皆用）
    # 改用 yf.download 批次下載（50/批），大幅加速雲端全市場掃描
    def _scan_via_yfinance(tickers, suffix="", is_tw_market=False,
                           batch_size=80):
        import yfinance as _yf
        import pandas as _pd
        import numpy as _np
        total = len(tickers)
        out = []
        if is_tw_market:
            tw_names = _get_tw_names()
            # 補入靜態 fallback（雲端 twstock 載不齊時）
            tw_names = {**(st.session_state.get('_tw_static_names') or {}),
                        **tw_names}
        else:
            # 美股：合併內建 + 靜態檔
            us_names_full = {**US_NAMES,
                             **(st.session_state.get('_us_static_names') or {})}

        def _eval_one(t, h):
            try:
                if h is None or len(h) < 60: return None
                # 必要欄位：Close + High + Low（算 ADX）
                if 'High' not in h.columns or 'Low' not in h.columns:
                    return None
                df_x = h[['Open','High','Low','Close']].dropna()
                if len(df_x) < 60: return None
                pr = df_x['Close'].values
                hi = df_x['High'].values
                lo = df_x['Low'].values
                s = _pd.Series(pr)
                e20 = s.ewm(span=20, adjust=False).mean().values
                e60 = s.ewm(span=60, adjust=False).mean().values
                delta = s.diff()
                gain = delta.where(delta > 0, 0.0).rolling(14).mean()
                loss = -delta.where(delta < 0, 0.0).rolling(14).mean()
                rs = gain / loss
                rsi = (100 - 100/(1+rs)).values

                # 計算 ADX（標準 14 期）— 排除假多頭關鍵指標
                try:
                    from ta.trend import ADXIndicator
                    adx_arr = ADXIndicator(
                        high=df_x['High'], low=df_x['Low'],
                        close=df_x['Close'], window=14, fillna=False,
                    ).adx().values
                    cur_adx = float(adx_arr[-1]) if not _np.isnan(adx_arr[-1]) else None
                except Exception:
                    cur_adx = None

                ADX_TH = 22  # 趨勢強度門檻：< 22 視為盤整／假多頭

                last = len(pr) - 1
                is_bull = e20[last] > e60[last]
                is_t1 = (last >= 1 and e20[last-1] <= e60[last-1] and
                         e20[last] > e60[last])
                cur_rsi = float(rsi[last]) if not _np.isnan(rsi[last]) else None
                adx_ok = cur_adx is not None and cur_adx >= ADX_TH

                if is_t1 and adx_ok:
                    sig, score = "T1 黃金交叉", 90
                elif is_bull and adx_ok and cur_rsi and cur_rsi < 60:
                    sig, score = "🟢 多頭觀察", 50
                elif is_bull and adx_ok and cur_rsi and cur_rsi >= 75:
                    sig, score = "⚠️ 過熱（RSI≥75）", 30
                elif is_bull and not adx_ok:
                    # 假多頭：不輸出（後面 sig in scan_signal_filter 會過濾掉）
                    sig, score = "⚠️ 假多頭（ADX弱）", 15
                else:
                    sig, score = "🔴 空頭", 10
                # 假多頭直接丟棄
                if sig == "⚠️ 假多頭（ADX弱）": return None
                if sig not in scan_signal_filter: return None
                cur_pr = float(pr[last])
                prev_pr = float(pr[last-1]) if last >= 1 else cur_pr
                chg_pct = (cur_pr - prev_pr) / prev_pr * 100 if prev_pr else 0
                name = tw_names.get(t, "") if is_tw_market \
                       else us_names_full.get(t, "")
                # TradingView 連結（台股 .TW → TWSE/TPEX 前綴；美股直接用代號）
                if is_tw_market:
                    tv_url = f"https://www.tradingview.com/chart/?symbol=TWSE:{t}"
                else:
                    tv_url = f"https://www.tradingview.com/chart/?symbol={t}"
                return dict(
                    ticker=t, name=name,
                    date=str(h.index[-1].date()),
                    price=round(cur_pr, 2),
                    change_pct=round(chg_pct, 2),
                    signal=sig,
                    industry=_get_industry(t),
                    adx=round(cur_adx, 1) if cur_adx else None,
                    is_bull=is_bull,
                    tv_url=tv_url,
                )
            except Exception:
                return None

        # 批次下載
        first_err = None
        for batch_start in range(0, total, batch_size):
            batch_end = min(batch_start + batch_size, total)
            batch = tickers[batch_start:batch_end]
            yf_syms = [f"{t}{suffix}" if suffix else t for t in batch]
            progress.progress(
                batch_start / total,
                text=f"批次下載 {batch_start+1}-{batch_end}/{total}…"
            )
            try:
                df_all = _yf.download(
                    yf_syms,
                    period="1y", auto_adjust=True,
                    group_by='ticker', threads=True,
                    progress=False,
                )
            except Exception as e:
                if first_err is None:
                    first_err = f"download err: {type(e).__name__}: {e}"
                continue
            if df_all is None or df_all.empty:
                if first_err is None:
                    first_err = f"batch {batch_start}: 空 DataFrame"
                continue
            for t, sym in zip(batch, yf_syms):
                try:
                    if len(yf_syms) > 1:
                        if sym not in df_all.columns.get_level_values(0):
                            continue
                        h = df_all[sym].dropna(how='all')
                    else:
                        h = df_all
                    r = _eval_one(t, h)
                    if r: out.append(r)
                except Exception as e:
                    if first_err is None:
                        first_err = f"eval {t}: {type(e).__name__}: {e}"
                    continue
        progress.progress(1.0, text=f"完成 {total} 檔")
        # 回傳除錯資訊
        if not out and first_err:
            st.session_state['_scan_first_err'] = first_err
        return out

    # ── 載入台股全市場（優先從靜態檔，再嘗試 twstock）
    @st.cache_data(ttl=86400, show_spinner=False)
    def _get_all_tw_universe() -> tuple:
        """回傳 (tickers, name_map)。
        優先從 repo 內 tw_universe.txt 載入（雲端最可靠），
        失敗 fallback 到 twstock.codes。"""
        from pathlib import Path as _P2
        # ① 靜態檔（已 commit 至 repo）
        f = _P2(__file__).parent / 'tw_universe.txt'
        if f.exists():
            try:
                tickers, names = [], {}
                for line in f.read_text(encoding='utf-8').splitlines():
                    if not line or line.startswith('#'): continue
                    parts = line.split('|')
                    if len(parts) >= 2:
                        tickers.append(parts[0])
                        names[parts[0]] = parts[1]
                if tickers:
                    return sorted(set(tickers)), names
            except Exception:
                pass
        # ② twstock 動態載入
        try:
            import twstock
            tickers = []; names = {}
            for k, v in twstock.codes.items():
                if v.type in ('股票', 'ETF', 'ETN', '臺灣存託憑證(TDR)', '特別股'):
                    tickers.append(str(k))
                    names[str(k)] = v.name
            return sorted(set(tickers)), names
        except Exception:
            return [], {}

    # ── 動態載入：美股全市場 + 名稱對照
    @st.cache_data(ttl=86400, show_spinner=False)
    def _get_all_us_universe() -> tuple:
        """回傳 (tickers, name_map)。
        優先從 repo us_names.txt 讀（含名稱），失敗 fallback 到 GitHub 純代號清單。"""
        from pathlib import Path as _P3
        # ① 靜態檔（已 commit，含 8000+ 代號 + 名稱）
        f = _P3(__file__).parent / 'us_names.txt'
        if f.exists():
            try:
                tickers, names = [], {}
                for line in f.read_text(encoding='utf-8').splitlines():
                    if not line or line.startswith('#'): continue
                    parts = line.split('|', 1)
                    if len(parts) >= 2:
                        sym = parts[0].strip()
                        # yfinance 不支援的特殊代號跳過
                        if any(c in sym for c in ('$','/','^','.','=')): continue
                        if not sym.replace('-','').isalpha(): continue
                        tickers.append(sym)
                        names[sym] = parts[1].strip()
                if tickers:
                    return sorted(set(tickers)), names
            except Exception:
                pass
        # ② GitHub 純代號 fallback
        url = ("https://raw.githubusercontent.com/rreichel3/"
               "US-Stock-Symbols/main/all/all_tickers.txt")
        try:
            r = requests.get(url, timeout=10)
            if r.status_code != 200: return [], {}
            out = set()
            for line in r.text.splitlines():
                s = line.strip().upper()
                if not s or len(s) > 6: continue
                if any(c in s for c in ('$','/','^','.','=')): continue
                if not s.replace('-','').isalpha(): continue
                out.add(s)
            return sorted(out), {}
        except Exception:
            return [], {}

    # 個股精選清單（雲端 fallback；twstock 失敗時使用）
    TW_INDIVIDUAL_TICKERS = [
        # 半導體
        "2330","2454","2303","3711","6669","3034","3017","3661","6488","8081",
        "6552","2382","3037","2379","6770","3231","2451","2441","8299","6239",
        "2408","6515","3035","3014","6147","6438","3325","3658","3105","6285",
        # 電子/PC/伺服器/網通/EMS
        "2317","2308","2474","2376","2356","3596","2353","2324","2354",
        "3702","3045","2412","4904","3045","2383","2385","2360","6230",
        "8261","6669","6781","2492","3653","6285","8210","3596","3088",
        # 金融
        "2881","2882","2891","2884","2885","2886","2887","2890","2892","5880",
        "2883","2880","2812","2823","2845","2849","2855","2867","2888","2889",
        # 傳產/食品/塑化/紡織/水泥/航運/鋼鐵
        "1101","1102","2002","1216","1301","1303","2105","6505","9904","2912",
        "2207","1326","1722","1227","1402","1907","2027","2014","2030",
        "2603","2609","2615","2618","2606","2606","9904","9910","9921",
        "1605","2812","2371","2362","2438","6121","2049","2049","6213",
        # 生技/光電/車用/其他
        "3008","4123","6446","4128","6691","4736","6446","4906","8454",
        "1519","2059","2049","8086","5483","3443","2059","1907","6121",
    ]

    # 動態取得所有上市/上櫃 ETF（00 開頭）
    @st.cache_data(ttl=3600, show_spinner=False)
    def _get_all_tw_etfs() -> list:
        try:
            import twstock
            etfs = []
            for code, info in twstock.codes.items():
                # ETF 篩選：代號以 00 開頭或 group=='ETF'
                if (str(code).startswith('00') or
                    getattr(info, 'group', '') == 'ETF' or
                    'ETF' in (getattr(info, 'name', '') or '')):
                    etfs.append(str(code))
            return sorted(set(etfs))
        except Exception:
            # fallback：手動清單
            return [
                "0050","0051","0052","0053","0055","0056","0057","0061",
                "006201","006203","006204","006208",
                "00631L","00632R","00633L","00634R","00635U","00636",
                "00637L","00638R","00640L","00641R","00642U","00643K",
                "00646","00650L","00651R","00652","00655L","00656R","00657",
                "00660","00661","00662","00663L","00664R","00665L","00666R",
                "00668","00670L","00671R","00672L","00673R","00674R",
                "00675L","00676R","00678","00679B","00680L","00681R","00682U",
                "00687B","00688L","00689R","00690","00691R","00692","00693U",
                "00694B","00695B","00696B","00697B","00700","00701","00709",
                "00710B","00711B","00712","00713","00714","00715L","00717",
                "00718B","00720B","00724B","00725B","00727B","00730","00731",
                "00733","00735","00736","00737","00739","00740B","00742",
                "00750","00751B","00752","00755","00757","00762","00763U",
                "00770","00771","00773B","00775B","00777","00778B","00783",
                "00850","00851","00861","00865B","00876","00878","00881",
                "00885","00891","00892","00893","00894","00895","00896",
                "00900","00901","00904","00905","00906","00907","00909",
                "00911","00912","00913","00915","00916","00918","00919",
                "00920","00921","00922","00923","00925","00927B","00929",
                "00930","00931B","00932","00933","00934B","00935B","00936",
                "00937B","00939","00940","00941","00943","00944","00945",
                "00946B","00947B","00948","00949","00951","00952","00953B",
                "00954B","00955B","00956B","00957B",
            ]

    # 美股大幅擴充：個股 ~150 + ETF ~80
    US_INDIVIDUAL_TICKERS = [
        # 科技七雄 + 半導體
        "AAPL","MSFT","NVDA","GOOGL","GOOG","AMZN","META","TSLA",
        "AVGO","AMD","INTC","QCOM","TXN","MU","AMAT","LRCX","KLAC","ADI",
        "MRVL","ASML","TSM","ARM","SMCI","ON","NXPI","MCHP","SWKS","QRVO",
        # 軟體 / 雲 / SaaS
        "ORCL","CRM","ADBE","NOW","INTU","WDAY","SNOW","PLTR","DDOG","NET",
        "MDB","CRWD","ZS","OKTA","TEAM","ZM","PANW","FTNT","CYBR",
        "SHOP","SQ","PYPL","COIN","HOOD","SOFI","AFRM","UPST",
        # 網路 / 媒體 / 通訊
        "NFLX","DIS","CMCSA","T","VZ","TMUS","CHTR","WBD","PARA","ROKU",
        "SPOT","PINS","SNAP","RBLX","EA","TTWO","UBER","LYFT","ABNB","DASH",
        # 金融
        "JPM","BAC","WFC","C","GS","MS","BLK","SCHW","AXP","COF",
        "USB","PNC","TFC","BK","STT","MET","PRU","AIG","TRV","ALL",
        "V","MA","FIS","FISV","ICE","CME","SPGI","MCO","MSCI",
        # 消費 / 零售 / 餐飲
        "WMT","COST","HD","LOW","TGT","DG","DLTR","BJ","KR",
        "MCD","SBUX","CMG","YUM","DPZ","WEN","QSR","MNST","KO","PEP",
        "PG","CL","KMB","CHD","CLX","UL","EL","NKE","LULU","UAA","DECK",
        # 工業 / 國防 / 運輸
        "BA","CAT","DE","GE","HON","LMT","RTX","NOC","GD","BAH",
        "UNP","CSX","NSC","UPS","FDX","JBHT","CHRW",
        # 能源 / 原物料
        "XOM","CVX","COP","PSX","SLB","HAL","BKR","OXY","EOG","DVN",
        "FCX","NEM","GOLD","AA","CLF","X","MP","LIN","APD","SHW",
        # 醫療 / 藥廠 / 生技
        "JNJ","PFE","MRK","ABBV","BMY","LLY","TMO","ABT","DHR","UNH",
        "CVS","CI","HUM","ANTM","MDT","SYK","BSX","ISRG","ZBH","BAX",
        "AMGN","GILD","BIIB","REGN","VRTX","MRNA","BNTX","NVAX",
        # 公用 / REIT
        "NEE","DUK","SO","D","AEP","EXC","SRE","XEL","ED",
        "AMT","PLD","CCI","EQIX","DLR","SPG","O","WELL","PSA","AVB",
    ]

    US_ETF_TICKERS = [
        # 大盤指數
        "SPY","VOO","IVV","QQQ","DIA","IWM","VTI","VT","VEA","VWO",
        # 板塊 SPDR
        "XLK","XLF","XLE","XLV","XLI","XLY","XLP","XLU","XLB","XLRE","XLC",
        # 半導體 / 科技主題
        "SMH","SOXX","SOXL","TQQQ","SQQQ","TECL","TECS","FNGU","FNGD",
        "BOTZ","ROBO","ARKK","ARKW","ARKG","ARKQ","ARKF","ARKX",
        "IGV","HACK","CIBR","CLOU","SKYY","WCLD","FDN",
        # 因子 / 風格
        "MTUM","QUAL","USMV","VLUE","SIZE","MOAT","SPLV","SPHQ",
        # 國際
        "EFA","EEM","FXI","INDA","EWJ","EWZ","EWG","EWU","EWT","RSX",
        # 固定收益
        "AGG","BND","TLT","IEF","SHY","LQD","HYG","JNK","MUB","TIP",
        # 商品 / 黃金 / 能源
        "GLD","IAU","SLV","USO","UNG","DBA","DBC","UUP","FXE","FXY",
        # 房地產 / REIT
        "VNQ","SCHH","REM","MORT",
        # 高股息 / 收益
        "SCHD","VYM","HDV","DVY","SPHD","PFF","DGRO","NOBL",
        # 中小型 / 微型
        "VB","IJH","IJR","SCHA","SCHM",
    ]
    US_HOT_TICKERS = US_INDIVIDUAL_TICKERS + US_ETF_TICKERS

    if is_tw:
        # 台股：優先使用本地 data_cache，雲端 fallback 至 yfinance
        local_ok = False
        try:
            import data_loader as _dl
            import variant_strategy as _vs
            import daily_scanner as _ds
            cache_dir = _dl.CACHE_DIR
            files = sorted(cache_dir.glob('*.parquet'))
            # 產業篩選（本地入口）
            if files and scan_industry_filter:
                files = [fp for fp in files
                         if _get_industry(fp.stem) in scan_industry_filter]
            if files:
                tw_names = _get_tw_names()
                mode = style_info['mode']
                total = len(files)
                for i, fp in enumerate(files):
                    if i % 20 == 0:
                        progress.progress(i / total,
                                          text=f"掃描中 {i+1}/{total}…（本地快取）")
                    ticker = fp.stem
                    try:
                        r = _ds.scan_one(ticker, str(fp), mode=mode)
                        if r is not None and r.get('signal') in scan_signal_filter:
                            # 補上名稱與漲跌
                            try:
                                df_local = pd.read_parquet(fp)
                                if len(df_local) >= 2:
                                    cur_p = float(df_local['Close'].iloc[-1])
                                    prev_p = float(df_local['Close'].iloc[-2])
                                    r['change_pct'] = round(
                                        (cur_p - prev_p) / prev_p * 100, 2) \
                                        if prev_p else 0
                            except Exception:
                                r['change_pct'] = None
                            r['name'] = tw_names.get(ticker, "")
                            r['price'] = round(r.get('price', 0), 2)
                            r['industry'] = _get_industry(ticker)
                            r['tv_url'] = f"https://www.tradingview.com/chart/?symbol=TWSE:{ticker}"
                            # 移除掃描表不需要的欄位
                            r.pop('score', None)
                            r.pop('rsi', None)
                            scan_results.append(r)
                    except Exception:
                        continue
                progress.progress(1.0, text=f"完成 {total} 檔（本地快取）")
                local_ok = True
        except ImportError:
            pass

        if not local_ok:
            # 雲端：載入台股全市場（優先 tw_universe.txt → twstock → 手動）
            full, tw_static_names = _get_all_tw_universe()
            if not full:
                full = list(dict.fromkeys(
                    TW_INDIVIDUAL_TICKERS + _get_all_tw_etfs()))
                tw_static_names = {}
            st.session_state['_tw_static_names'] = tw_static_names
            # 產業篩選（雲端入口）
            if scan_industry_filter:
                full = [t for t in full if _get_industry(t) in scan_industry_filter]
                st.info(
                    f"📡 雲端模式：依產業篩選後掃描 **{len(full)}** 檔"
                    f"（產業：{'、'.join(scan_industry_filter)}）"
                )
            else:
                st.info(
                    f"📡 雲端模式：批次掃描台股全市場 **{len(full)}** 檔"
                    f"（含上市/上櫃/ETF/ETN/TDR/特別股）"
                )
            scan_results = _scan_via_yfinance(full, suffix=".TW",
                                              is_tw_market=True)
    else:
        # 美股：靜態檔 us_names.txt（~8000 檔，含名稱）
        full_us, us_static_names = _get_all_us_universe()
        if full_us and len(full_us) > 1000:
            extra_hot = list(dict.fromkeys(US_HOT_TICKERS + full_us))
            st.session_state['_us_static_names'] = us_static_names
            # 產業篩選（美股）
            if scan_industry_filter:
                extra_hot = [t for t in extra_hot
                             if _get_industry(t) in scan_industry_filter]
                st.info(
                    f"📡 依產業篩選後掃描美股 **{len(extra_hot)}** 檔"
                    f"（產業：{'、'.join(scan_industry_filter)}）"
                )
            else:
                st.info(
                    f"📡 即時抓取美股全市場 **{len(extra_hot)}** 檔"
                    f"（NASDAQ + NYSE + AMEX，預估 3-5 分鐘）"
                )
            scan_results = _scan_via_yfinance(extra_hot, is_tw_market=False)
        else:
            st.info(
                f"📡 即時抓取美股 {len(US_HOT_TICKERS)} 檔"
                f"（個股 {len(US_INDIVIDUAL_TICKERS)} + ETF {len(US_ETF_TICKERS)}）"
                f"（無法載入 NASDAQ 全清單，使用內建熱門池）"
            )
            scan_results = _scan_via_yfinance(US_HOT_TICKERS, is_tw_market=False)

    if scan_results:
        df_scan = pd.DataFrame(scan_results)
        if 'score' in df_scan.columns:
            df_scan = df_scan.sort_values('score', ascending=False)
        else:
            df_scan = df_scan.sort_values('change_pct', ascending=False)
        df_scan = df_scan.reset_index(drop=True)
        # 移除掃描表不需要顯示的欄位
        for col in ('score', 'rsi'):
            if col in df_scan.columns:
                df_scan = df_scan.drop(columns=[col])
        # 欄位順序：tv_url 在 name 後面（讓 LinkColumn 顯示為「📈 連結」）
        preferred_cols = ['ticker', 'name', 'tv_url',
                          'price', 'change_pct',
                          'industry',
                          'signal', 'adx', 'is_bull',
                          'cross_days', 'date']
        ordered = [c for c in preferred_cols if c in df_scan.columns]
        ordered += [c for c in df_scan.columns if c not in ordered]
        df_scan = df_scan[ordered]

        st.markdown(
            f'<div style="font-size:.85rem;color:#3dbb6a;margin:.5rem 0">'
            f'✓ 找到 <b>{len(df_scan)}</b> 檔候選。勾選左側方框可多選後加入清單。</div>',
            unsafe_allow_html=True)

        # 持久化已存自選股清單（雲端 localStorage）
        _wls_for_scan = _load_watchlists()

        # 用 st.dataframe + selection_mode 達成多選
        sel_event = st.dataframe(
            df_scan,
            use_container_width=True,
            height=600,
            on_select="rerun",
            selection_mode="multi-row",
            key="scan_table_select",
            column_config={
                "ticker": st.column_config.TextColumn("代號", width="small"),
                "name": st.column_config.TextColumn("名稱", width="medium"),
                "tv_url": st.column_config.LinkColumn(
                    "圖表", display_text="📈 TV", width="small"),
                "price": st.column_config.NumberColumn("收盤價", format="%.2f"),
                "change_pct": st.column_config.NumberColumn(
                    "漲跌%", format="%+.2f%%"),
                "industry": st.column_config.TextColumn(
                    "產業別", width="small"),
                "signal": st.column_config.TextColumn("信號", width="medium"),
                "adx": st.column_config.NumberColumn("ADX", format="%.1f"),
                "is_bull": st.column_config.CheckboxColumn("多頭"),
                "cross_days": st.column_config.NumberColumn(
                    "交叉天", format="%.0f"),
                "date": st.column_config.TextColumn("日期"),
            },
        )

        # 取出選中的列
        try:
            sel_rows = sel_event.selection.rows  # type: ignore
        except Exception:
            sel_rows = []
        sel_tickers = [df_scan.iloc[i]['ticker'] for i in sel_rows] if sel_rows else []

        # ── 加入清單操作面板 ──────────────────────────────────────
        st.markdown(
            f'<div style="background:#0a1628;border:1px solid #1a2f48;'
            f'border-radius:10px;padding:12px 16px;margin-top:10px">'
            f'<div style="color:#7ab0d0;font-size:.78rem;margin-bottom:8px">'
            f'已選擇 <b style="color:#3dbb6a">{len(sel_tickers)}</b> 檔'
            f'{"：" + "、".join(sel_tickers[:8]) + ("…" if len(sel_tickers)>8 else "") if sel_tickers else ""}'
            f'</div></div>',
            unsafe_allow_html=True
        )

        if sel_tickers:
            _add_targets = ["（覆蓋目前清單）"] + sorted(_wls_for_scan.keys())
            ca, cb, cc = st.columns([2, 1, 1])
            with ca:
                _add_to = st.selectbox(
                    "加入到清單", options=_add_targets,
                    key="scan_add_target",
                    help="選擇現有清單將以「合併」方式加入（不會覆蓋既有代號）；選『覆蓋目前清單』則將目前文字框替換為選中項。"
                )
            with cb:
                if st.button("➕ 加入清單", use_container_width=True,
                             type="primary", key="scan_append_btn"):
                    if _add_to == "（覆蓋目前清單）":
                        # 覆蓋目前 textarea
                        st.session_state[f"stock_input_{_selected_wl}"] = \
                            "\n".join(sel_tickers)
                        st.success(f"✓ 已覆蓋目前清單為 {len(sel_tickers)} 檔")
                    else:
                        # 合併到指定 saved 清單
                        existing = _wls_for_scan.get(_add_to, "")
                        existing_set = {l.strip() for l in existing.splitlines()
                                        if l.strip() and not l.strip().startswith('#')}
                        new_set = set(sel_tickers) - existing_set
                        merged = (existing.rstrip() + "\n" if existing.strip() else "") + \
                                 "\n".join(sel_tickers)
                        # 去重保序
                        seen = set(); out_lines = []
                        for ln in merged.splitlines():
                            s = ln.strip()
                            if not s or s.startswith('#'):
                                out_lines.append(ln); continue
                            if s in seen: continue
                            seen.add(s); out_lines.append(ln)
                        _wls_for_scan[_add_to] = "\n".join(out_lines)
                        _save_watchlists(_wls_for_scan)
                        st.success(
                            f"✓ 加入「{_add_to}」（新增 {len(new_set)} 檔，"
                            f"原 {len(existing_set)} → {len(seen)}）"
                        )
                    st.rerun()
            with cc:
                if st.button("📋 全部填入文字框", use_container_width=True,
                             key="scan_fill_btn"):
                    st.session_state[f"stock_input_{_selected_wl}"] = \
                        "\n".join(df_scan['ticker'].tolist())
                    st.success(f"✓ 已填入 {len(df_scan)} 檔")
                    st.rerun()
        else:
            st.caption("👆 在表格左側勾選想加入清單的股票（可多選）")
    else:
        err = st.session_state.pop('_scan_first_err', None)
        if err:
            st.error(
                f"⚠️ 沒有符合條件的股票（且批次下載出錯）\n\n"
                f"第一個錯誤：`{err}`\n\n"
                f"可能原因：yfinance 在 Streamlit Cloud 被限速、"
                f"或 yfinance 版本與此程式不相容。"
            )
        else:
            st.warning("沒有符合條件的股票（資料下載成功，但全市場都不符合篩選條件）")
    st.stop()

# ── 用 session_state 儲存結果，避免下拉選單觸發重跑時資料消失 ──
if fetch_btn:
    # 歷史模式快取永不過期（數據不變）；即時模式才清快取
    if not selected_end_date:
        fetch_indicators.clear()
        fetch_news.clear()
    for k in list(st.session_state.keys()):
        if k.startswith("ai_") or k == "results" or k == "debug_msgs":
            del st.session_state[k]
    st.session_state["fetch_end_date"] = selected_end_date
    stocks = parse_input(stock_input)
    if not stocks:
        st.error("股票清單為空，請重新輸入"); st.stop()

    progress_bar = st.progress(0, text="準備中...")
    status_ph    = st.empty()
    results    = []
    debug_msgs = []

    for i, (ticker, market) in enumerate(stocks):
        progress_bar.progress(i / len(stocks), text=f"抓取 {ticker}  ({i+1}/{len(stocks)})")
        status_ph.markdown(
            f'<div style="font-size:.78rem;color:#5a8ab0;text-align:center">'
            f'正在抓取 <b style="color:#8ab8d8">{ticker}</b>...</div>',
            unsafe_allow_html=True)
        _end = st.session_state.get("fetch_end_date", "")
        d = fetch_indicators(ticker, market, _end, _cache_ver=APP_VERSION)
        if d and d.get("_error"):
            debug_msgs.append(f"❌ {ticker} ({get_yf_symbol(ticker)}): {d['_error']}")
            d = None
        if d and d.get("close"):
            g_trend    = judge_trend(d)
            g_position = judge_position(d)
            g_momentum = judge_momentum(d)
            g_aux      = judge_aux(d)
            ts = calc_summary(g_trend,    TREND_W)
            ps = calc_summary(g_position, POSITION_W)
            mom_grade  = compute_momentum_grade(d)
            ms_b, ms_s, ms_n, _ = calc_summary(g_momentum, MOMENTUM_W)
            ms = (ms_b, ms_s, ms_n, mom_grade)   # direct verdict overrides sum
            xs = _calc_aux_summary(g_aux, AUX_W)
            tb = round(ts[0]+ps[0]+ms[0]+xs[0], 1)
            ts_= round(ts[1]+ps[1]+ms[1]+xs[1], 1)
            tn_= round(ts[2]+ps[2]+ms[2]+xs[2], 1)
            raw_verdict = _rec(tb, ts_)
            verdict, cap = apply_cap(raw_verdict, d, mom_grade)
            # 保留舊格式供 Excel 相容
            osc = judge_oscillators(d)
            mas = judge_mas(d)
            results.append((ticker, market, d, False,
                            (g_trend, g_position, g_momentum, g_aux),
                            (ts, ps, ms, xs),
                            (tb, ts_, tn_, verdict), cap,
                            osc, mas))
        else:
            if not any(m.startswith(f"❌ {ticker}") for m in debug_msgs):
                debug_msgs.append(f"❌ {ticker} ({get_yf_symbol(ticker)}): 無資料或資料不足")
            results.append((ticker, market, None, True,
                            ([], [], [], []), ((0,0,0,"中立"),)*4,
                            (0, 0, 0, "中立"), None, [], []))

    progress_bar.progress(1.0, text="完成 ✓")
    status_ph.empty()

    # 儲存到 session_state
    st.session_state["results"]         = results
    st.session_state["debug_msgs"]      = debug_msgs
    st.session_state["results_version"] = _RESULTS_VERSION

# 從 session_state 讀取（保持結果在模型切換時不消失）
if "results" not in st.session_state:
    st.markdown("""
<div style="text-align:center;padding:30px 20px 12px">
  <div style="font-size:2rem;margin-bottom:8px">📈</div>
  <div style="font-size:.95rem;color:#3a6a9a">在左側輸入股票代號，點擊「開始抓取資料」</div>
  <div style="font-size:.72rem;color:#1e3a5f;margin-top:4px">支援台股 · NASDAQ · NYSE · 任何 Yahoo Finance 代號</div>
</div>""", unsafe_allow_html=True)
    # TOP 200 已移到健壯性驗證下方（永遠顯示），這裡不重複
    st.stop()

results    = st.session_state["results"]
debug_msgs = st.session_state.get("debug_msgs", [])

error_msgs = [m for m in debug_msgs if m.startswith("❌")]
if error_msgs:
    with st.expander(f"⚠️ {len(error_msgs)} 筆無法取得資料（點擊展開查看原因）", expanded=False):
        for msg in error_msgs:
            st.markdown(f"<div style='font-size:.8rem;color:#ff8080;font-family:monospace'>{msg}</div>",
                        unsafe_allow_html=True)
        st.markdown("<div style='font-size:.75rem;color:#556677;margin-top:8px'>可能原因：代號格式不正確、Yahoo Finance 暫時無法存取、或此標的在 Yahoo Finance 不存在</div>",
                    unsafe_allow_html=True)

total      = len(results)
ok         = sum(1 for r in results if not r[3])
# 🆕 改為 v8 操作分類 + 收集 ticker 清單
_entry_tks = []; _exit_tks = []; _hold_tks = []; _wait_tks = []
for r in results:
    if r[3] or not r[2]: continue
    a = classify_action(r[2])
    if   a == 'ENTRY': _entry_tks.append(r[0])
    elif a == 'EXIT':  _exit_tks.append(r[0])
    elif a == 'HOLD':  _hold_tks.append(r[0])
    elif a == 'WAIT':  _wait_tks.append(r[0])
entry_count = len(_entry_tks)
exit_count  = len(_exit_tks)
hold_count  = len(_hold_tks)
wait_count  = len(_wait_tks)


def _tickers_chips(tks, max_show=8, color='#a8cce8'):
    """卡片底下顯示 ticker 代碼（最多 max_show 個，超過用 +N more）"""
    if not tks:
        return '<div style="font-size:.62rem;color:#3a5a7a;margin-top:4px">—</div>'
    show = tks[:max_show]
    more = len(tks) - len(show)
    chips = ''.join(
        f'<span style="display:inline-block;padding:1px 4px;background:#0a1830;'
        f'color:{color};font-size:.62rem;font-family:\'IBM Plex Mono\',monospace;'
        f'border-radius:3px;margin:2px 2px 0 0">{t}</span>'
        for t in show
    )
    if more > 0:
        chips += (f'<span style="font-size:.6rem;color:#5a7a99;margin-left:2px">+{more}</span>')
    return f'<div style="margin-top:4px;line-height:1.5">{chips}</div>'


st.markdown(f"""
<div class="cards-row">
  <div class="card total"><div class="c-label">抓取完成</div>
    <div class="c-value">{ok}<span style="font-size:1rem;color:#3a5a7a">/{total}</span></div></div>
  <div class="card entry"><div class="c-label">可進場</div>
    <div class="c-value">{entry_count}</div>{_tickers_chips(_entry_tks, color='#3dbb6a')}</div>
  <div class="card exit"><div class="c-label">應出倉</div>
    <div class="c-value">{exit_count}</div>{_tickers_chips(_exit_tks, color='#ff7777')}</div>
  <div class="card hold"><div class="c-label">持倉中</div>
    <div class="c-value">{hold_count}</div>{_tickers_chips(_hold_tks, color='#7abadd')}</div>
  <div class="card neu"><div class="c-label">觀望</div>
    <div class="c-value">{wait_count}</div></div>
  <div class="card total"><div class="c-label">{'歷史日期' if st.session_state.get('fetch_end_date') else '更新時間'}</div>
    <div class="c-value" style="font-size:{'0.78rem' if st.session_state.get('fetch_end_date') else '.95rem'};color:{'#f0a030' if st.session_state.get('fetch_end_date') else '#f0f4ff'}">{st.session_state.get('fetch_end_date') or datetime.now().strftime("%H:%M")}</div></div>
</div>""", unsafe_allow_html=True)

platform_url_tpl = "https://www.perplexity.ai/search?q={prompt}"
selected_platform = "Perplexity"

# 🆕 📊 今日 Portfolio 推薦（嚴格依 TOP 200 tier，從 data_cache 全市場掃）
@st.cache_data(ttl=3600)
def _scan_top200_signals():
    """從 data_cache 直接讀全 TOP 200 個股的最新指標 + 分類動作。
    比起依賴 results（僅含使用者搜尋），這個函式掃整個 TOP 200 清單。
    使用 data_cache parquet（已含 e20/e60/rsi/adx 等指標），快速。
    """
    import pandas as _pd
    import numpy as _np
    from pathlib import Path as _P
    tier_data = _load_vwap_applicable()
    top_tickers = [t for t, info in tier_data.items() if info.get('tier') == 'TOP']

    # 載入 stock 名稱映射
    name_map = {}
    try:
        sj = _P(__file__).parent / 'tw_stock_list.json'
        if sj.exists():
            import json as _json
            with open(sj, encoding='utf-8') as _f:
                _data = _json.load(_f)
            if isinstance(_data, dict):
                if 'tickers' in _data:
                    _data = _data['tickers']
                for k, v in _data.items():
                    if isinstance(v, dict):
                        name_map[k] = v.get('name', '')
    except Exception:
        pass

    # 概念股名稱備援
    if not name_map:
        for t in top_tickers:
            name_map[t] = ''

    entry, exit_, hold = [], [], []
    for t in top_tickers:
        try:
            cache_path = _P(__file__).parent / 'data_cache' / f'{t}.parquet'
            if not cache_path.exists():
                continue
            df = _pd.read_parquet(cache_path)
            if df is None or len(df) < 30:
                continue
            last_idx = -1
            close = float(df['Close'].iloc[last_idx])

            # 建構 d dict (符合 classify_action 介面)
            ema20 = float(df['e20'].iloc[last_idx]) if 'e20' in df.columns and not _pd.isna(df['e20'].iloc[last_idx]) else None
            ema60 = float(df['e60'].iloc[last_idx]) if 'e60' in df.columns and not _pd.isna(df['e60'].iloc[last_idx]) else None
            rsi = float(df['rsi'].iloc[last_idx]) if 'rsi' in df.columns and not _pd.isna(df['rsi'].iloc[last_idx]) else None
            rsi_prev = float(df['rsi'].iloc[last_idx-1]) if len(df) >= 2 and 'rsi' in df.columns and not _pd.isna(df['rsi'].iloc[last_idx-1]) else None
            rsi_prev2 = float(df['rsi'].iloc[last_idx-2]) if len(df) >= 3 and 'rsi' in df.columns and not _pd.isna(df['rsi'].iloc[last_idx-2]) else None
            adx = float(df['adx'].iloc[last_idx]) if 'adx' in df.columns and not _pd.isna(df['adx'].iloc[last_idx]) else None
            atr14 = float(df['atr'].iloc[last_idx]) if 'atr' in df.columns and not _pd.isna(df['atr'].iloc[last_idx]) else None

            # 計算 ema20_cross_days
            cross_days = None
            if 'e20' in df.columns and 'e60' in df.columns and len(df) >= 30:
                e20s = df['e20'].values
                e60s = df['e60'].values
                cur_bull = e20s[last_idx] > e60s[last_idx] if not (_np.isnan(e20s[last_idx]) or _np.isnan(e60s[last_idx])) else None
                if cur_bull is not None:
                    for k in range(1, min(60, len(df))):
                        if _np.isnan(e20s[last_idx-k]) or _np.isnan(e60s[last_idx-k]):
                            continue
                        prev_bull = e20s[last_idx-k] > e60s[last_idx-k]
                        if prev_bull != cur_bull:
                            cross_days = k if cur_bull else -k
                            break

            d = {
                'close': close,
                'ema20': ema20, 'ema60': ema60,
                'rsi': rsi, 'rsi_prev': rsi_prev, 'rsi_prev2': rsi_prev2,
                'adx': adx, 'atr14': atr14,
                'ema20_cross_days': cross_days,
            }
            action = classify_action(d)
            tier_info = tier_data.get(t, {})
            delta_v = tier_info.get('delta', 0)
            row = (t, name_map.get(t, t), d, delta_v)
            if action == 'ENTRY':
                entry.append(row)
            elif action == 'EXIT':
                exit_.append(row)
            elif action == 'HOLD':
                hold.append(row)
        except Exception:
            continue

    entry.sort(key=lambda x: -x[3])
    exit_.sort(key=lambda x: -x[3])
    hold.sort(key=lambda x: -x[3])
    return entry, exit_, hold


_top_entry_picks, _top_exit_picks, _top_hold_picks = _scan_top200_signals()

if _top_entry_picks or _top_exit_picks or _top_hold_picks:
    def _pick_html(picks, max_show=8, accent='#3dbb6a'):
        if not picks:
            return '<div style="color:#3a5a7a;font-size:.72rem;padding:4px 8px">— 暫無 —</div>'
        out = ''
        for t, name, d, delta in picks[:max_show]:
            close = d.get('close')
            close_str = f'{close:.2f}' if close else '—'
            rsi = d.get('rsi')
            rsi_str = f'RSI {rsi:.0f}' if rsi else ''
            per = d.get('per')
            pe_str = f'PE {per:.1f}' if per else ''
            cd = d.get('ema20_cross_days')
            cd_str = ''
            if cd is not None and 0 < cd <= 10:
                cd_str = f'<span style="color:#5a8ab0;font-size:.62rem">T1 {cd}d</span>　'
            elif rsi is not None and rsi < 50:
                cd_str = f'<span style="color:#5a8ab0;font-size:.62rem">T3 拉回</span>　'
            out += (
                f'<div style="display:flex;align-items:baseline;gap:6px;'
                f'padding:5px 10px;font-size:.78rem;border-bottom:1px solid #1a2a3f">'
                f'<span style="color:{accent};font-weight:700;font-family:monospace;'
                f'min-width:48px">{t}</span>'
                f'<span style="color:#a8cce8;flex:1;overflow:hidden;text-overflow:ellipsis;'
                f'white-space:nowrap;max-width:90px">{name}</span>'
                f'<span style="color:#e8f4fd;font-family:monospace;font-size:.74rem">{close_str}</span>'
                f'<span style="color:#7a8899;font-size:.7rem">{rsi_str}</span>'
                f'<span style="color:#7a8899;font-size:.7rem">{pe_str}</span>'
                f'{cd_str}'
                f'<span style="color:#3dbb6a;font-size:.65rem;background:#0a3a1f;'
                f'padding:1px 5px;border-radius:3px">Δ {delta:+.0f}%</span>'
                f'</div>'
            )
        if len(picks) > max_show:
            out += (f'<div style="text-align:center;padding:5px;color:#5a8ab0;font-size:.7rem">'
                    f'＋{len(picks)-max_show} 檔在表格內</div>')
        return out

    # Header with explanation
    st.markdown(
        '<div style="background:#0a1a2a;border:1px solid #3dbb6a55;border-radius:10px;'
        'padding:10px 14px;margin:8px 0;display:flex;align-items:center;gap:14px;'
        'flex-wrap:wrap">'
        '<div style="font-size:1.05rem;font-weight:700;color:#3dbb6a">📊 今日 Portfolio 推薦</div>'
        f'<div style="color:#7abadd;font-size:.78rem">'
        f'進場 <b>{len(_top_entry_picks)}</b>　│　'
        f'出倉 <b>{len(_top_exit_picks)}</b>　│　'
        f'持倉 <b>{len(_top_hold_picks)}</b></div>'
        '<div style="color:#7a8899;font-size:.72rem;flex:1">'
        '嚴格從 ⭐ <b style="color:#3dbb6a">TOP 200 tier</b> 篩選'
        '（v8+P5+VWAPEXEC 適用清單，預期年化 ~52%）'
        '</div></div>',
        unsafe_allow_html=True
    )

    # 三欄
    cols = st.columns(3)
    with cols[0]:
        st.markdown(
            '<div style="background:#0a1e10;border:1px solid #3dbb6a55;border-radius:8px;'
            'padding:6px 4px;margin-bottom:8px">'
            '<div style="color:#3dbb6a;font-weight:700;padding:0 8px 4px;font-size:.85rem">'
            f'🚀 進場（{len(_top_entry_picks)}）</div>'
            + _pick_html(_top_entry_picks, accent='#3dbb6a')
            + '</div>', unsafe_allow_html=True)
    with cols[1]:
        st.markdown(
            '<div style="background:#1a0808;border:1px solid #ff555555;border-radius:8px;'
            'padding:6px 4px;margin-bottom:8px">'
            '<div style="color:#ff7777;font-weight:700;padding:0 8px 4px;font-size:.85rem">'
            f'🚪 出倉（{len(_top_exit_picks)}）</div>'
            + _pick_html(_top_exit_picks, accent='#ff7777')
            + '</div>', unsafe_allow_html=True)
    with cols[2]:
        st.markdown(
            '<div style="background:#0a1830;border:1px solid #5a8ab055;border-radius:8px;'
            'padding:6px 4px;margin-bottom:8px">'
            '<div style="color:#7abadd;font-weight:700;padding:0 8px 4px;font-size:.85rem">'
            f'📌 持倉中（{len(_top_hold_picks)}）</div>'
            + _pick_html(_top_hold_picks, accent='#7abadd')
            + '</div>', unsafe_allow_html=True)

# 🆕 📊 Portfolio Simulator —————————————————————————————————
with st.expander("📊 投組模擬器（基於 1028 檔 22 個月 out-of-sample 結果）", expanded=False):
    @st.cache_data(ttl=86400)
    def _load_portfolio_data():
        """載入 portfolio simulation 基礎資料"""
        import json as _json
        try:
            from pathlib import Path as _P
            with open(_P(__file__).parent / 'full_market_results.json', encoding='utf-8') as f:
                data = _json.load(f)
            with open(_P(__file__).parent / 'vwap_applicable.json', encoding='utf-8') as f:
                tier_data = _json.load(f)
            vwap_test = data.get('B VWAPEXEC|TEST', {})
            base_test = data.get('A baseline|TEST', {})
            pnl = dict(zip(vwap_test['tickers'], vwap_test['pnl_pcts']))
            base_pnl = dict(zip(base_test['tickers'], base_test['pnl_pcts']))
            return pnl, base_pnl, tier_data
        except Exception:
            return {}, {}, {}

    _pnl_t, _base_t, _tier_d = _load_portfolio_data()

    if not _pnl_t:
        st.markdown('<div style="color:#7a8899;padding:8px">portfolio 資料尚未產生（需 full_market_results.json）</div>',
                    unsafe_allow_html=True)
    else:
        import numpy as np
        # 上方：說明
        st.markdown(
            '<div style="font-size:.8rem;color:#5a8ab0;line-height:1.7;padding:6px 0">'
            '基於回測 1028 檔 × 22 月（2024.6-2026.4）TEST 期 out-of-sample 報酬。'
            '<b style="color:#f0a030">⚠️ Top N 是事後 cherry-pick</b>，未來無法複製；'
            '實務可用是 <b style="color:#3dbb6a">⭐ TOP 200 tier</b>（前向確定的清單）。'
            '</div>', unsafe_allow_html=True
        )

        # 互動：投入金額 + 組合選擇
        col_a, col_b, col_c = st.columns([1, 1.2, 1.5])
        with col_a:
            inv_amount = st.number_input("投入金額（萬）",
                                          min_value=10, max_value=10000,
                                          value=100, step=10, key="pf_amt")
        with col_b:
            combo_choice = st.selectbox(
                "組合策略",
                ["⭐ TOP 200 tier（前向可用）",
                 "全市場 1028 檔（無篩選）",
                 "Top 50（事後 cherry-pick）",
                 "Top 100（事後 cherry-pick）",
                 "Tier OK 675 檔",
                 "Tier NA 153 檔（不適用）"],
                key="pf_combo")
        with col_c:
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:.7rem;color:#7a8899;line-height:1.6">'
                '對照：22 個月 0050 BH ≈ <b style="color:#3dbb6a">+90%</b>'
                '（年化 +49%）'
                '</div>', unsafe_allow_html=True)

        # 計算組合報酬
        deltas = []
        for t in _pnl_t:
            if t in _base_t:
                deltas.append((t, _pnl_t[t] - _base_t[t], _pnl_t[t]))
        deltas.sort(key=lambda x: -x[1])

        if combo_choice.startswith("⭐ TOP"):
            picks = [_pnl_t[t] for t, info in _tier_d.items()
                     if info.get('tier') == 'TOP' and t in _pnl_t]
            label = "TOP 200"
        elif "全市場" in combo_choice:
            picks = list(_pnl_t.values())
            label = "全市場"
        elif "Top 50" in combo_choice:
            picks = [d[2] for d in deltas[:50]]
            label = "Top 50（事後）"
        elif "Top 100" in combo_choice:
            picks = [d[2] for d in deltas[:100]]
            label = "Top 100（事後）"
        elif "OK" in combo_choice:
            picks = [_pnl_t[t] for t, info in _tier_d.items()
                     if info.get('tier') == 'OK' and t in _pnl_t]
            label = "OK"
        else:
            picks = [_pnl_t[t] for t, info in _tier_d.items()
                     if info.get('tier') == 'NA' and t in _pnl_t]
            label = "NA"

        if picks:
            arr = np.array(picks)
            avg_ret = arr.mean() / 100
            initial = inv_amount * 10000
            end_val = initial * (1 + avg_ret)
            years = 22 / 12
            ann = ((end_val/initial)**(1/years) - 1) * 100 if years > 0 else 0
            win_rate = (arr > 0).mean() * 100
            worst = arr.min()
            best = arr.max()
            std = arr.std()

            color_main = '#3dbb6a' if avg_ret > 0 else '#ff7777'

            # 結果卡片
            st.markdown(
                f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;'
                f'margin:12px 0">'
                # 終值
                f'<div style="background:#0a1e10;border:1px solid {color_main}55;'
                f'border-radius:8px;padding:14px;text-align:center">'
                f'<div style="color:#5a8ab0;font-size:.7rem">22 月終值</div>'
                f'<div style="color:{color_main};font-size:1.4rem;font-weight:700">'
                f'{end_val/10000:,.1f} 萬</div>'
                f'<div style="color:#7a8899;font-size:.7rem">初始 {inv_amount} 萬</div>'
                f'</div>'
                # 22 月報酬
                f'<div style="background:#0a1830;border:1px solid #5a8ab055;'
                f'border-radius:8px;padding:14px;text-align:center">'
                f'<div style="color:#5a8ab0;font-size:.7rem">22 月報酬</div>'
                f'<div style="color:{color_main};font-size:1.4rem;font-weight:700">'
                f'{avg_ret*100:+.1f}%</div>'
                f'<div style="color:#7a8899;font-size:.7rem">{label}</div>'
                f'</div>'
                # 年化
                f'<div style="background:#0a1830;border:1px solid #5a8ab055;'
                f'border-radius:8px;padding:14px;text-align:center">'
                f'<div style="color:#5a8ab0;font-size:.7rem">年化報酬</div>'
                f'<div style="color:{color_main};font-size:1.4rem;font-weight:700">'
                f'{ann:+.1f}%</div>'
                f'<div style="color:#7a8899;font-size:.7rem">CAGR</div>'
                f'</div>'
                # 勝率
                f'<div style="background:#0a1830;border:1px solid #5a8ab055;'
                f'border-radius:8px;padding:14px;text-align:center">'
                f'<div style="color:#5a8ab0;font-size:.7rem">個股勝率</div>'
                f'<div style="color:#3dbb6a;font-size:1.4rem;font-weight:700">'
                f'{win_rate:.1f}%</div>'
                f'<div style="color:#7a8899;font-size:.7rem">n={len(picks)}</div>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True)

            # 風險細節
            st.markdown(
                f'<div style="background:#080f1c;border:1px solid #1e3a5f;border-radius:8px;'
                f'padding:8px 14px;font-size:.78rem;color:#a8cce8">'
                f'📈 最佳個股 <b style="color:#3dbb6a">+{best:.0f}%</b> ｜ '
                f'📉 最差個股 <b style="color:#ff7777">{worst:.1f}%</b> ｜ '
                f'波動 σ <b>{std:.0f}%</b> ｜ '
                f'Sharpe~ <b>{(avg_ret*100/std if std else 0):.3f}</b>'
                f'</div>',
                unsafe_allow_html=True)

            # 警示框
            if 'cherry-pick' in combo_choice or 'Top 50' in combo_choice or 'Top 100' in combo_choice:
                st.markdown(
                    '<div style="background:#2a1a05;border:1px solid #f0a030;border-radius:8px;'
                    'padding:8px 14px;margin-top:8px;font-size:.78rem;color:#f0a030">'
                    '⚠️ <b>事後 cherry-pick 警告</b>：此結果是「歷史最佳 N 檔」回推，'
                    '實際操作不可能事先知道。建議參考但不可作為前向預期。'
                    '</div>',
                    unsafe_allow_html=True)
            elif 'NA' in combo_choice:
                st.markdown(
                    '<div style="background:#2a0808;border:1px solid #ff5555;border-radius:8px;'
                    'padding:8px 14px;margin-top:8px;font-size:.78rem;color:#ff7777">'
                    '⚠️ <b>NA tier 警告</b>：此 153 檔 VWAPEXEC 反而傷害，'
                    '不建議使用 P5+VWAPEXEC 策略。'
                    '</div>',
                    unsafe_allow_html=True)
            elif 'TOP 200' in combo_choice:
                st.markdown(
                    '<div style="background:#0a2a18;border:1px solid #3dbb6a;border-radius:8px;'
                    'padding:8px 14px;margin-top:8px;font-size:.78rem;color:#3dbb6a">'
                    '✅ <b>前向可用</b>：TOP 200 tier 是事先確定的 200 檔清單，'
                    '實際操作可實踐。預期年化 ~52%（含適度 alpha）。'
                    '</div>',
                    unsafe_allow_html=True)

st.markdown("#### 完整指標一覽表")

# 🆕 圖示說明區
st.markdown(
    '<div style="background:#080f1c;border:1px solid #1e3a5f;border-radius:8px;'
    'padding:8px 14px;margin:4px 0 10px;font-size:.74rem;color:#7abadd;'
    'display:flex;flex-wrap:wrap;gap:14px;align-items:center">'
    '<b style="color:#5a8ab0">圖示說明：</b>'
    '<span><span style="display:inline-block;padding:1px 5px;background:#0a3a1f;color:#3dbb6a;'
    'border-radius:3px;font-size:.65rem;font-weight:700">⭐</span>'
    '<span style="color:#a8cce8">　= VWAPEXEC TOP 200（歷史 Δ 顯著正向，最該用 VWAP 限價）</span></span>'
    '<span><span style="display:inline-block;padding:1px 5px;background:#3a0a0a;color:#ff8888;'
    'border-radius:3px;font-size:.65rem;font-weight:700">⚠️</span>'
    '<span style="color:#a8cce8">　= VWAPEXEC NA（Δ ≤ 0，VWAP 反而傷害，避開）</span></span>'
    '<span style="color:#7a8899">（無徽章 = OK 一般適用 / 不在歷史測試樣本）</span>'
    '</div>'
    '<div style="background:#080f1c;border:1px solid #1e3a5f;border-radius:8px;'
    'padding:8px 14px;margin-bottom:10px;font-size:.74rem;color:#7abadd;'
    'display:flex;flex-wrap:wrap;gap:14px;align-items:center">'
    '<b style="color:#5a8ab0">P/E 顏色：</b>'
    '<span><span style="color:#3dbb6a">綠</span> &lt; 20 合理偏低</span>'
    '<span><span style="color:#c8b87a">黃</span> 20-30 合理</span>'
    '<span><span style="color:#e8a020">橘</span> 30-50 偏高（成長股）</span>'
    '<span><span style="color:#ff5555">紅</span> &gt; 50 或虧損 過熱</span>'
    '<span><span style="color:#3dbb6a">▼▼</span> = 60 日 PER 大降（盈餘上修信號）</span>'
    '<span><span style="color:#ff5555">▲▲</span> = 60 日 PER 大漲（盈餘下修風險）</span>'
    '</div>',
    unsafe_allow_html=True)
st.markdown(render_table(results, platform_url_tpl, selected_platform), unsafe_allow_html=True)

st.markdown("<br>#### 個股指標詳細", unsafe_allow_html=True)
for item in results:
    ticker, market, d, error = item[0], item[1], item[2], item[3]
    groups      = item[4]   # (g_trend, g_position, g_momentum, g_aux)
    group_summs = item[5]   # (ts, ps, ms, xs)
    tsumm       = item[6]   # (tb, ts_, tn_, verdict)
    cap         = item[7]
    _, _, _, tr_ = tsumm
    name  = d.get("name", ticker) if d else ticker
    label = f"{ticker}  {name}" if name and name != ticker else ticker
    _rlabel_exp, _ = get_rec_label(d, ticker) if (d and not error) else ("無資料", "")
    title = f"{label}  {_rlabel_exp}" if not error else f"{label}  —  無資料"
    with st.expander(title, expanded=False):
        if error or not d:
            st.markdown('<div style="color:#334455;padding:12px">無法取得資料，請確認代號是否正確</div>',
                        unsafe_allow_html=True)
        else:
            # 🆕 v9.32：detail card 上方放 intraday-style plotly 互動 ZigZag chart（1d）
            try:
                from intraday.charts import build_zigzag_chart_plotly
                from intraday.settings import get_zigzag_atr_mult
                # 從 _swing_history 重建 df（fetch_indicators 已存 252 bars）
                _sh = d.get('_swing_history') or {}
                _closes = _sh.get('close') or []
                if _closes and len(_closes) >= 30:
                    _dates = _sh.get('dates') or []
                    if not _dates or len(_dates) != len(_closes):
                        _dates = pd.bdate_range(end=pd.Timestamp.now().normalize(),
                                                  periods=len(_closes))
                    _df_for_chart = pd.DataFrame({
                        'Open':   _sh.get('open') or _closes,
                        'High':   _sh.get('high') or _closes,
                        'Low':    _sh.get('low')  or _closes,
                        'Close':  _closes,
                        'Volume': _sh.get('volume') or [0] * len(_closes),
                    }, index=pd.to_datetime(_dates))
                    _atr_v = get_zigzag_atr_mult()
                    _fig = build_zigzag_chart_plotly(
                        _df_for_chart,
                        atr_mult=_atr_v,
                        title=f'{ticker} 1d — ZigZag (ATR×{_atr_v:.2f}) + BB + EMA  ｜ hover 看 OHLC',
                        max_bars=180,
                        show_bb=True,
                        show_emas=[5, 20, 60, 150, 200],
                        show_macd=True,
                        theme='dark',
                    )
                    if _fig is not None:
                        st.plotly_chart(_fig, use_container_width=True,
                                          key=f'_tvapp_zz_{ticker}')

                    # 🆕 v9.38：波段戰法訊號 banner
                    try:
                        from intraday.strategy import detect_swing_signal
                        _sig = detect_swing_signal(_df_for_chart, market=market, tf='1d')
                        if not _sig.get('error'):
                            _entry_v9 = _sig.get('entry', {})
                            _exit_v9 = _sig.get('exit', {})
                            _sepa_v9 = _sig.get('sepa', {})

                            _es = _entry_v9.get('state', 'NO_SETUP')
                            if _es == 'ENTER':
                                _ebg, _ec = '#0d2a14', '#3dbb6a'
                                _e_lbl = _entry_v9.get('label')
                            elif _es == 'WAIT_BB_P1':
                                _ebg, _ec = '#0a1828', '#5dccdd'
                                _e_lbl = _entry_v9.get('label')
                            else:
                                _ebg, _ec = '#0a1422', '#7a8899'
                                _e_lbl = '⚪ 無進場訊號 (EMA5≤EMA20 或 趨勢未全揚)'

                            _xs = _exit_v9.get('state', 'HOLD')
                            if _xs == 'EXIT':
                                _xbg, _xc = '#2a0a0a', '#ff5555'
                            elif _xs in ('WARN_PRICE', 'WARN_EMA'):
                                _xbg, _xc = '#1a1500', '#e8a020'
                            else:
                                _xbg, _xc = '#0a1422', '#7a8899'
                            _x_lbl = _exit_v9.get('label', '⚪ 持有')

                            _sepa_html_tv = ''
                            if _sepa_v9.get('available'):
                                _sc = _sepa_v9.get('score', 0)
                                if _sc == 7:    _sbg, _scol = '#0a2a14', '#3dbb6a'
                                elif _sc >= 5:  _sbg, _scol = '#0a2014', '#5dccdd'
                                elif _sc >= 3:  _sbg, _scol = '#1a1500', '#e8a020'
                                else:           _sbg, _scol = '#2a0a0a', '#ff5555'
                                _stip = ' ｜ '.join(
                                    f'{"✓" if _c2["pass"] else "✗"} {_c2["name"]}'
                                    for _c2 in _sepa_v9.get('conditions', [])
                                ).replace('"', '&quot;')
                                _sepa_html_tv = (
                                    f'<div style="background:{_sbg};color:{_scol};'
                                    f'padding:4px 8px;border-radius:4px;border-left:3px solid {_scol};'
                                    f'font-weight:700;font-size:.76rem;margin-bottom:4px"'
                                    f' title="{_stip}">{_sepa_v9["label"]}</div>'
                                )

                            # 🆕 v9.39：加碼訊號 banner
                            _re_v9 = _sig.get('reentry', {})
                            _re_html_tv = ''
                            if _re_v9:
                                _re_fired_tv = _re_v9.get('fired', [])
                                _re_cnt_tv = _re_v9.get('count', 0)
                                _re_abbrev_tv = {
                                    'r_p1sig_redo': 'P1', 'r_20d_high': 'HI',
                                    'r_mid_bounce': 'MB', 'r_ema5_pull': 'E5',
                                    'r_ema20': 'E20',
                                }
                                _re_full_tv = {
                                    'r_p1sig_redo': '重觸 BB+1σ',
                                    'r_20d_high': '破 20b 新高',
                                    'r_mid_bounce': 'BB 中軌反彈',
                                    'r_ema5_pull': 'EMA5 觸碰',
                                    'r_ema20': 'EMA20 觸碰',
                                }
                                if _re_cnt_tv >= 3:
                                    _re_col, _re_bg2 = '#3dbb6a', '#0d2a14'
                                elif _re_cnt_tv >= 2:
                                    _re_col, _re_bg2 = '#5dccdd', '#0a1828'
                                elif _re_cnt_tv == 1:
                                    _re_col, _re_bg2 = '#e8a020', '#1a1500'
                                else:
                                    _re_col, _re_bg2 = '#7a8899', '#0a1422'
                                if _re_cnt_tv > 0:
                                    _re_short = ' '.join(
                                        _re_abbrev_tv.get(k, k) for k in _re_fired_tv)
                                    _re_lbl_tv = f'💪 加碼 ×{_re_cnt_tv}: {_re_short}'
                                    _re_tip = ' ｜ '.join(
                                        f'{_re_abbrev_tv.get(k, k)}={_re_full_tv.get(k, k)}'
                                        for k in _re_fired_tv).replace('"', '&quot;')
                                else:
                                    _re_lbl_tv = '💤 無加碼訊號'
                                    _re_tip = '5 規則皆未觸發'
                                _re_html_tv = (
                                    f'<div style="background:{_re_bg2};color:{_re_col};'
                                    f'padding:4px 8px;border-radius:4px;border-left:3px solid {_re_col};'
                                    f'margin-top:3px;font-size:.76rem"'
                                    f' title="{_re_tip}">{_re_lbl_tv}</div>'
                                )

                            st.markdown(
                                f'<div style="background:#080f1c;border:1px solid #1e3a5f;'
                                f'border-radius:8px;padding:10px 14px;margin:8px 0">'
                                f'<div style="color:#5dccdd;font-weight:700;font-size:.8rem;'
                                f'margin-bottom:6px">🎯 波段戰法訊號 v9.39 '
                                f'<span style="color:#5a8ab0;font-weight:400;font-size:.7rem">'
                                f'(進場: EMA5&gt;EMA20 + 5EMA全上揚 + Close≥BB+1σ｜'
                                f'出場: Close&lt;BB Mid + EMA5/EMA20 下行)</span></div>'
                                f'{_sepa_html_tv}'
                                f'<div style="background:{_ebg};color:{_ec};padding:4px 8px;'
                                f'border-radius:4px;border-left:3px solid {_ec};margin-bottom:3px;'
                                f'font-size:.78rem">{_e_lbl}</div>'
                                f'<div style="background:{_xbg};color:{_xc};padding:4px 8px;'
                                f'border-radius:4px;border-left:3px solid {_xc};'
                                f'font-size:.78rem">{_x_lbl}</div>'
                                f'{_re_html_tv}'
                                f'</div>',
                                unsafe_allow_html=True)
                    except Exception:
                        pass
            except Exception:
                pass   # plotly 失敗就跳過，detail card 仍可看
            # detail card（v9.32 起 ZigZag PNG 由上面 plotly 取代、新聞已關閉）
            st.markdown(render_detail(ticker, d, groups, group_summs, tsumm, cap, market=market),
                        unsafe_allow_html=True)

            # ── 歷史區間 Excel ────────────────────────────────────────
            st.markdown(
                '<div style="border-top:1px solid #1a2f48;margin:14px 0 10px;padding-top:10px">'
                '<span style="font-size:.75rem;color:#5a8ab0;font-weight:700;letter-spacing:.05em">'
                '📊 歷史區間 Excel</span>'
                '<span style="font-size:.68rem;color:#334455;margin-left:8px">'
                '選擇起訖日後點「產生」，完成後出現下載按鈕</span></div>',
                unsafe_allow_html=True)
            _col_s, _col_e, _col_btn = st.columns([2, 2, 1])
            import datetime as _dt   # 確保已載入（sidebar 已 import，此行為保險）
            _today = _dt.date.today()
            with _col_s:
                _r_start = st.date_input(
                    "起始日",
                    value=_today - _dt.timedelta(days=90),
                    min_value=_dt.date(2010, 1, 1),
                    max_value=_today - _dt.timedelta(days=2),
                    key=f"rng_s_{ticker}",
                )
            with _col_e:
                _r_end = st.date_input(
                    "結束日",
                    value=_today - _dt.timedelta(days=1),
                    min_value=_dt.date(2010, 1, 2),
                    max_value=_today - _dt.timedelta(days=1),
                    key=f"rng_e_{ticker}",
                )
            with _col_btn:
                st.markdown("<div style='margin-top:26px'>", unsafe_allow_html=True)
                _gen_btn = st.button("🔄 產生", key=f"gen_rng_{ticker}",
                                     use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

            if _gen_btn:
                if _r_start >= _r_end:
                    st.error("起始日必須早於結束日")
                else:
                    with st.spinner(
                        f"正在計算 {ticker} {_r_start}～{_r_end} 歷史資料，請稍候…"
                    ):
                        _xl = build_stock_range_excel(
                            ticker, market,
                            _r_start.isoformat(), _r_end.isoformat()
                        )
                    st.session_state[f"rng_xl_{ticker}"]       = _xl
                    st.session_state[f"rng_xl_dates_{ticker}"] = (
                        _r_start.isoformat(), _r_end.isoformat())
                    st.success("✅ 產生完成，點擊下方按鈕下載")

            if f"rng_xl_{ticker}" in st.session_state:
                _sd, _ed = st.session_state.get(f"rng_xl_dates_{ticker}", ("", ""))
                st.download_button(
                    label=f"📥 下載 {ticker}  {_sd} ～ {_ed}",
                    data=st.session_state[f"rng_xl_{ticker}"],
                    file_name=f"{ticker}_{_sd}_{_ed}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_rng_{ticker}",
                    use_container_width=True,
                )


st.markdown("---")
# 🆕 v9.10p：底部下載按鈕 — PDF 為主、Excel 為備份
_, col_pdf, col_xls, _ = st.columns([1, 2, 2, 1])
with col_pdf:
    try:
        pdf_bytes = build_pdf(results)
        pdf_filename = f"Stock001_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        st.download_button(
            label="📕  下載 PDF 完整報告",
            data=pdf_bytes,
            file_name=pdf_filename,
            mime="application/pdf",
            use_container_width=True,
            type="primary",
            help="包含每股完整指標 + 操作建議 + 估值參考"
        )
    except Exception as _e:
        st.error(f"PDF 生成失敗：{str(_e)[:100]}")
with col_xls:
    excel_bytes = build_excel(results)
    xls_filename = f"Indicators_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        label="📊  Excel 表格（備份）",
        data=excel_bytes,
        file_name=xls_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
st.markdown(
    f'<div style="text-align:center;font-size:.7rem;color:#334455;margin-top:6px">'
    f'{total} 支股票 · {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} · '
    f'PDF 含每股 1 頁完整分析｜Excel 為原始指標表</div>',
    unsafe_allow_html=True)
